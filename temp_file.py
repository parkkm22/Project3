import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import os
import google.generativeai as genai
import io
import re
import pdfplumber
from pdf2image import convert_from_bytes
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import requests
import json
import xml.etree.ElementTree as ET
import urllib.parse
from io import BytesIO
import openpyxl
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

# ?섏씠吏 ?ㅼ젙
st.set_page_config(
    page_title="AI 怨듭궗愿由??먯씠?꾪듃",
    page_icon="??,
)

def parse_cell_address(cell_address):
    """
    ? 二쇱냼瑜??뚯떛?섏뿬 ?됯낵 ???몃뜳?ㅻ? 諛섑솚?⑸땲??
    
    Args:
        cell_address: ? 二쇱냼 (?? 'A1', 'BC123')
    
    Returns:
        tuple: (row_idx, col_idx)
    """
    import re
    
    # ?뺢퇋?앹쑝濡?? 二쇱냼 ?뚯떛
    match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
    if not match:
        raise ValueError(f"?섎せ??? 二쇱냼: {cell_address}")
    
    col_str, row_str = match.groups()
    
    # ???몃뜳??怨꾩궛 (A=1, B=2, ..., Z=26, AA=27, ...)
    col_idx = 0
    for char in col_str:
        col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
    
    row_idx = int(row_str)
    
    return row_idx, col_idx

def get_cell_value_safely(worksheet, cell_address):
    """
    蹂묓빀??????ы븿?섏뿬 ?덉쟾?섍쾶 ? 媛믪쓣 ?쎌뒿?덈떎.
    
    Args:
        worksheet: ?뚰겕?쒗듃 媛앹껜
        cell_address: ? 二쇱냼 (?? 'A1')
    
    Returns:
        ? 媛??먮뒗 None
    """
    try:
        # ? 二쇱냼瑜??뚯떛?섏뿬 吏곸젒 ?묎렐
        row_idx, col_idx = parse_cell_address(cell_address)
        
        # 吏곸젒 ? 媛믪뿉 ?묎렐
        cell = worksheet.cell(row=row_idx, column=col_idx)
        
        # 蹂묓빀????몄? ?뺤씤
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 蹂묓빀??踰붿쐞??泥?踰덉㎏ ??먯꽌 媛믪쓣 媛?몄샂
                top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        
        return cell.value
    except Exception as e:
        st.warning(f"? {cell_address} ?쎄린 ?ㅽ뙣: {e}")
        return None

def extract_cell_data_from_excel(excel_bytes, date_str):
    """
    ?묒? ?뚯씪?먯꽌 ?뱀젙 ? ?곗씠?곕? 異붿텧?⑸땲??
    
    Args:
        excel_bytes: ?묒? ?뚯씪 諛붿씠???곗씠??        date_str: ?좎쭨 臾몄옄??    
    Returns:
        dict: 異붿텧???곗씠??    """
    try:
        # ?묒? ?뚯씪 濡쒕뱶
        workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        
        extracted_data = {
            "date": date_str,
            "construction_data": {},
            "personnel_data": {},
            "equipment_data": {}
        }
        
        # 1. ?쒓났?꾪솴 ?곗씠??異붿텧 (A11~43, T11~43)
        for row in range(11, 44):
            category_cell = f"A{row}"
            cumulative_cell = f"T{row}"
            
            category = get_cell_value_safely(worksheet, category_cell)
            cumulative = get_cell_value_safely(worksheet, cumulative_cell)
            
            if category and str(category).strip():
                extracted_data["construction_data"][str(category)] = {
                    "?꾧퀎": cumulative if cumulative else 0
                }
        
        # 2. ?몄썝 ?곗씠??異붿텧 (A66~87, L66~87, N66~87, Y66~87)
        for row in range(66, 88):
            category_cell = f"A{row}"
            previous_cell = f"L{row}"
            today_cell = f"N{row}"
            cumulative_cell = f"Y{row}"
            
            category = get_cell_value_safely(worksheet, category_cell)
            previous = get_cell_value_safely(worksheet, previous_cell)
            today = get_cell_value_safely(worksheet, today_cell)
            cumulative = get_cell_value_safely(worksheet, cumulative_cell)
            
            if category and str(category).strip():
                extracted_data["personnel_data"][str(category)] = {
                    "?꾩씪源뚯?": previous if previous else 0,
                    "湲덉씪": today if today else 0,
                    "?꾧퀎": cumulative if cumulative else 0
                }
        
        # 3. ?λ퉬 ?곗씠??異붿텧 (A91~119, L91~119, N91~119, Y91~119)
        for row in range(91, 120):
            category_cell = f"A{row}"
            previous_cell = f"L{row}"
            today_cell = f"N{row}"
            cumulative_cell = f"Y{row}"
            
            category = get_cell_value_safely(worksheet, category_cell)
            previous = get_cell_value_safely(worksheet, previous_cell)
            today = get_cell_value_safely(worksheet, today_cell)
            cumulative = get_cell_value_safely(worksheet, cumulative_cell)
            
            if category and str(category).strip():
                extracted_data["equipment_data"][str(category)] = {
                    "?꾩씪源뚯?": previous if previous else 0,
                    "湲덉씪": today if today else 0,
                    "?꾧퀎": cumulative if cumulative else 0
                }
        
        return extracted_data
        
    except Exception as e:
        st.error(f"?묒? ?곗씠??異붿텧 以??ㅻ쪟: {e}")
        import traceback
        st.error(f"?곸꽭 ?ㅻ쪟: {traceback.format_exc()}")
        return None

def get_previous_day_data(current_date):
    """
    ?꾩씪 ?곗씠?곕? 媛?몄샃?덈떎.
    
    Args:
        current_date: ?꾩옱 ?좎쭨 (YYYY-MM-DD)
    
    Returns:
        dict: ?꾩씪 ?곗씠???먮뒗 None
    """
    try:
        # ?꾩씪 ?좎쭨 怨꾩궛
        current_dt = datetime.strptime(current_date, "%Y-%m-%d")
        previous_dt = current_dt - timedelta(days=1)
        previous_date = previous_dt.strftime("%Y-%m-%d")
        
        # Supabase?먯꽌 ?꾩씪 ?곗씠??議고쉶
        result = supabase_client.table("daily_report").select("*").eq("date", previous_date).execute()
        
        if result.data:
            st.info(f"?뵇 ?꾩씪 ?곗씠??諛쒓껄: {previous_date}")
            return result.data[0]
        else:
            st.info(f"?뱄툘 ?꾩씪 ?곗씠???놁쓬: {previous_date}")
        return None
        
    except Exception as e:
        st.error(f"?꾩씪 ?곗씠??議고쉶 以??ㅻ쪟: {e}")
        return None

def apply_previous_day_data_to_excel(excel_bytes, previous_data):
    """
    ?꾩씪 ?곗씠?곕? ?묒? ?뚯씪???곸슜?⑸땲??
    
    Args:
        excel_bytes: ?묒? ?뚯씪 諛붿씠???곗씠??        previous_data: ?꾩씪 ?곗씠??    
    Returns:
        bytes: ?섏젙???묒? ?뚯씪 諛붿씠???곗씠??    """
    try:
        # ?묒? ?뚯씪 濡쒕뱶
        workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        
        if not previous_data:
            return excel_bytes
        
        # 1. ?쒓났?꾪솴 ?꾩씪 ?곗씠???곸슜 (T11~43 ?꾧퀎 ??N11~43 ?꾩씪源뚯?)
        construction_data = previous_data.get("construction_data", {})
        row = 11
        for category, data in construction_data.items():
            if row <= 43:
                cumulative_value = data.get("?꾧퀎", 0)
                worksheet[f"N{row}"] = cumulative_value
                row += 1
        
        # 2. ?몄썝 ?꾩씪 ?곗씠???곸슜 (L66~87, Y66~87)
        personnel_data = previous_data.get("personnel_data", {})
        row = 66
        for category, data in personnel_data.items():
            if row <= 87:
                previous_value = data.get("?꾩씪源뚯?", 0)
                cumulative_value = data.get("?꾧퀎", 0)
                worksheet[f"L{row}"] = previous_value
                worksheet[f"Y{row}"] = cumulative_value
                row += 1
        
        # 3. ?λ퉬 ?꾩씪 ?곗씠???곸슜 (L91~119, Y91~119)
        equipment_data = previous_data.get("equipment_data", {})
        row = 91
        for category, data in equipment_data.items():
            if row <= 119:
                previous_value = data.get("?꾩씪源뚯?", 0)
                cumulative_value = data.get("?꾧퀎", 0)
                worksheet[f"L{row}"] = previous_value
                worksheet[f"Y{row}"] = cumulative_value
                row += 1
        
        # ?섏젙???묒? ?뚯씪??諛붿씠?몃줈 蹂??        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"?꾩씪 ?곗씠???곸슜 以??ㅻ쪟: {e}")
        return excel_bytes

# --- CONFIG & SETUP ---
st.set_page_config(
    page_title="?묒뾽?쇰낫 ?먮룞??,
    layout="wide"
)

# Supabase ?대씪?댁뼵?몃뒗 ?⑥닔 ?뺤쓽 ?꾩뿉 珥덇린?붾맗?덈떎

# 湲곗긽泥?API ?ㅼ젙
# 湲곗긽泥?API ?ㅼ젙
WEATHER_API_KEY = (
    "srgpo0t7uDjbNhm4WllX4RVzvVowMmqeSsJ7Y0Sg2XmHWjTUu%2BXou%2FuSFiLcKEvKpAo"
    "JhlKsNRVlcXcNh%2Fjm1Q%3D%3D"
)

# ?쒓컙蹂?ASOS ?뺣낫瑜?議고쉶?섎뒗 湲곗긽泥?OpenAPI
WEATHER_API_URL = "https://apis.data.go.kr/1360000/AsosHourlyInfoService/getWthrDataList"

# --- STYLING ---
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    

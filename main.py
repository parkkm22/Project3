import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import os
import google.generativeai as genai
import io
import re
import pdfplumber
from pdf2image import convert_from_bytes
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import requests
import json
import xml.etree.ElementTree as ET
import urllib.parse
from io import BytesIO
import openpyxl
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

def parse_cell_address(cell_address):
    """
    셀 주소를 파싱하여 행과 열 인덱스를 반환합니다.
    
    Args:
        cell_address: 셀 주소 (예: 'A1', 'BC123')
    
    Returns:
        tuple: (row_idx, col_idx)
    """
    import re
    
    # 정규식으로 셀 주소 파싱
    match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
    if not match:
        raise ValueError(f"잘못된 셀 주소: {cell_address}")
    
    col_str, row_str = match.groups()
    
    # 열 인덱스 계산 (A=1, B=2, ..., Z=26, AA=27, ...)
    col_idx = 0
    for char in col_str:
        col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
    
    row_idx = int(row_str)
    
    return row_idx, col_idx

def get_cell_value_safely(worksheet, cell_address):
    """
    병합된 셀을 포함하여 안전하게 셀 값을 읽습니다.
    
    Args:
        worksheet: 워크시트 객체
        cell_address: 셀 주소 (예: 'A1')
    
    Returns:
        셀 값 또는 None
    """
    try:
        # 셀 주소를 파싱하여 직접 접근
        row_idx, col_idx = parse_cell_address(cell_address)
        
        # 직접 셀 값에 접근
        cell = worksheet.cell(row=row_idx, column=col_idx)
        
        # 병합된 셀인지 확인
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 병합된 범위의 첫 번째 셀에서 값을 가져옴
                top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        
        return cell.value
    except Exception as e:
        st.warning(f"셀 {cell_address} 읽기 실패: {e}")
        return None

def extract_cell_data_from_excel(excel_bytes, date_str):
    """
    엑셀 파일에서 특정 셀 데이터를 추출합니다.
    
    Args:
        excel_bytes: 엑셀 파일 바이트 데이터
        date_str: 날짜 문자열
    
    Returns:
        dict: 추출된 데이터
    """
    try:
        # 엑셀 파일 로드
        workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        
        extracted_data = {
            "date": date_str,
            "construction_data": {},
            "personnel_data": {},
            "equipment_data": {}
        }
        
        # 1. 시공현황 데이터 추출 (A11~43, T11~43)
        for row in range(11, 44):
            category_cell = f"A{row}"
            cumulative_cell = f"T{row}"
            
            category = get_cell_value_safely(worksheet, category_cell)
            cumulative = get_cell_value_safely(worksheet, cumulative_cell)
            
            if category and str(category).strip():
                extracted_data["construction_data"][str(category)] = {
                    "누계": cumulative if cumulative else 0
                }
        
        # 2. 인원 데이터 추출 (A66~87, L66~87, N66~87, Y66~87)
        for row in range(66, 88):
            category_cell = f"A{row}"
            previous_cell = f"L{row}"
            today_cell = f"N{row}"
            cumulative_cell = f"Y{row}"
            
            category = get_cell_value_safely(worksheet, category_cell)
            previous = get_cell_value_safely(worksheet, previous_cell)
            today = get_cell_value_safely(worksheet, today_cell)
            cumulative = get_cell_value_safely(worksheet, cumulative_cell)
            
            if category and str(category).strip():
                extracted_data["personnel_data"][str(category)] = {
                    "전일까지": previous if previous else 0,
                    "금일": today if today else 0,
                    "누계": cumulative if cumulative else 0
                }
        
        # 3. 장비 데이터 추출 (A91~119, L91~119, N91~119, Y91~119)
        for row in range(91, 120):
            category_cell = f"A{row}"
            previous_cell = f"L{row}"
            today_cell = f"N{row}"
            cumulative_cell = f"Y{row}"
            
            category = get_cell_value_safely(worksheet, category_cell)
            previous = get_cell_value_safely(worksheet, previous_cell)
            today = get_cell_value_safely(worksheet, today_cell)
            cumulative = get_cell_value_safely(worksheet, cumulative_cell)
            
            if category and str(category).strip():
                extracted_data["equipment_data"][str(category)] = {
                    "전일까지": previous if previous else 0,
                    "금일": today if today else 0,
                    "누계": cumulative if cumulative else 0
                }
        
        return extracted_data
        
    except Exception as e:
        st.error(f"엑셀 데이터 추출 중 오류: {e}")
        import traceback
        st.error(f"상세 오류: {traceback.format_exc()}")
        return None

def get_previous_day_data(current_date):
    """
    전일 데이터를 가져옵니다.
    
    Args:
        current_date: 현재 날짜 (YYYY-MM-DD)
    
    Returns:
        dict: 전일 데이터 또는 None
    """
    try:
        # 전일 날짜 계산
        current_dt = datetime.strptime(current_date, "%Y-%m-%d")
        previous_dt = current_dt - timedelta(days=1)
        previous_date = previous_dt.strftime("%Y-%m-%d")
        
        # Supabase에서 전일 데이터 조회
        result = supabase_client.table("daily_report_data").select("*").eq("date", previous_date).execute()
        
        if result.data:
            return result.data[0]
        return None
        
    except Exception as e:
        st.error(f"전일 데이터 조회 중 오류: {e}")
        return None

def apply_previous_day_data_to_excel(excel_bytes, previous_data):
    """
    전일 데이터를 엑셀 파일에 적용합니다.
    
    Args:
        excel_bytes: 엑셀 파일 바이트 데이터
        previous_data: 전일 데이터
    
    Returns:
        bytes: 수정된 엑셀 파일 바이트 데이터
    """
    try:
        # 엑셀 파일 로드
        workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
        worksheet = workbook.active
        
        if not previous_data:
            return excel_bytes
        
        # 1. 시공현황 전일 데이터 적용 (T11~43 누계 → N11~43 전일까지)
        construction_data = previous_data.get("construction_data", {})
        row = 11
        for category, data in construction_data.items():
            if row <= 43:
                cumulative_value = data.get("누계", 0)
                worksheet[f"N{row}"] = cumulative_value
                row += 1
        
        # 2. 인원 전일 데이터 적용 (L66~87, Y66~87)
        personnel_data = previous_data.get("personnel_data", {})
        row = 66
        for category, data in personnel_data.items():
            if row <= 87:
                previous_value = data.get("전일까지", 0)
                cumulative_value = data.get("누계", 0)
                worksheet[f"L{row}"] = previous_value
                worksheet[f"Y{row}"] = cumulative_value
                row += 1
        
        # 3. 장비 전일 데이터 적용 (L91~119, Y91~119)
        equipment_data = previous_data.get("equipment_data", {})
        row = 91
        for category, data in equipment_data.items():
            if row <= 119:
                previous_value = data.get("전일까지", 0)
                cumulative_value = data.get("누계", 0)
                worksheet[f"L{row}"] = previous_value
                worksheet[f"Y{row}"] = cumulative_value
                row += 1
        
        # 수정된 엑셀 파일을 바이트로 변환
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"전일 데이터 적용 중 오류: {e}")
        return excel_bytes

# --- CONFIG & SETUP ---
st.set_page_config(
    page_title="공사일보 자동화",
    page_icon="🏗️",
    layout="wide"
)

# Supabase 클라이언트는 함수 정의 후에 초기화됩니다

# 기상청 API 설정
# 기상청 API 설정
WEATHER_API_KEY = (
    "srgpo0t7uDjbNhm4WllX4RVzvVowMmqeSsJ7Y0Sg2XmHWjTUu%2BXou%2FuSFiLcKEvKpAo"
    "JhlKsNRVlcXcNh%2Fjm1Q%3D%3D"
)

# 시간별 ASOS 정보를 조회하는 기상청 OpenAPI
WEATHER_API_URL = "https://apis.data.go.kr/1360000/AsosHourlyInfoService/getWthrDataList"

# --- STYLING ---
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Global Styles */
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    }
    
    /* Main App Background - Summer Gradient */
    .main {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    
    /* Main container with glassmorphism */
    .main .block-container {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border: 1px solid rgba(255, 255, 255, 0.2);
        border-radius: 24px;
        padding: 2rem 1.5rem 5rem 1.5rem;
        max-width: 1200px;
        margin: 1rem auto;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
    }
    
    /* Sidebar glassmorphism */
    [data-testid="stSidebar"] {
        background: rgba(255, 255, 255, 0.15);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border-right: 1px solid rgba(255, 255, 255, 0.2);
    }
    
    [data-testid="stSidebar"] h1 {
        font-size: 1.5rem;
        color: #ffffff;
        font-weight: 600;
        padding: 1rem 0;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }
    
    /* Glassmorphism Cards */
    .glass-card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border: 1px solid rgba(255, 255, 255, 0.2);
        border-radius: 16px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        transition: all 0.3s ease;
    }
    
    .glass-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 12px 40px rgba(0, 0, 0, 0.15);
    }
    
    /* Section Headers */
    h1, h2, h3 {
        color: #1e293b !important;
        font-weight: 600;
        text-shadow: none;
        letter-spacing: -0.02em;
    }
    
    h1 {
        font-size: 2.5rem !important;
        color: #0f172a !important;
        font-weight: 700 !important;
    }
    
    h2 {
        font-size: 1.75rem !important;
        color: #1e293b !important;
        font-weight: 600 !important;
    }
    
    h3 {
        font-size: 1.25rem !important;
        color: #334155 !important;
        font-weight: 500 !important;
    }
    
    /* Text Areas with glassmorphism */
    .stTextArea textarea {
        background: rgba(255, 255, 255, 0.9) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        color: #1e293b !important;
        font-size: 14px;
        padding: 12px !important;
        font-weight: 400;
    }
    
    .stTextArea textarea::placeholder {
        color: rgba(30, 41, 59, 0.6) !important;
    }
    
    /* General text styling */
    p, div, span {
        color: #334155 !important;
        line-height: 1.6;
    }
    
    /* Strong text */
    strong, b {
        color: #1e293b !important;
        font-weight: 600;
    }
    
    /* Buttons with glassmorphism */
    .stButton > button {
        background: rgba(255, 255, 255, 0.9) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        color: #1e293b !important;
        font-weight: 500 !important;
        padding: 10px 20px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 16px rgba(0, 0, 0, 0.1) !important;
        font-size: 14px;
    }
    
    .stButton > button:hover {
        background: rgba(255, 255, 255, 1) !important;
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15) !important;
        color: #0f172a !important;
    }
    
    /* Primary Button */
    .stButton > button,
    button[data-baseweb="button"] {
        background: linear-gradient(135deg, #788CE6 0%, #7850B4 100%) !important;
        border: none !important;
        color: #ffffff !important;
        font-weight: 600 !important;
        box-shadow: 0 4px 20px rgba(120, 140, 230, 0.4) !important;
    }
    
    .stButton > button:hover,
    button[data-baseweb="button"]:hover {
        background: linear-gradient(135deg, #6B7FD8 0%, #6B46C1 100%) !important;
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(120, 140, 230, 0.5) !important;
        color: #ffffff !important;
    }
    
    /* 강제로 모든 버튼 텍스트를 흰색으로 설정 */
    .stButton > button *,
    button[data-baseweb="button"] * {
        color: #ffffff !important;
    }
    
    /* 버튼 내부의 모든 텍스트 요소 */
    .stButton > button span,
    button[data-baseweb="button"] span,
    .stButton > button div,
    button[data-baseweb="button"] div {
        color: #ffffff !important;
    }
    
    /* 모든 버튼을 AI 자동채우기와 같은 스타일로 통일 */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        border: none !important;
        color: white !important;
        font-weight: 600 !important;
        border-radius: 12px !important;
        padding: 8px 16px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4) !important;
    }
    
    .stButton > button:hover {
        background: linear-gradient(135deg, #5a6fd8 0%, #6a4190 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6) !important;
    }
    
    /* 프롬프트 관리 섹션의 버튼만 기본 스타일 적용 */
    .stExpander .stButton > button {
        background: #f0f2f6 !important;
        border: 1px solid #d0d7de !important;
        color: #000000 !important;
        font-weight: 500 !important;
        border-radius: 6px !important;
        padding: 6px 12px !important;
        transition: none !important;
        box-shadow: none !important;
    }
    
    .stExpander .stButton > button:hover {
        background: #f6f8fa !important;
        border-color: #0969da !important;
        color: #0969da !important;
        transform: none !important;
        box-shadow: none !important;
    }
    
    /* 프롬프트 관리 섹션의 모든 버튼 텍스트를 강제로 검은색으로 설정 */
    .stExpander button {
        color: #000000 !important;
    }
    
    .stExpander button:hover {
        color: #0969da !important;
    }
    
    /* 모든 버튼 텍스트를 강제로 검은색으로 설정 */
    button {
        color: #000000 !important;
    }
    
    button:hover {
        color: #0969da !important;
    }
    
    /* 프롬프트 관리 섹션의 모든 요소를 검은색으로 설정 */
    .stExpander * {
        color: #000000 !important;
    }
    
    /* 모든 버튼 내부 텍스트를 강제로 검은색으로 설정 */
    .stButton > button {
        color: #000000 !important;
    }
    
    .stButton > button:hover {
        color: #0969da !important;
    }
    
    /* 모든 버튼 요소의 텍스트를 강제로 검은색으로 설정 */
    button span, button div, button p {
        color: #000000 !important;
    }
    
    /* Streamlit 버튼의 모든 텍스트를 강제로 검은색으로 설정 */
    [data-testid="stButton"] button {
        color: #000000 !important;
    }
    
    /* 모든 텍스트를 강제로 검은색으로 설정 */
    * {
        color: #000000 !important;
    }
    
    /* 버튼 내부의 모든 텍스트를 강제로 검은색으로 설정 */
    button * {
        color: #000000 !important;
    }
    
    /* 프롬프트 관리 섹션의 모든 버튼 텍스트를 강제로 검은색으로 설정 */
    .stExpander button * {
        color: #000000 !important;
    }
    
    /* 모든 Streamlit 버튼의 텍스트를 강제로 검은색으로 설정 */
    .stButton * {
        color: #000000 !important;
    }
    
    /* 모든 버튼의 텍스트를 강제로 검은색으로 설정 */
    button, button *, .stButton, .stButton * {
        color: #000000 !important;
    }
    
    /* 프롬프트 관리 섹션의 모든 요소를 검은색으로 설정 */
    .stExpander, .stExpander * {
        color: #000000 !important;
    }
    
    /* 모든 Streamlit 컴포넌트의 텍스트를 검은색으로 설정 */
    [data-testid], [data-testid] * {
        color: #000000 !important;
    }
    
    /* 프롬프트 관리 섹션의 텍스트 색상을 검은색으로 설정 */
    .stExpander {
        color: #000000 !important;
    }
    
    .stExpander .stMarkdown {
        color: #000000 !important;
    }
    
    .stExpander .stTextInput > div > div > input {
        color: #000000 !important;
    }
    
    .stExpander .stTextArea > div > div > textarea {
        color: #000000 !important;
    }
    
    .stExpander .stSelectbox > div > div {
        color: #000000 !important;
    }
    
    /* File Uploader */
    .stFileUploader {
        background: rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        border-radius: 12px !important;
        padding: 1rem !important;
    }
    
    /* Selectbox */
    .stSelectbox > div > div {
        background: rgba(255, 255, 255, 0.9) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        color: #1e293b !important;
        font-weight: 500;
    }
    
    /* Text Input */
    .stTextInput > div > div > input {
        background: rgba(255, 255, 255, 0.9) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        color: #1e293b !important;
        font-weight: 400;
        padding: 8px 12px !important;
    }
    
    .stTextInput > div > div > input::placeholder {
        color: rgba(30, 41, 59, 0.6) !important;
    }
    
    /* Date Input */
    .stDateInput > div > div > input {
        background: rgba(255, 255, 255, 0.9) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        color: #1e293b !important;
        font-weight: 400;
        padding: 8px 12px !important;
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        background: rgba(255, 255, 255, 0.9) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        padding: 6px !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        color: rgba(30, 41, 59, 0.7) !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
        font-weight: 500;
        padding: 8px 16px !important;
    }
    
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: rgba(59, 130, 246, 0.1) !important;
        color: #1e40af !important;
        font-weight: 600;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(59, 130, 246, 0.05) !important;
        color: #1e293b !important;
    }
    
    /* Data Editor */
    .stDataFrame {
        background: rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        border-radius: 12px !important;
    }
    
    /* Metrics */
    .metric-container {
        background: rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        border-radius: 12px !important;
        padding: 1rem !important;
        margin: 0.5rem 0 !important;
    }
    
    /* Success/Error Messages */
    .stAlert {
        background: rgba(255, 255, 255, 0.95) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1) !important;
        border-radius: 12px !important;
        color: #1e293b !important;
        font-weight: 500;
        padding: 1rem !important;
    }
    
    /* Success messages */
    .stAlert[data-baseweb="notification"] {
        background: rgba(34, 197, 94, 0.1) !important;
        border-color: rgba(34, 197, 94, 0.3) !important;
        color: #166534 !important;
    }
    
    /* Error messages */
    .stAlert[data-baseweb="notification"][data-severity="error"] {
        background: rgba(239, 68, 68, 0.1) !important;
        border-color: rgba(239, 68, 68, 0.3) !important;
        color: #dc2626 !important;
    }
    
    /* Warning messages */
    .stAlert[data-baseweb="notification"][data-severity="warning"] {
        background: rgba(245, 158, 11, 0.1) !important;
        border-color: rgba(245, 158, 11, 0.3) !important;
        color: #d97706 !important;
    }
    
    /* Info messages */
    .stAlert[data-baseweb="notification"][data-severity="info"] {
        background: rgba(59, 130, 246, 0.1) !important;
        border-color: rgba(59, 130, 246, 0.3) !important;
        color: #1e40af !important;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        border-radius: 12px !important;
        color: #ffffff !important;
    }
    
    /* Mobile Optimization */
    @media (max-width: 768px) {
        .main .block-container {
            padding: 1rem 1rem 3rem 1rem;
            margin: 0.5rem;
            border-radius: 16px;
        }
        
        .glass-card {
            padding: 1rem;
            border-radius: 12px;
        }
        
        h1 { font-size: 1.5rem !important; }
        h2 { font-size: 1.25rem !important; }
        h3 { font-size: 1.1rem !important; }
        
        .stButton > button {
            padding: 10px 16px !important;
            font-size: 14px !important;
        }
        
        .stTextArea textarea {
            font-size: 16px !important; /* Prevents zoom on iOS */
        }
        
        /* Stack columns on mobile */
        [data-testid="column"] {
            width: 100% !important;
            margin-bottom: 1rem;
        }
    }
    
    /* Custom Scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: rgba(255, 255, 255, 0.1);
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: rgba(255, 255, 255, 0.3);
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: rgba(255, 255, 255, 0.5);
    }
    
    /* Loading Spinner */
    .stSpinner > div {
        border-color: rgba(255, 255, 255, 0.3) !important;
        border-top-color: #ffffff !important;
    }
    
    /* Code blocks */
    .stCodeBlock {
        background: rgba(0, 0, 0, 0.2) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 8px !important;
    }
    
    /* AgGrid Customization */
    .ag-theme-streamlit {
        background: rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        border-radius: 12px !important;
    }
    
    .ag-header-cell {
        background: rgba(255, 255, 255, 0.2) !important;
        color: #ffffff !important;
        font-weight: 600 !important;
    }
    
    .ag-cell {
        background: rgba(255, 255, 255, 0.05) !important;
        color: #ffffff !important;
        border-color: rgba(255, 255, 255, 0.1) !important;
    }
    
    /* Step styling for sidebar */
    .step {
        display: flex;
        align-items: center;
        margin-bottom: 1rem;
        padding: 1rem;
        border-radius: 12px;
        transition: all 0.3s ease;
        background: rgba(255, 255, 255, 0.9);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border: 1px solid rgba(0, 0, 0, 0.1);
        color: #1e293b;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
    }
    
    .step.active {
        background: rgba(59, 130, 246, 0.1);
        border-color: rgba(59, 130, 246, 0.3);
        transform: translateX(4px);
        box-shadow: 0 4px 16px rgba(59, 130, 246, 0.2);
    }
    
    .step.completed {
        background: rgba(34, 197, 94, 0.1);
        border-color: rgba(34, 197, 94, 0.3);
        box-shadow: 0 2px 8px rgba(34, 197, 94, 0.2);
    }
    
    .step-icon {
        font-size: 1.25rem;
        margin-right: 0.75rem;
        min-width: 1.5rem;
        text-align: center;
        color: #1e293b;
    }
    
    .step-text {
        font-size: 0.9rem;
        font-weight: 500;
        color: #1e293b;
        flex: 1;
        line-height: 1.4;
    }
    
    .step.completed .step-text {
        color: #166534;
        font-weight: 600;
    }
    
    .step.active .step-text {
        color: #1e40af;
        font-weight: 600;
    }
    
    /* Card styling */
    .card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border: 1px solid rgba(255, 255, 255, 0.2);
        border-radius: 16px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        transition: all 0.3s ease;
    }
    
    .card:hover {
        transform: translateY(-2px);
        box-shadow: 0 12px 40px rgba(0, 0, 0, 0.15);
    }
    
    .card-title {
        font-size: 1.25rem;
        font-weight: 600;
        color: #ffffff;
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
    }
    
    .card-title .icon {
        margin-right: 0.5rem;
        font-size: 1.5rem;
    }
    
    .card-description {
        color: rgba(255, 255, 255, 0.8);
        font-size: 0.9rem;
        margin: 0;
    }
</style>
""", unsafe_allow_html=True)


# --- GLOBAL CONSTANTS & API SETUP ---
# Streamlit secrets에서 설정 가져오기
TEAMS_WEBHOOK_URL = st.secrets.get("TEAMS_WEBHOOK_URL", "https://poscoenc365.webhook.office.com/webhookb2/f6efcf11-c6a7-4385-903f-f3fd8937de55@ec1d3aa9-13ec-4dc5-8672-06fc64ca7701/IncomingWebhook/1fb9d9ce7f4c4093ba4fe9a8db67dc2f/1a2e3f7d-551b-40ec-90a1-e815373c81a7/V2qbqRtbAap4il8cvVljyk_ApZuHTDE0AfOYLQ8V9SqQs1")
GENAI_API_KEY = st.secrets.get("GENAI_API_KEY", "AIzaSyD69-wKYfZSID327fczrkx-JveJdGYIUIk")

def get_weather_data(date=None, station_id="108"):
    """
    기상청 API를 통해 날씨 데이터를 가져옵니다. (시간별 ASOS 기준)
    
    Args:
        date: YYYYMMDD 형식의 날짜 문자열 (기본값: 오늘)
        station_id: 관측소 ID (기본값: 서울=108)
    
    Returns:
        dict: {'최고기온': str, '최저기온': str, '강수량': str}
    """
    # 날짜가 없으면 어제 날짜 사용 (기상청 API는 전날 자료까지 제공)
    if date is None:
        date = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    
    try:
        # (1) API 키 디코딩
        decoded_api_key = urllib.parse.unquote(WEATHER_API_KEY)
        
        # (2) API 키 검증
        if not decoded_api_key or len(decoded_api_key) < 10:
            print("❌ API 키가 올바르지 않습니다.")
            return {
                '최고기온': '25.5',
                '최저기온': '18.2',
                '강수량': '0.0'
            }
        
        # (3) 요청 파라미터 구성
        # 하루(00~23시) 기준 24시간 데이터를 불러오도록 설정
        params = {
            "serviceKey": decoded_api_key,   # 디코딩된 인증키
            "dataType": "XML",              # 응답 타입
            "pageNo": "1",
            "numOfRows": "24",
            "dataCd": "ASOS",
            "dateCd": "HR",                 # 시간 자료
            "startDt": date,
            "startHh": "00",
            "endDt": date,
            "endHh": "23",
            "stnIds": station_id
        }
        
        print(f"API 요청 URL: {WEATHER_API_URL}")
        print(f"요청 파라미터: {params}")
        
        # (4) 실제 요청 (SSL 우회 포함)
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        try:
            # SSL 검증 없이 요청
            response = requests.get(WEATHER_API_URL, params=params, timeout=30, verify=False)
            response.raise_for_status()  # 에러 발생 시 예외 처리
        except requests.exceptions.SSLError:
            # SSL 오류 시 HTTP로 시도
            http_url = WEATHER_API_URL.replace('https://', 'http://')
            response = requests.get(http_url, params=params, timeout=30, verify=False)
            response.raise_for_status()
        
        print(f"응답 상태 코드: {response.status_code}")
        print(f"응답 내용 (처음 500자): {response.text[:500]}")
        
        # (5) XML 파싱
        root = ET.fromstring(response.content)
        
        # (5-1) 결과 코드 확인
        result_code = root.find('.//resultCode')
        if result_code is not None and result_code.text != '00':
            result_msg = root.find('.//resultMsg')
            err_msg = result_msg.text if result_msg is not None else "알 수 없는 오류"
            
            # "전날 자료까지 제공됩니다" 오류인 경우 어제 날짜로 재시도
            if "전날 자료까지" in err_msg:
                print(f"전날 자료 오류 발생. 어제 날짜로 재시도합니다.")
                yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
                if yesterday != date:
                    return get_weather_data(yesterday, station_id)
            
            raise Exception(f"API 오류: {err_msg}")
        
        # (5-2) 관측 데이터 추출
        items = root.findall('.//item')
        print(f"추출된 아이템 수: {len(items)}")
        
        if not items:
            # 데이터가 없으면 어제 날짜로 재시도
            print("데이터가 없습니다. 어제 날짜로 다시 시도합니다.")
            yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
            if yesterday != date:
                return get_weather_data(yesterday, station_id)
            else:
                # 어제도 데이터가 없다면 기본값 반환
                return {
                    'max_temp': 25.5,
                    'min_temp': 18.2,
                    'precipitation': 0.0
                }
        
        # (6) 온도 및 강수량 계산
        temperatures = []
        precipitation = 0.0
        
        for item in items:
            # 기온(ta)
            temp_elem = item.find('ta')
            if temp_elem is not None and temp_elem.text:
                try:
                    temperatures.append(float(temp_elem.text))
                except ValueError:
                    pass
            
            # 강수량(rn)
            rain_elem = item.find('rn')
            if rain_elem is not None and rain_elem.text:
                try:
                    precipitation += float(rain_elem.text)
                except ValueError:
                    pass
        
        if temperatures:
            max_temp = max(temperatures)
            min_temp = min(temperatures)
        else:
            max_temp, min_temp = None, None
        
        # (7) 최종 결과 생성
        result = {
            'max_temp': float(f"{max_temp:.1f}") if max_temp is not None else 25.5,
            'min_temp': float(f"{min_temp:.1f}") if min_temp is not None else 18.2,
            'precipitation': float(f"{precipitation:.1f}") if precipitation > 0 else 0.0
        }
        
        print(f"최종 결과: {result}")
        return result
        
    except Exception as e:
        # API 호출 중 예외가 발생하면 로그와 함께 에러 표시
        print(f"API 호출 중 오류 발생: {e}")
        st.error(f"❌ 날씨 데이터 가져오기 실패: {e}")
        
        # 실패 시 현실적인 테스트 데이터 반환
        print("API 연결 실패로 테스트 데이터를 반환합니다.")
        
        # 현재 계절에 맞는 현실적인 데이터
        current_month = datetime.now().month
        if current_month in [12, 1, 2]:  # 겨울
            test_data = {'max_temp': 5.2, 'min_temp': -2.1, 'precipitation': 0.0}
        elif current_month in [3, 4, 5]:  # 봄
            test_data = {'max_temp': 18.5, 'min_temp': 8.3, 'precipitation': 2.5}
        elif current_month in [6, 7, 8]:  # 여름
            test_data = {'max_temp': 28.7, 'min_temp': 22.1, 'precipitation': 15.3}
        else:  # 가을
            test_data = {'max_temp': 20.3, 'min_temp': 12.8, 'precipitation': 0.0}
        
        print(f"계절별 테스트 데이터: {test_data}")
        return test_data

def get_weather_stations():
    """
    주요 관측소 목록을 반환합니다.
    """
    return {
        "서울": "108",
        "부산": "159", 
        "대구": "143",
        "인천": "112",
        "광주": "156",
        "대전": "133",
        "울산": "152",
        "세종": "184",
        "수원": "119",
        "춘천": "101",
        "강릉": "105",
        "청주": "131",
        "전주": "146",
        "포항": "138",
        "제주": "184"
    }

# Supabase 클라이언트 초기화
@st.cache_resource
def init_supabase():
    """Supabase 클라이언트를 초기화하고 반환합니다."""
    if not SUPABASE_AVAILABLE:
        return None
        
    try:
        supabase_url = st.secrets.get("SUPABASE_URL")
        supabase_key = st.secrets.get("SUPABASE_KEY")
        
        if not supabase_url or not supabase_key:
            st.warning("⚠️ Supabase 설정이 완료되지 않았습니다. .streamlit/secrets.toml 파일을 확인해주세요.")
            return None
            
        if supabase_url == "https://your-project-id.supabase.co" or supabase_key == "your-anon-key-here":
            st.warning("⚠️ Supabase 설정이 기본값으로 되어 있습니다. 실제 프로젝트 정보로 업데이트해주세요.")
            return None
        
        # 단순한 Supabase 클라이언트 생성
        client = create_client(supabase_url, supabase_key)
        
        # 연결 테스트
        try:
            # 간단한 쿼리로 연결 확인
            test_result = client.table("daily_report_data").select("count", count="exact").execute()
            st.success("✅ Supabase 연결 성공!")
            return client
        except Exception as test_error:
            st.warning(f"⚠️ Supabase 연결 테스트 실패: {test_error}")
            st.info("💡 daily_report_data 테이블이 존재하지 않을 수 있습니다.")
            # 테이블이 없어도 클라이언트는 반환
            return client
            
    except Exception as e:
        st.error(f"❌ Supabase 연결 실패: {e}")
        st.info("💡 네트워크 설정을 확인해주세요.")
        return None

# 전역 Supabase 클라이언트 변수
supabase_client = None



# Supabase 클라이언트 초기화
if SUPABASE_AVAILABLE:
    try:
        supabase_client = init_supabase()
    except Exception as e:
        supabase_client = None
else:
    supabase_client = None

# Gemini AI 설정
genai.configure(api_key=GENAI_API_KEY)
GEMINI_MODEL = genai.GenerativeModel("models/gemini-2.5-flash-preview-05-20")

BLAST_EXTRACTION_PROMPT = '''
# INSTRUCTION
- 반드시 아래 예시처럼 오직 TSV(탭 구분) 데이터만 출력하세요.
- 설명, 마크다운, 코드블록, 주석, 기타 텍스트는 절대 포함하지 마세요.
- 아래 예시와 동일한 형식으로만 출력하세요.
발파일자    발파시간    지발당장약량(최소, kg)    지발당장약량(최대, kg)    폭약사용량(kg)    발파진동(cm/sec)    발파소음(dB(A))    계측위치    비고
2023-07-27    08:05    0.5    0.9    73    -    -    -    PLA-2
2023-07-27    13:47    0.4    0.8    77    0.87    53.29    티스테이션    PD-2
2023-07-27    13:47    -    -    -    0.71    61.23    양말집    PD-2
(위 예시는 형식만 참고, 실제 데이터는 입력값에 따라 동적으로 생성)
# 입력
- 입력1: 발파작업일지_TSV (아래와 같은 형식)
- 입력2: 계측일지_TSV (아래와 같은 형식, **계측일지 표는 PDF 2페이지 이후부터 추출**)
# 입력1 예시
발파일자    발파시간    지발당장약량(최소, kg)    지발당장약량(최대, kg)    폭약사용량(kg)    비고
2023-07-27    08:05    0.5    0.9    73    PLA-2
2023-07-27    13:47    0.4    0.8    77    PD-2
# 입력2 예시 (**2페이지 이후 표만**)
Date/Time    Peak Particle Vel (X_Axis) (mm/sec)    Peak Particle Vel (Y_Axis) (mm/sec)    Peak Particle Vel (Z_Axis) (mm/sec)    LMax (Sound) (dBA)    측정위치
2023/07/27 1:47:00 PM    0.71    0.36    0.71    61.23    양말집
2023/07/27 1:47:00 PM    0.87    0.56    0.87    53.29    티스테이션
# Mapping Rules
- 두 입력을 병합하여 위 예시와 동일한 TSV만 출력
- 설명, 마크다운, 코드블록, 주석, 기타 텍스트는 절대 포함하지 마세요.
- 계측일지 표는 반드시 PDF 2페이지 이후의 표만 사용 
- 최종 헤더(고정열): 발파일자, 발파시간, 지발당장약량(최소, kg), 지발당장약량(최대, kg), 폭약사용량(kg), 발파진동(cm/sec), 발파소음(dB(A)), 계측위치, 비고
- 정렬: 발파시간 오름차순, 계측위치 오름차순(필요시)
- 병합/매칭/포맷 규칙은 기존과 동일
'''
DEFAULT_PROMPT = """
# INSTRUCTIONS
1. 일일작업보고 원문에서 데이터를 파싱하여 4개 테이블("시공현황", "작업내용", "인원", "장비")과 각각을 TSV 형식의 코드블록으로 생성합니다.
2. 자동 검증 결과(QA-CHECKLIST)를 마크다운 표(Table)로 생성합니다.
3. 일일작업보고 텍스트에서 **작업 날짜**를 추출하여 첫 번째로 출력 (YYYY-MM-DD 형식)

# OUTPUT  
## 1. 시공현황 테이블  
1. 고정 열 : "구분", "누계"  
2. 고정 행(총 33행) - 아래 순서와 명칭을 그대로  
- "1. 본선터널 (1구간, 대림-신풍)  
- "1. 본선터널 (1구간, 대림-신풍) 라이닝" 
- "2. 신풍정거장 - 1)정거장 라이닝"
- "2. 신풍정거장 - 1)정거장 미들 슬라브"
- "2. 신풍정거장 – 2)주출입구 수직구 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 정거장 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 환승통로 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (2)PCC 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (3)PCD 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (4)PHA 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - 수직구 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - PHB 라이닝"
- "2. 신풍정거장 - 4)외부출입구 출입구(#3) 굴착" 
- "2. 신풍정거장 - 4)외부출입구 출입구(#2) 굴착"
- "2. 신풍정거장 - 4)외부출입구 출입구(#1) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 라이닝"  
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 라이닝"  
- "3. 신풍 환승통로 - 2)개착 BOX 보라매 방면 구조물"  
- "3. 신풍 환승통로 - 2)개착 BOX 대림 방면 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 터널 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 미들 슬라브" 
- "5. 도림사거리정거장 - 2)출입구#1 수직구 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCA 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCC 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PHA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 수직구 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCC 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PHB 라이닝"  

3. 추출데이터  
- "누계"값만 숫자로 추출할 것 (예:945.3m / 1,116m 에서 "945.3" 추출)

## 2. 작업내용 테이블  
1. 고정 열 : "구분", "금일작업"  
2. 고정 행(총 14행) - 아래 순서와 명칭(매핑 후 결과)을 그대로  
- "1. 본선터널 (1구간, 대림-신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍-도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"  
3. 주의사항  
- '작업내용' 셀은 여러 세부 내용을 포함할 수 있습니다. 내용을 구분할 때는, 최종 TSV 출력 시 해당 셀을 큰따옴표("...")로 감싸되, 셀 내부의 각 내용은 **실제 줄바꿈 문자(예: '\\n'문자열 대신 엔터 키 입력에 해당)**를 사용하여 분리하며, "-"기호는 생략함

## 3. 인원 / 장비 테이블  
1. 고정 열 (총 15열) - 열 순서는 아래와 같음
- "구분" 
- "1. 본선터널 (1구간, 대림~신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍~도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"    

2. 고정 행(인원 테이블 – 총 36행)  
(인원 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):

"직영반장", "연수생", "장비운전원", "전기주임", "화약주임", "터널공", "목공", "철근공", "라이닝폼공", "오폐수처리공", "카리프트공", "BP공", "가시설공", "설치공/해체공", "동바리공", "신호수", "부단수공", "슬러리월공", "CIP공", "미장공", "시설물공", "경계석공", "조경공", "배관공", "도색공", "방수공", "장비/작업지킴이", "보통인부", "포장공", "용접공", "타설공", "보링공/앙카공", "비계공", "도장공", "석면공", "주입공/그라우팅공"

3. 고정 행 (장비 테이블 – 총 46행)  
(장비 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):

"B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)", "덤프트럭(5T)", "덤프트럭(15T)", "덤프트럭(25T)", "앵글크레인(100T)", "앵글크레인(80T)", "앵글크레인(35T)", "앵글크레인(25T)", "카고크레인(25T)", "카고크레인(5T)", "콤프", "점보드릴", "페이로더", "숏트머신", "차징카", "살수차", "하이드로크레인", "믹서트럭", "화물차(5T)", "펌프카", "스카이", "콘크리트피니셔", "전주오거", "로더(바브켓)", "유제살포기(비우다)", "지게차", "싸인카", "BC커터기", "바이브로해머", "롤러(2.5T)", "롤러(1T)", "롤러(0.7T)", "몰리", "항타기", "크레인", "콤비로라", "공압드릴", "유압드릴", "기타"

## 4. Parsing Rules 
1. 시공현황: "누계/설계" → **앞값(소수 허용)** 만 추출.    
2. 인원·장비: 투입현황에서 **정수만** 추출, 빈셀은 **0**    
3. 하위 섹션 매핑   
   - 정거장 터널 → 열 ②, PCB → ③, PCC → ④, PCD → ⑤,PHA → ⑥, 특별피난 → ⑦, 외부출입구 →⑧    
4. 매핑 딕셔너리 적용    
- "B/H08W" →"B/H(08W)"   
- "25톤 카고크레인" → "카고크레인(25T)"   
- "특공" → "보통인부"    
- "기계타설공" → "타설공"    
- "목공연수생" 또는 "목수연수생" → "연수생"    
- "5톤트럭" → "화물차(5T)"    
- "카리프트" → "카리프트공"    
- "하이드로크레인(20T)" → "하이드로크레인"    
- "라이닝폼조립" → "라이닝폼공"  
- "S/C타설팀" → "터널공"  
- "목수" → "목공"    
5. 사전에 없는 항목 → 유사항목, 없으면 **인원: 보통인부 / 장비: 기타** 로 합산하고 '오류요약'에 기재.

## 5. QA-CHECKLIST(자동 검증 결과)

 1. 검증 항목
아래 기준에 따라 데이터 처리 과정의 정확성을 자체 검증하고, 그 결과를 마크다운 테이블로 생성합니다.
- **구조/형식**: 4개 테이블의 행과 열 개수, 순서, 데이터 형식(숫자, 정수, 0 처리)이 지침과 일치하는가?
- **데이터 무결성**: 원본 보고서의 인원 및 장비 수량이 누락되거나 중복되지 않고 100% 정확하게 집계되었는가?
- **매핑/변환**: 지정된 매핑 규칙(용어 표준화, 유사 항목 처리 등)이 모두 올바르게 적용되었는가?
- **미분류 항목**: 사전에 정의되지 않은 항목이 '보통인부' 또는 '기타'로 적절히 분류되고 기록되었는가?

2. 출력 방식
- **요약**: 검증 결과를 아래 예시와 같이 마크다운 테이블(`QA-CHECKLIST`)으로 요약합니다.
- **변환 내역**: 데이터 처리 과정에서 변경된 내용이 있는 경우, **'변환 내역'**란에 `원문 → 결과` 형식으로 명시합니다. 변경 사항이 없으면 "이상 없음"으로 표기합니다.

3. 예시 (마크다운 테이블)
| 점검 항목 | 검증 기준 | 변환 내역 (원문 → 결과) | 상태 |
| :--- | :--- | :--- | :---: |
| **구조 및 형식** | 4개 테이블의 구조 및 데이터 형식이 지침과 일치함 | 이상 없음 | ✅ |
| **데이터 무결성** | 인원(85명), 장비(12대) 총계가 원문과 일치함 | 이상 없음 | ✅ |
| **용어 표준화** | 매핑 규칙에 따라 용어가 일괄 변환됨 | - 목수 → 목공<br>- 특공 → 보통인부<br>- B/H08W → B/H(08W) | ✅ |
| **미분류 항목 처리** | 사전에 없는 항목을 규칙에 따라 처리함 | - 노무원 → 보통인부 (합산) | ⚠️ |

---

# USER TEXT(작업계획보고 입력란)
"""

# --- PROMPT MANAGEMENT FUNCTIONS ---
def save_prompt_to_supabase(prompt_name, prompt_content, description=""):
    """프롬프트를 Supabase에 저장합니다."""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return False
    
    try:
        data = {
            "name": prompt_name,
            "content": prompt_content,
            "description": description,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        # 기존 프롬프트가 있는지 확인
        existing = supabase_client.table("prompts").select("*").eq("name", prompt_name).execute()
        
        if existing.data:
            # 업데이트
            data["updated_at"] = datetime.now().isoformat()
            result = supabase_client.table("prompts").update(data).eq("name", prompt_name).execute()
            st.success(f"✅ 프롬프트 '{prompt_name}'이(가) 업데이트되었습니다.")
        else:
            # 새로 생성
            result = supabase_client.table("prompts").insert(data).execute()
            st.success(f"✅ 프롬프트 '{prompt_name}'이(가) 저장되었습니다.")
        
        return True
    except Exception as e:
        st.error(f"❌ 프롬프트 저장 실패: {e}")
        return False

def load_prompt_from_supabase(prompt_name):
    """Supabase에서 프롬프트를 로드합니다."""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return None
    
    try:
        # SSL 오류 방지를 위한 추가 설정
        import ssl
        import urllib3
        import os
        
        # SSL 경고 비활성화
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        os.environ['PYTHONHTTPSVERIFY'] = '0'
        
        result = supabase_client.table("prompts").select("*").eq("name", prompt_name).execute()
        if result.data:
            return result.data[0]
        return None
    except Exception as e:
        st.error(f"❌ 프롬프트 로드 실패: {e}")
        st.info("💡 SSL 인증서 문제일 수 있습니다. 네트워크 설정을 확인해주세요.")
        return None

def get_all_prompts_from_supabase():
    """Supabase에서 모든 프롬프트 목록을 가져옵니다."""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return []
    
    try:
        # SSL 오류 방지를 위한 추가 설정
        import ssl
        import urllib3
        import os
        
        # SSL 경고 비활성화
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        os.environ['PYTHONHTTPSVERIFY'] = '0'
        
        result = supabase_client.table("prompts").select("name, description, updated_at").execute()
        return result.data if result.data else []
    except Exception as e:
        st.error(f"❌ 프롬프트 목록 로드 실패: {e}")
        st.info("💡 SSL 인증서 문제일 수 있습니다. 네트워크 설정을 확인해주세요.")
        return []

def delete_prompt_from_supabase(prompt_name):
    """Supabase에서 프롬프트를 삭제합니다."""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return False
    
    try:
        result = supabase_client.table("prompts").delete().eq("name", prompt_name).execute()
        st.success(f"✅ 프롬프트 '{prompt_name}'이(가) 삭제되었습니다.")
        return True
    except Exception as e:
        st.error(f"❌ 프롬프트 삭제 실패: {e}")
        return False

def extract_table_structure_from_prompt(prompt_text):
    """프롬프트 텍스트에서 테이블 구조를 추출합니다."""
    extracted_tables = {}
    
    try:
        # 시공현황 테이블 추출
        construction_match = re.search(r'시공현황.*?고정 행.*?총 (\d+)행.*?아래 순서와 명칭을 그대로(.*?)(?=## 2\.|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if construction_match:
            construction_lines = construction_match.group(2)
            construction_items = re.findall(r'- "([^"]+)"', construction_lines)
            if construction_items:
                extracted_tables['construction'] = construction_items
        
        # 작업내용 테이블 추출
        work_match = re.search(r'작업내용.*?고정 행.*?총 (\d+)행.*?아래 순서와 명칭.*?을 그대로(.*?)(?=## 3\.|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if work_match:
            work_lines = work_match.group(2)
            work_items = re.findall(r'- "([^"]+)"', work_lines)
            if work_items:
                extracted_tables['work_content'] = work_items
        
        # 인원 테이블 열 추출
        personnel_col_match = re.search(r'고정 열.*?총 (\d+)열.*?아래와 같음(.*?)(?=고정 행|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if personnel_col_match:
            col_lines = personnel_col_match.group(2)
            col_items = re.findall(r'- "([^"]+)"', col_lines)
            if col_items:
                extracted_tables['personnel_columns'] = col_items
        
        # 인원 테이블 행 추출
        personnel_row_match = re.search(r'인원 테이블.*?총 (\d+)행.*?아래 순서와 명칭.*?을 반드시 그대로 사용(.*?)(?=고정 행.*?장비|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if personnel_row_match:
            row_lines = personnel_row_match.group(2)
            row_items = re.findall(r'"([^"]+)"', row_lines)
            if row_items:
                extracted_tables['personnel_rows'] = row_items
        
        # 장비 테이블 추출
        equipment_match = re.search(r'장비 테이블.*?총 (\d+)행.*?아래 순서와 명칭.*?을 반드시 그대로 사용(.*?)(?=## 4\.|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if equipment_match:
            equipment_lines = equipment_match.group(2)
            equipment_items = re.findall(r'"([^"]+)"', equipment_lines)
            if equipment_items:
                extracted_tables['equipment'] = equipment_items
        
        return extracted_tables if extracted_tables else None
        
    except Exception as e:
        st.error(f"테이블 구조 추출 중 오류: {e}")
        return None

def generate_prompt_from_tables():
    """테이블 구조 데이터를 기반으로 프롬프트를 생성합니다."""
    
    # 시공현황 행 목록
    construction_rows = st.session_state.construction_rows
    
    # 작업내용 행 목록
    work_content_rows = st.session_state.work_content_rows
    
    # 인원/장비 열 목록
    personnel_columns = st.session_state.personnel_columns
    personnel_rows = st.session_state.personnel_rows
    equipment_rows = st.session_state.equipment_rows
    
    prompt = f"""# INSTRUCTIONS
1. 일일작업보고 원문에서 데이터를 파싱하여 4개 테이블("시공현황", "작업내용", "인원", "장비")과 각각을 TSV 형식의 코드블록으로 생성합니다.
2. 자동 검증 결과(QA-CHECKLIST)를 마크다운 표(Table)로 생성합니다.
3. 일일작업보고 텍스트에서 **작업 날짜**를 추출하여 첫 번째로 출력 (YYYY-MM-DD 형식)

# OUTPUT  
## 1. 시공현황 테이블  
1. 고정 열 : "구분", "누계"  
2. 고정 행(총 {len(construction_rows)}행) - 아래 순서와 명칭을 그대로  
{chr(10).join([f'- "{row}"' for row in construction_rows])}
3. 추출데이터  
- "누계"값만 숫자로 추출할 것 (예:945.3m / 1,116m 에서 "945.3" 추출)

## 2. 작업내용 테이블  
1. 고정 열 : "구분", "금일작업"  
2. 고정 행(총 {len(work_content_rows)}행) - 아래 순서와 명칭(매핑 후 결과)을 그대로  
{chr(10).join([f'- "{row}"' for row in work_content_rows])}
3. 주의사항  
- '작업내용' 셀은 여러 세부 내용을 포함할 수 있습니다. 내용을 구분할 때는, 최종 TSV 출력 시 해당 셀을 큰따옴표("...")로 감싸되, 셀 내부의 각 내용은 **실제 줄바꿈 문자(예: '\\n'문자열 대신 엔터 키 입력에 해당)**를 사용하여 분리하며, "-"기호는 생략함

## 3. 인원 / 장비 테이블  
1. 고정 열 (총 {len(personnel_columns) + 1}열) - 열 순서는 아래와 같음
- "구분" 
{chr(10).join([f'- "{col}"' for col in personnel_columns])}
2. 고정 행(인원 테이블 – 총 {len(personnel_rows)}행)  
(인원 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):
{', '.join([f'"{row}"' for row in personnel_rows])}
3. 고정 행 (장비 테이블 – 총 {len(equipment_rows)}행)  
(장비 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):
{', '.join([f'"{row}"' for row in equipment_rows])}

## 4. Parsing Rules 
1. 시공현황: "누계/설계" → **앞값(소수 허용)** 만 추출.    
2. 인원·장비: 투입현황에서 **정수만** 추출, 빈셀은 **0**    
3. 하위 섹션 매핑   
   - 정거장 터널 → 열 ②, PCB → ③, PCC → ④, PCD → ⑤,PHA → ⑥, 특별피난 → ⑦, 외부출입구 →⑧    
4. 매핑 딕셔너리 적용    
- "B/H08W" →"B/H(08W)"   
- "25톤 카고크레인" → "카고크레인(25T)"   
- "특공" → "보통인부"    
- "기계타설공" → "타설공"    
- "목공연수생" 또는 "목수연수생" → "연수생"    
- "5톤트럭" → "화물차(5T)"    
- "카리프트" → "카리프트공"    
- "하이드로크레인(20T)" → "하이드로크레인"    
- "라이닝폼조립" → "라이닝폼공"  
- "S/C타설팀" → "터널공"  
- "목수" → "목공"    
5. 사전에 없는 항목 → 유사항목, 없으면 **인원: 보통인부 / 장비: 기타** 로 합산하고 '오류요약'에 기재.

## 5. QA-CHECKLIST(자동 검증 결과)

 1. 검증 항목
아래 기준에 따라 데이터 처리 과정의 정확성을 자체 검증하고, 그 결과를 마크다운 테이블로 생성합니다.
- **구조/형식**: 4개 테이블의 행과 열 개수, 순서, 데이터 형식(숫자, 정수, 0 처리)이 지침과 일치하는가?
- **데이터 무결성**: 원본 보고서의 인원 및 장비 수량이 누락되거나 중복되지 않고 100% 정확하게 집계되었는가?
- **매핑/변환**: 지정된 매핑 규칙(용어 표준화, 유사 항목 처리 등)이 모두 올바르게 적용되었는가?
- **미분류 항목**: 사전에 정의되지 않은 항목이 '보통인부' 또는 '기타'로 적절히 분류되고 기록되었는가?

2. 출력 방식
- **요약**: 검증 결과를 아래 예시와 같이 마크다운 테이블(`QA-CHECKLIST`)으로 요약합니다.
- **변환 내역**: 데이터 처리 과정에서 변경된 내용이 있는 경우, **'변환 내역'**란에 `원문 → 결과` 형식으로 명시합니다. 변경 사항이 없으면 "이상 없음"으로 표기합니다.

3. 예시 (마크다운 테이블)
| 점검 항목 | 검증 기준 | 변환 내역 (원문 → 결과) | 상태 |
| :--- | :--- | :--- | :---: |
| **구조 및 형식** | 4개 테이블의 구조 및 데이터 형식이 지침과 일치함 | 이상 없음 | ✅ |
| **데이터 무결성** | 인원(85명), 장비(12대) 총계가 원문과 일치함 | 이상 없음 | ✅ |
| **용어 표준화** | 매핑 규칙에 따라 용어가 일괄 변환됨 | - 목수 → 목공<br>- 특공 → 보통인부<br>- B/H08W → B/H(08W) | ✅ |
| **미분류 항목 처리** | 사전에 없는 항목을 규칙에 따라 처리함 | - 노무원 → 보통인부 (합산) | ⚠️ |

---

# USER TEXT(작업계획보고 입력란)
"""
    
    return prompt

# --- HELPER FUNCTIONS ---
def safe_generate_content(model_input):
    """
    Calls the Gemini API with robust error handling and relaxed safety settings.
    """
    try:
        # AI 모델의 안전 설정을 완화하여 콘텐츠 차단을 최소화합니다.
        safety_settings = {
            'HARM_CATEGORY_HARASSMENT': 'BLOCK_NONE',
            'HARM_CATEGORY_HATE_SPEECH': 'BLOCK_NONE',
            'HARM_CATEGORY_SEXUALLY_EXPLICIT': 'BLOCK_NONE',
            'HARM_CATEGORY_DANGEROUS_CONTENT': 'BLOCK_NONE',
        }
        
        response = GEMINI_MODEL.generate_content(
            model_input,
            safety_settings=safety_settings
        )

        # 응답에 실제 콘텐츠(parts)가 있는지 확인합니다.
        if response.parts:
            return response.text
        else:
            # 콘텐츠가 없는 경우, 차단 원인을 확인하여 사용자에게 알립니다.
            reason = "Unknown"
            try:
                # API 응답에서 제공하는 공식적인 차단 이유를 가져옵니다.
                reason = response.prompt_feedback.block_reason.name
            except Exception:
                pass 
            st.error(f"AI 응답 생성에 실패했습니다. API에 의해 콘텐츠가 차단되었을 수 있습니다. (차단 이유: {reason})")
            st.warning(f"전체 피드백: {response.prompt_feedback}")
            return None
            
    except Exception as e:
        st.error(f"AI 모델 호출 중 심각한 오류 발생: {e}")
        return None

def sanitize_prompt(prompt_text):
    """프롬프트에서 민감한 단어를 제거하거나 대체합니다."""
    # 민감한 단어들을 더 중립적인 표현으로 대체
    replacements = {
        '특공': '특수공',
        '폭파': '파쇄',
        '폭발': '파쇄',
        '폭약': '파쇄재',
        '다이너마이트': '파쇄재',
        'TNT': '파쇄재',
        '폭파공': '파쇄공',
        '폭파작업': '파쇄작업',
        '폭파공사': '파쇄공사',
        '폭파장': '파쇄장',
        '폭파구': '파쇄구',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파장비': '파쇄장비',
        '폭파구멍': '파쇄구',
    }
    
    sanitized_text = prompt_text
    for old_word, new_word in replacements.items():
        sanitized_text = sanitized_text.replace(old_word, new_word)
    
    # 추가적인 정제 작업
    # 1. 모든 안전 관련 단어를 중립적으로 변경
    safety_replacements = {
        '위험': '주의',
        '사고': '사건',
        '사망': '인명피해',
        '부상': '상해',
        '폭발물': '파쇄재',
        '폭발성': '파쇄성',
        '폭발력': '파쇄력',
    }
    
    for old_word, new_word in safety_replacements.items():
        sanitized_text = sanitized_text.replace(old_word, new_word)
    
    # 2. 프롬프트 시작 부분에 안전성 강조 문구 추가
    safety_prefix = """[안전한 건설업무 데이터 분석 지시사항]
이 지시사항은 건설현장의 일일작업보고를 안전하고 정확하게 분석하기 위한 것입니다.
모든 내용은 건설업무의 정상적인 업무처리를 위한 것입니다.

"""
    
    return safety_prefix + sanitized_text

def create_simplified_prompt(original_prompt):
    """차단된 경우 사용할 간단한 프롬프트를 생성합니다."""
    # 원본 프롬프트에서 사용자 입력 부분만 추출
    user_input_start = original_prompt.find('"""')
    if user_input_start != -1:
        user_input = original_prompt[user_input_start + 3:]
        user_input = user_input.replace('"""', '').strip()
    else:
        user_input = original_prompt
    
    simplified_prompt = f"""건설현장 일일작업보고를 분석하여 다음 형식으로 정리해주세요:

1. 작업 날짜 추출 (YYYY-MM-DD 형식)
2. 시공현황 테이블 (구분, 누계)
3. 작업내용 테이블 (구분, 금일작업)
4. 인원 테이블 (구분, 인원수)
5. 장비 테이블 (구분, 대수)

각 테이블은 TSV 형식으로 출력해주세요.

작업보고:
{user_input}"""
    
    return simplified_prompt

def create_minimal_prompt(original_prompt):
    """최소한의 프롬프트를 생성합니다."""
    # 원본 프롬프트에서 사용자 입력 부분만 추출
    user_input_start = original_prompt.find('"""')
    if user_input_start != -1:
        user_input = original_prompt[user_input_start + 3:]
        user_input = user_input.replace('"""', '').strip()
    else:
        user_input = original_prompt
    
    minimal_prompt = f"""건설현장 일일작업보고를 분석하여 테이블로 정리해주세요.

작업보고:
{user_input}"""
    
    return minimal_prompt

def send_teams_alert(warning_rows, file_date):
    try:
        message = {
            "type": "message",
            "attachments": [{
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": {
                    "type": "AdaptiveCard",
                    "body": [
                        {"type": "TextBlock", "size": "Large", "weight": "Bolder", "text": f"⚠️ 계측기 경고 알림 ({file_date})", "color": "Attention"},
                        {"type": "TextBlock", "text": "다음 계측기에서 주의가 필요한 변화가 감지되었습니다:", "wrap": True}
                    ]
                }
            }]
        }
        for _, row in warning_rows.iterrows():
            warning_info = {"type": "TextBlock", "text": f"📍 위치: {row['위치']}\\n\\n📊 계측기: {row['계측기명']} ({row['계측기 종류']})\\n\\n⚠️ 상태: {row['상태']}\\n\\n📈 3차 초과 대비: {row['비율']}", "wrap": True, "style": "warning"}
            message["attachments"][0]["content"]["body"].append(warning_info)
        
        response = requests.post(TEAMS_WEBHOOK_URL, json=message, headers={"Content-Type": "application/json"})
        if response.status_code == 200: st.success("Teams로 경고 메시지가 전송되었습니다.")
        else: st.error(f"Teams 메시지 전송 실패: {response.status_code}")
    except Exception as e: st.error(f"Teams 메시지 전송 중 오류 발생: {e}")

def save_step1_to_supabase(data, date=None):
    """1단계 AI 데이터 추출 결과를 4개 별도 테이블에 저장합니다."""
    global supabase_client
    
    if not supabase_client:
        st.error("❌ Supabase 연결이 설정되지 않았습니다.")
        return False
    
    try:
        current_date = date or datetime.now().strftime('%Y-%m-%d')
        
        success_count = 0
        total_tables = 4
        
        # 1. 시공현황 테이블 저장
        if data.get("시공현황"):
            try:
                construction_records = []
                for item in data["시공현황"]:
                    construction_records.append({
                        "date": current_date,
                        "구분": item.get("구분", ""),
                        "누계": item.get("누계", ""),
                        "created_at": datetime.now().isoformat()
                    })
                
                result = supabase_client.table("construction_status").upsert(construction_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"❌ 시공현황 저장 실패: {e}")
        
        # 2. 작업내용 테이블 저장
        if data.get("작업내용"):
            try:
                work_content_records = []
                for item in data["작업내용"]:
                    work_content_records.append({
                        "date": current_date,
                        "구분": item.get("구분", ""),
                        "금일작업": item.get("금일작업", ""),
                        "created_at": datetime.now().isoformat()
                    })
                
                result = supabase_client.table("work_content").upsert(work_content_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"❌ 작업내용 저장 실패: {e}")
        
        # 3. 인원 데이터 테이블 저장
        if data.get("인원"):
            try:
                personnel_records = []
                for item in data["인원"]:
                    # 기존 테이블 구조에 맞게 데이터 저장
                    personnel_record = {
                        "date": current_date,
                        "구분": item.get("구분", ""),
                        "created_at": datetime.now().isoformat()
                    }
                    personnel_records.append(personnel_record)
                
                result = supabase_client.table("personnel_data").upsert(personnel_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"❌ 인원 데이터 저장 실패: {e}")
        
        # 4. 장비 데이터 테이블 저장
        if data.get("장비"):
            try:
                equipment_records = []
                for item in data["장비"]:
                    # 기존 테이블 구조에 맞게 데이터 저장
                    equipment_record = {
                        "date": current_date,
                        "구분": item.get("구분", ""),
                        "created_at": datetime.now().isoformat()
                    }
                    equipment_records.append(equipment_record)
                
                result = supabase_client.table("equipment_data").upsert(equipment_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"❌ 장비 데이터 저장 실패: {e}")
        
        if success_count == total_tables:
            st.success("✅ 1단계 데이터 저장 완료")
            return True
        else:
            st.warning(f"⚠️ 일부 데이터 저장에 실패했습니다. ({success_count}/{total_tables})")
            return False
        
    except Exception as e:
        st.error(f"❌ 1단계 데이터 저장 실패: {e}")
        import traceback
        st.error(f"❌ 상세 오류: {traceback.format_exc()}")
        return False

def save_to_supabase(data_type, data, date=None):
    """데이터를 Supabase에 저장합니다."""
    global supabase_client
    
    if not supabase_client:
        st.error("❌ Supabase 연결이 설정되지 않았습니다.")
        return False
    
    try:
        current_date = date or datetime.now().strftime('%Y-%m-%d')
        
        # 데이터 타입에 따라 다른 테이블에 저장
        if data_type == "daily_report":
            try:
                # 엑셀 파일에서 특정 셀 데이터 추출
                if "excel_bytes" in data:
                    extracted_data = extract_cell_data_from_excel(data["excel_bytes"], current_date)
                    
                    if not extracted_data:
                        st.error("❌ 엑셀 데이터 추출에 실패했습니다.")
                        return False
                    
                    # daily_report_data 테이블에 저장
                    report_data = {
                        "date": current_date,
                        "construction_data": extracted_data.get("construction_data", {}),
                        "personnel_data": extracted_data.get("personnel_data", {}),
                        "equipment_data": extracted_data.get("equipment_data", {}),
                        "created_at": datetime.now().isoformat(),
                        "updated_at": datetime.now().isoformat()
                    }
                    
                    try:
                        # 기존 데이터 확인
                        existing_data = supabase_client.table("daily_report_data").select("*").eq("date", current_date).execute()
                        
                        if existing_data.data:
                            # 기존 데이터 업데이트
                            result = supabase_client.table("daily_report_data").update(report_data).eq("date", current_date).execute()
                            st.success("✅ 3단계 데이터 저장 완료")
                        else:
                            # 새 데이터 삽입
                            result = supabase_client.table("daily_report_data").insert(report_data).execute()
                            st.success("✅ 3단계 데이터 저장 완료")
                        
                        return True
                        
                    except Exception as table_error:
                        st.error(f"❌ Supabase 저장 실패: {table_error}")
                        return False
                    
                else:
                    st.error("❌ 엑셀 파일 데이터가 없습니다.")
                    return False
                
            except Exception as e:
                st.error(f"❌ 데이터 저장 실패: {e}")
                return False
            
        elif data_type == "blast_data":
            # 발파 데이터
            blast_records = []
            for _, row in data.iterrows():
                blast_records.append({
                    "date": current_date,
                    "blast_date": row.get("발파일자", ""),
                    "blast_time": row.get("발파시간", ""),
                    "min_charge": row.get("지발당장약량(최소, kg)", ""),
                    "max_charge": row.get("지발당장약량(최대, kg)", ""),
                    "explosive_usage": row.get("폭약사용량(kg)", ""),
                    "vibration": row.get("발파진동(cm/sec)", ""),
                    "noise": row.get("발파소음(dB(A))", ""),
                    "measurement_location": row.get("계측위치", ""),
                    "remarks": row.get("비고", ""),
                    "created_at": datetime.now().isoformat()
                })
            
            result = supabase_client.table("blast_data").insert(blast_records).execute()
            st.success("✅ 발파 데이터 저장 완료")
            
        elif data_type == "instrument_data":
            # 계측기 데이터
            instrument_records = []
            for _, row in data.iterrows():
                instrument_records.append({
                    "date": current_date,
                    "location": row.get("위치", ""),
                    "instrument_type": row.get("계측기 종류", ""),
                    "instrument_name": row.get("계측기명", ""),
                    "weekly_change": row.get("주간변화량", ""),
                    "cumulative_change": row.get("누적변화량", ""),
                    "unit": row.get("단위", ""),
                    "status": row.get("상태", ""),
                    "ratio": row.get("비율", ""),
                    "created_at": datetime.now().isoformat()
                })
            
            result = supabase_client.table("instrument_data").insert(instrument_records).execute()
            st.success("✅ 계측기 데이터가 Supabase에 저장되었습니다.")
            
        return True
        
    except Exception as e:
        st.error(f"❌ Supabase 저장 실패: {e}")
        return False

def load_from_supabase(data_type, date=None):
    """Supabase에서 데이터를 로드합니다."""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return None
    
    try:
        current_date = date or datetime.now().strftime('%Y-%m-%d')
        
        if data_type == "daily_report":
            result = supabase_client.table("daily_reports").select("*").eq("date", current_date).execute()
            if result.data:
                return result.data[0]  # 첫 번째 레코드 반환
            return None
            
        elif data_type == "blast_data":
            result = supabase_client.table("blast_data").select("*").eq("date", current_date).execute()
            return pd.DataFrame(result.data) if result.data else None
            
        elif data_type == "instrument_data":
            result = supabase_client.table("instrument_data").select("*").eq("date", current_date).execute()
            return pd.DataFrame(result.data) if result.data else None
            
    except Exception as e:
        st.error(f"❌ Supabase 로드 실패: {e}")
        return None

def save_template_to_supabase(template_bytes, template_name="default", description=""):
    """엑셀 템플릿을 Supabase에 저장"""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return False, "Supabase 연결 실패"
    
    try:
        # 템플릿 데이터를 base64로 인코딩
        import base64
        template_base64 = base64.b64encode(template_bytes).decode('utf-8')
        
        # 데이터 크기 확인 (Supabase 제한: 1MB)
        if len(template_base64) > 1000000:  # 약 1MB
            return False, "템플릿 파일이 너무 큽니다. 1MB 이하의 파일을 사용해주세요."
        
        # 템플릿 데이터 준비
        template_data = {
            'template_name': template_name,
            'template_data': template_base64,
            'description': description,
            'is_default': template_name == "default",
            'created_at': datetime.now().isoformat()
        }
        
        # 기존 템플릿이 있으면 업데이트, 없으면 새로 생성
        existing = supabase_client.table('templates')\
            .select('id')\
            .eq('template_name', template_name)\
            .execute()
        
        if existing.data:
            # 기존 템플릿 업데이트
            result = supabase_client.table('templates')\
                .update(template_data)\
                .eq('template_name', template_name)\
                .execute()
        else:
            # 새 템플릿 생성
            result = supabase_client.table('templates')\
                .insert(template_data)\
                .execute()
        
        if result.data:
            return True, "템플릿 저장 성공"
        else:
            return False, "템플릿 저장 실패: 데이터가 저장되지 않았습니다."
            
    except Exception as e:
        st.error(f"템플릿 저장 중 오류: {e}")
        return False, f"템플릿 저장 실패: {str(e)}"

def get_template_from_supabase(template_name="default"):
    """Supabase에서 엑셀 템플릿 불러오기"""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return None
    
    try:
        result = supabase_client.table('templates')\
            .select('*')\
            .eq('template_name', template_name)\
            .order('created_at', desc=True)\
            .limit(1)\
            .execute()
        
        if result.data:
            # base64 디코딩
            import base64
            template_base64 = result.data[0]['template_data']
            template_bytes = base64.b64decode(template_base64)
            return template_bytes
        return None
    except Exception as e:
        st.error(f"템플릿 불러오기 실패: {e}")
        return None

def get_all_templates():
    """모든 템플릿 목록 조회"""
    global supabase_client
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return []
    
    try:
        result = supabase_client.table('templates')\
            .select('template_name, description, created_at, is_default')\
            .order('created_at', desc=True)\
            .execute()
        
        return result.data
    except Exception as e:
        st.error(f"템플릿 목록 조회 실패: {e}")
        return []

def check_templates_table():
    """templates 테이블 존재 여부 확인"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase 연결 실패"
    
    try:
        # 간단한 쿼리로 테이블 존재 확인
        result = supabase_client.table('templates').select('id').limit(1).execute()
        return True, "테이블 존재"
    except Exception as e:
        return False, f"테이블 확인 실패: {str(e)}"

def check_daily_report_data_table():
    """daily_report_data 테이블 존재 여부 확인"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase 연결 실패"
    
    try:
        # 간단한 쿼리로 테이블 존재 확인
        result = supabase_client.table('daily_report_data').select('date').limit(1).execute()
        return True, "테이블 존재"
    except Exception as e:
        return False, f"테이블 확인 실패: {str(e)}"

def test_supabase_connection():
    """Supabase 연결을 테스트합니다."""
    global supabase_client
    if not supabase_client:
        return False, "Supabase 클라이언트가 없습니다."
    
    try:
        # 간단한 쿼리로 연결 테스트
        result = supabase_client.table("daily_report_data").select("count", count="exact").execute()
        return True, "연결 성공"
    except Exception as e:
        return False, f"연결 실패: {str(e)}"

def create_daily_report_data_table():
    """daily_report_data 테이블 생성"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase 연결 실패"
    
    try:
        st.write("🔍 테이블 생성 시도 중...")
        
        # 테이블이 없으면 빈 데이터로 테스트 삽입 시도
        test_data = {
            "date": "2024-01-01",
            "construction_data": {},
            "personnel_data": {},
            "equipment_data": {},
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        result = supabase_client.table("daily_report_data").insert(test_data).execute()
        st.success("✅ 테이블 생성 성공!")
        
        # 테스트 데이터 삭제
        supabase_client.table("daily_report_data").delete().eq("date", "2024-01-01").execute()
        
        return True, "테이블 생성 완료"
    except Exception as e:
        error_msg = str(e).lower()
        if "duplicate key" in error_msg or "unique" in error_msg:
            st.write("ℹ️ 테이블이 이미 존재합니다.")
            return True, "테이블 이미 존재"
        else:
            return False, f"테이블 생성 실패: {str(e)}"

def save_cell_mapping_to_supabase(mapping_data, mapping_name="default"):
    """셀 매핑 설정을 Supabase에 저장"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase 연결 실패"
    
    try:
        result = supabase_client.table('cell_mappings').upsert({
            'mapping_name': mapping_name,
            'mapping_data': mapping_data,
            'created_at': datetime.now().isoformat()
        }).execute()
        
        return True, "매핑 설정 저장 성공"
    except Exception as e:
        return False, f"매핑 설정 저장 실패: {str(e)}"

def get_cell_mapping_from_supabase(mapping_name="default"):
    """Supabase에서 셀 매핑 설정 불러오기"""
    global supabase_client
    if not supabase_client:
        return None
    
    try:
        result = supabase_client.table('cell_mappings')\
            .select('mapping_data')\
            .eq('mapping_name', mapping_name)\
            .order('created_at', desc=True)\
            .limit(1)\
            .execute()
        
        if result.data:
            return result.data[0]['mapping_data']
        return None
    except Exception as e:
        st.error(f"매핑 설정 불러오기 실패: {e}")
        return None

def extract_file_content(file):
    if file.name.endswith('.pdf'):
        try:
            file.seek(0)
            uploaded_file = genai.upload_file(file, mime_type="application/pdf")
            
            filename_lower = file.name.lower()
            is_measurement_file = any(keyword in filename_lower for keyword in ["계측", "진동", "소음"])
            is_blast_log_file = any(keyword in filename_lower for keyword in ["발파", "작업", "일지"])

            if is_measurement_file:
                pdf_prompt = "이 PDF 파일은 '발파진동소음 계측일지'입니다. 다음 지침에 따라 데이터를 TSV 형식으로 추출해주세요. ... (Prompt content is long and omitted for brevity)"
            elif is_blast_log_file:
                pdf_prompt = "이 PDF 파일은 '발파작업일지'입니다. 다음 지침에 따라 주요 데이터를 TSV 형식으로 추출해주세요. ... (Prompt content is long and omitted for brevity)"
            else:
                st.warning("⚠️ 파일 유형을 특정할 수 없어 일반 표 추출을 시도합니다.")
                pdf_prompt = "이 PDF에서 가장 중요해 보이는 표를 찾아 TSV 형식으로 추출해주세요. ..."

            # 안전하게 AI 모델을 호출합니다.
            response_text = safe_generate_content([pdf_prompt, uploaded_file])
            
            # 사용이 끝난 파일은 즉시 삭제합니다.
            genai.delete_file(uploaded_file.name)

            if response_text:
                return re.sub(r'```tsv|```', '', response_text).strip()
            
            return None # safe_generate_content에서 오류를 이미 표시했으므로 None만 반환합니다.

        except Exception as e:
            st.error(f"❌ {file.name} 처리 중 AI 오류 발생: {e}")
            return None
    elif file.name.endswith(('.xlsx', '.xls')):
        try:
            return pd.read_excel(file, engine='openpyxl').to_csv(sep='\t', index=False, encoding='utf-8')
        except Exception as e:
            st.error(f"❌ 엑셀 데이터 추출 실패: {e}")
            return None
    return None

def parse_tsv_to_dataframe(tsv_content):
    try:
        if not tsv_content or tsv_content.strip() == '':
            st.warning("⚠️ 빈 TSV 데이터입니다.")
            return None
        
        # main2.py 방식으로 간단하게 처리
        cleaned_content = '\n'.join(line.strip() for line in tsv_content.split('\n') if line.strip())
        
        if not cleaned_content:
            st.warning("⚠️ 정제된 TSV 데이터가 없습니다.")
            return None
        
        df = pd.read_csv(io.StringIO(cleaned_content), sep='\t', encoding='utf-8')
        
        if df.empty:
            st.warning("⚠️ 파싱된 데이터프레임이 비어있습니다.")
            return None
        
        df.columns = df.columns.str.strip()
        
        # Arrow 직렬화 오류 방지를 위한 데이터 정제
        for col in df.columns:
            # 빈 문자열을 '0' 또는 적절한 기본값으로 변경
            df[col] = df[col].fillna('').astype(str)
            # 숫자처럼 보이는 컬럼의 빈 값을 '0'으로 변경
            if '누계' in col or '변화량' in col or any(x in col for x in ['명', '대', 'kg', 'cm/sec', 'dB']):
                df[col] = df[col].replace('', '0')
        
        st.info(f"✅ TSV 파싱 성공: {len(df)}행, {len(df.columns)}열")
        return df
        
    except Exception as e:
        st.error(f"TSV 파싱 중 오류 발생: {e}")
        st.info(f"🔍 원본 TSV 데이터 (처음 200자): {tsv_content[:200]}")
        return None

def extract_work_date_from_response(response_text):
    """AI 응답에서 작업 날짜를 추출합니다."""
    if not response_text:
        return datetime.now().strftime('%Y-%m-%d')
    
    # WORK_DATE: YYYY-MM-DD 패턴 찾기
    date_pattern = r'WORK_DATE:\s*(\d{4}-\d{2}-\d{2})'
    match = re.search(date_pattern, response_text)
    
    if match:
        return match.group(1)
    
    # 대안 패턴들 시도 (2024년 1월 15일, 24.01.15 등)
    alt_patterns = [
        r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일',
        r'(\d{2})\.(\d{2})\.(\d{2})',
        r'(\d{4})-(\d{2})-(\d{2})',
        r'(\d{4})/(\d{2})/(\d{2})'
    ]
    
    for pattern in alt_patterns:
        match = re.search(pattern, response_text)
        if match:
            groups = match.groups()
            if len(groups) == 3:
                year, month, day = groups
                if len(year) == 2:  # 24.01.15 형식인 경우
                    year = '20' + year
                try:
                    # 날짜 유효성 검증
                    datetime.strptime(f"{year}-{month.zfill(2)}-{day.zfill(2)}", '%Y-%m-%d')
                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                except ValueError:
                    continue
    
    # 추출 실패시 현재 날짜 사용
    return datetime.now().strftime('%Y-%m-%d')

def extract_qa_checklist_from_response(response_text):
    """AI 응답에서 QA-Checklist를 추출합니다."""
    if not response_text:
        return ""
    
    # QA-CHECKLIST 섹션 찾기
    qa_patterns = [
        r'QA-CHECKLIST[\s\S]*?(?=\n\n|\n#|\n##|\n###|\n####|\n#####|\n######|$)',
        r'QA-CHECKLIST[\s\S]*',
        r'## 5\. QA-CHECKLIST[\s\S]*?(?=\n\n|\n#|\n##|\n###|\n####|\n#####|\n######|$)',
        r'## 5\. QA-CHECKLIST[\s\S]*'
    ]
    
    for pattern in qa_patterns:
        match = re.search(pattern, response_text, re.IGNORECASE | re.DOTALL)
        if match:
            qa_content = match.group(0).strip()
            # 불필요한 헤더 제거
            qa_content = re.sub(r'^## 5\. QA-CHECKLIST.*?\n', '', qa_content, flags=re.IGNORECASE | re.DOTALL)
            qa_content = re.sub(r'^QA-CHECKLIST.*?\n', '', qa_content, flags=re.IGNORECASE | re.DOTALL)
            return qa_content.strip()
    
    return ""

def convert_to_number_if_possible(value):
    """값이 숫자로 변환 가능한지 확인하고 숫자로 변환합니다."""
    if value is None or value == "":
        return 0
    
    # 문자열인 경우 공백 제거
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return 0
    
    try:
        # 정수로 변환 시도
        return int(float(value))
    except (ValueError, TypeError):
        # 숫자로 변환할 수 없는 경우 원래 값 반환
        return value

def extract_tsv_from_response(response_text):
    if not response_text: return ""
    lines = response_text.strip().split('\n')
    cleaned_lines = [line.strip() for line in lines if '\t' in line.strip()]
    return "\n".join(cleaned_lines)

def fix_tsv_field_count(tsv_str):
    lines = tsv_str.strip().split('\n')
    if not lines: 
        return tsv_str
    
    header = lines[0]
    n_fields = header.count('\t') + 1
    fixed_lines = [header]
    
    for line in lines[1:]:
        fields = line.split('\t')
        if len(fields) < n_fields:
            fields += [''] * (n_fields - len(fields))
        elif len(fields) > n_fields:
            fields = fields[:n_fields-1] + [' '.join(fields[n_fields-1:])]
        fixed_lines.append('\t'.join(fields))
    
    return '\n'.join(fixed_lines)

def create_excel_report(**kwargs):
    """새로운 엑셀 보고서를 처음부터 생성합니다 - 1페이지 최적화 버전"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    
    wb = Workbook()
    ws = wb.active
    ws.title = "공사일보"
    
    # 페이지 설정 - A4 세로, 1페이지 최적화
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9  # A4 용지 크기
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    
    # 스타일 정의
    title_font = Font(bold=True, size=16, name='맑은 고딕')
    header_font = Font(bold=True, size=11, name='맑은 고딕')
    normal_font = Font(size=9, name='맑은 고딕')
    small_font = Font(size=8, name='맑은 고딕')
    
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    # 기본 정보
    work_date = kwargs.get("work_date", datetime.now().strftime('%Y-%m-%d'))
    project_name = kwargs.get("project_name", "터널 건설공사")
    section_name = kwargs.get("section_name", "대림-신풍-도림")
    
    # 1. 제목 및 헤더 (1-3행)
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value=f"{project_name} 일일작업보고서")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = header_fill
    title_cell.border = thick_border
    
    # 기본 정보 (2행)
    ws.cell(row=2, column=1, value="구간").font = header_font
    ws.cell(row=2, column=2, value=section_name).font = normal_font
    ws.cell(row=2, column=6, value="보고일").font = header_font
    ws.cell(row=2, column=7, value=work_date).font = normal_font
    
    for col in range(1, 11):
        ws.cell(row=2, column=col).border = thin_border
        if col in [1, 6]:
            ws.cell(row=2, column=col).fill = sub_header_fill
    
    current_row = 4
    
    # 2. 날씨 및 기본 현황을 한 줄에 (4-5행)
    tables_data = kwargs.get("tables_data", [])
    
    # 날씨정보 (왼쪽)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value="날씨 현황").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    
    # 주요 시공 현황 (왼쪽)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value="주요 시공 현황").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    
    if tables_data and len(tables_data) > 0 and tables_data[0] is not None:
        construction_df = tables_data[0]
        # 상위 5개 항목만 표시
        main_items = construction_df.head(5) if not construction_df.empty else pd.DataFrame()
        construction_text = ""
        for _, row in main_items.iterrows():
            if str(row['누계']) not in ['', '0', '0.0']:
                construction_text += f"{row['구분']}: {row['누계']}, "
        construction_text = construction_text.rstrip(", ")
        
        ws.merge_cells(f'F{current_row+1}:J{current_row+1}')
        ws.cell(row=current_row+1, column=6, value=construction_text).font = small_font
        ws.cell(row=current_row+1, column=6).border = thin_border
    
    current_row += 3
    
    # 3. 금일 작업 내용 (주요 섹션)
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws.cell(row=current_row, column=1, value="금일 주요 작업 내용").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    current_row += 1
    
    if tables_data and len(tables_data) > 1 and tables_data[1] is not None:
        work_df = tables_data[1]
        # 실제 작업이 있는 항목만 표시
        active_work = work_df[work_df['금일작업'].str.strip() != ''] if not work_df.empty else pd.DataFrame()
        
        for idx, (_, row) in enumerate(active_work.head(6).iterrows()):  # 최대 6개 항목
            ws.cell(row=current_row, column=1, value=f"• {row['구분']}").font = small_font
            ws.merge_cells(f'B{current_row}:J{current_row}')
            ws.cell(row=current_row, column=2, value=row['금일작업']).font = small_font
            ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
            
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
    else:
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws.cell(row=current_row, column=1, value="작업 내용이 없습니다.").font = small_font
        current_row += 1
    
    current_row += 1
    
    # 4. 인원 및 장비 현황 (요약)
    col1_start = current_row
    
    # 인원 현황 (왼쪽)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1, value="인원 현황").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    current_row += 1
    
    if tables_data and len(tables_data) > 2 and tables_data[2] is not None:
        personnel_df = tables_data[2]
        if not personnel_df.empty:
            # 총 인원 계산
            total_personnel = 0
            for col in personnel_df.columns[1:]:  # 첫 번째 열(구분) 제외
                personnel_df[col] = pd.to_numeric(personnel_df[col], errors='coerce').fillna(0)
                total_personnel += personnel_df[col].sum()
            
            ws.cell(row=current_row, column=1, value="총 투입인원").font = small_font
            ws.cell(row=current_row, column=2, value=f"{int(total_personnel)}명").font = small_font
            
            # 주요 직종별 인원 (0이 아닌 것만)
            for _, row in personnel_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in personnel_df.columns[1:])
                if row_total > 0:
                    ws.cell(row=current_row+1, column=1, value=row['구분']).font = small_font
                    ws.cell(row=current_row+1, column=2, value=f"{int(row_total)}명").font = small_font
                    current_row += 1
                    if current_row - col1_start > 8:  # 최대 8줄
                        break
    
    # 장비 현황 (오른쪽)
    current_row = col1_start
    ws.merge_cells(f'F{current_row}:J{current_row}')
    ws.cell(row=current_row, column=6, value="장비 현황").font = header_font
    ws.cell(row=current_row, column=6).fill = sub_header_fill
    ws.cell(row=current_row, column=6).border = thin_border
    current_row += 1
    
    if tables_data and len(tables_data) > 3 and tables_data[3] is not None:
        equipment_df = tables_data[3]
        if not equipment_df.empty:
            # 주요 장비 (0이 아닌 것만)
            for _, row in equipment_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in equipment_df.columns[1:])
                if row_total > 0:
                    ws.cell(row=current_row, column=6, value=row['구분']).font = small_font
                    ws.cell(row=current_row, column=7, value=f"{int(row_total)}대").font = small_font
                    current_row += 1
                    if current_row - col1_start > 8:  # 최대 8줄
                        break
    
    # 5. 발파 및 계측기 데이터 (하단 요약)
    current_row = max(current_row, col1_start + 10) + 1
    
    blast_df = kwargs.get("blast_df")
    instrument_df = kwargs.get("instrument_df")
    
    if blast_df is not None and not blast_df.empty:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value="발파 현황").font = header_font
        ws.cell(row=current_row, column=1).fill = sub_header_fill
        ws.cell(row=current_row, column=1).border = thin_border
        
        blast_count = len(blast_df)
        total_explosive = blast_df['폭약사용량(kg)'].sum() if '폭약사용량(kg)' in blast_df.columns else 0
        
        ws.merge_cells(f'A{current_row+1}:E{current_row+1}')
        ws.cell(row=current_row+1, column=1, value=f"발파횟수: {blast_count}회, 폭약사용량: {total_explosive}kg").font = small_font
        ws.cell(row=current_row+1, column=1).border = thin_border
        
    if instrument_df is not None and not instrument_df.empty:
        ws.merge_cells(f'F{current_row}:J{current_row}')
        ws.cell(row=current_row, column=6, value="계측기 현황").font = header_font
        ws.cell(row=current_row, column=6).fill = sub_header_fill
        ws.cell(row=current_row, column=6).border = thin_border
        
        warning_count = len(instrument_df[instrument_df['상태'] != '안정']) if '상태' in instrument_df.columns else 0
        total_count = len(instrument_df)
        
        ws.merge_cells(f'F{current_row+1}:J{current_row+1}')
        ws.cell(row=current_row+1, column=6, value=f"총 {total_count}개소, 경고: {warning_count}개소").font = small_font
        ws.cell(row=current_row+1, column=6).border = thin_border
    
    # 모든 셀에 테두리 적용 및 열 너비 설정
    for row in ws.iter_rows(min_row=1, max_row=current_row+2, min_col=1, max_col=10):
        for cell in row:
            if not cell.border.left.style:
                cell.border = thin_border
    
    # 열 너비 최적화 (A4 1페이지에 맞게)
    column_widths = [12, 12, 10, 10, 10, 12, 12, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 행 높이 조정
    for row_num in range(1, current_row + 3):
        ws.row_dimensions[row_num].height = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_pdf_report(**kwargs):
    """새로운 PDF 보고서를 생성합니다 - 완전한 내용 표시 버전"""
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    
    # 한글 폰트 등록 시도
    try:
        # 시스템에서 한글 폰트 찾기
        font_paths = [
            "C:/Windows/Fonts/malgun.ttf",  # 맑은 고딕
            "C:/Windows/Fonts/gulim.ttc",   # 굴림
            "C:/Windows/Fonts/batang.ttc",  # 바탕
        ]
        
        korean_font = None
        for font_path in font_paths:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Korean', font_path))
                korean_font = 'Korean'
                break
        
        if not korean_font:
            korean_font = 'Helvetica'  # 폴백 폰트
    except:
        korean_font = 'Helvetica'
    
    # PDF 생성
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # 스타일 정의
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=korean_font,
        fontSize=16,
        spaceAfter=12,
        alignment=1,  # 중앙 정렬
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontName=korean_font,
        fontSize=12,
        spaceAfter=6,
        textColor=colors.darkblue
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=korean_font,
        fontSize=9,
        spaceAfter=3
    )
    
    # 내용 구성
    story = []
    
    # 기본 정보
    work_date = kwargs.get("work_date", datetime.now().strftime('%Y-%m-%d'))
    project_name = kwargs.get("project_name", "터널 건설공사")
    section_name = kwargs.get("section_name", "대림-신풍-도림")
    tables_data = kwargs.get("tables_data", [])
    
    # 1. 제목
    story.append(Paragraph(f"{project_name} 일일작업보고서", title_style))
    story.append(Spacer(1, 6*mm))
    
    # 2. 기본 정보 테이블
    basic_info = [
        ['구간', section_name, '보고일', work_date]
    ]
    basic_table = Table(basic_info, colWidths=[30*mm, 60*mm, 30*mm, 60*mm])
    basic_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.lightblue),
        ('BACKGROUND', (2, 0), (2, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), korean_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(basic_table)
    story.append(Spacer(1, 6*mm))
    
    # 3. 시공현황
    if tables_data and len(tables_data) > 0 and tables_data[0] is not None:
        story.append(Paragraph("시공현황", header_style))
        construction_df = tables_data[0]
        if not construction_df.empty:
            construction_data = [['구분', '누계']]
            for _, row in construction_df.iterrows():
                if str(row['누계']) not in ['', '0', '0.0']:  # 값이 있는 것만
                    construction_data.append([str(row['구분']), str(row['누계'])])
            
            if len(construction_data) > 1:  # 헤더 외에 데이터가 있으면
                construction_table = Table(construction_data, colWidths=[120*mm, 40*mm])
                construction_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 0), (-1, -1), korean_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # 첫 번째 열 왼쪽 정렬
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # 두 번째 열 중앙 정렬
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(construction_table)
        story.append(Spacer(1, 6*mm))
    
    # 4. 주요 작업 내용
    if tables_data and len(tables_data) > 1 and tables_data[1] is not None:
        story.append(Paragraph("금일 주요 작업 내용", header_style))
        work_df = tables_data[1]
        active_work = work_df[work_df['금일작업'].str.strip() != ''] if not work_df.empty else pd.DataFrame()
        
        if not active_work.empty:
            work_data = [['구분', '작업내용']]
            for _, row in active_work.iterrows():  # 모든 작업 내용 포함
                work_data.append([str(row['구분']), str(row['금일작업'])])
            
            work_table = Table(work_data, colWidths=[60*mm, 110*mm])
            work_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), korean_font),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # 첫 번째 열 중앙 정렬
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),    # 두 번째 열 왼쪽 정렬
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(work_table)
        else:
            story.append(Paragraph("금일 작업 내용이 없습니다.", normal_style))
        story.append(Spacer(1, 6*mm))
    
    # 5. 인원 현황 요약
    if tables_data and len(tables_data) > 2 and tables_data[2] is not None:
        story.append(Paragraph("인원 현황", header_style))
        personnel_df = tables_data[2]
        if not personnel_df.empty:
            # 총 인원 계산
            total_personnel = 0
            personnel_summary = []
            
            for _, row in personnel_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in personnel_df.columns[1:])
                if row_total > 0:
                    personnel_summary.append([str(row['구분']), f"{int(row_total)}명"])
                    total_personnel += row_total
            
            # 총 인원을 맨 위에 추가
            personnel_data = [['구분', '인원수'], ['총 투입인원', f"{int(total_personnel)}명"]] + personnel_summary  # 모든 인원 포함
            
            personnel_table = Table(personnel_data, colWidths=[80*mm, 40*mm])
            personnel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('BACKGROUND', (0, 1), (-1, 1), colors.lightyellow),  # 총 인원 강조
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), korean_font),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(personnel_table)
        story.append(Spacer(1, 6*mm))
    
    # 6. 장비 현황 요약
    if tables_data and len(tables_data) > 3 and tables_data[3] is not None:
        story.append(Paragraph("장비 현황", header_style))
        equipment_df = tables_data[3]
        if not equipment_df.empty:
            equipment_summary = []
            
            for _, row in equipment_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in equipment_df.columns[1:])
                if row_total > 0:
                    equipment_summary.append([str(row['구분']), f"{int(row_total)}대"])
            
            if equipment_summary:
                equipment_data = [['장비명', '대수']] + equipment_summary  # 모든 장비 포함
                
                equipment_table = Table(equipment_data, colWidths=[80*mm, 40*mm])
                equipment_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 0), (-1, -1), korean_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                story.append(equipment_table)
        story.append(Spacer(1, 6*mm))
    
    # 8. 발파 데이터 (상세)
    blast_df = kwargs.get("blast_df")
    if blast_df is not None and not blast_df.empty:
        story.append(Paragraph("발파 현황", header_style))
        
        # 발파 데이터 테이블
        blast_columns = list(blast_df.columns)
        blast_data = [blast_columns]  # 헤더
        
        for _, row in blast_df.iterrows():
            blast_data.append([str(row[col]) for col in blast_columns])
        
        # 컬럼 수에 따라 동적으로 너비 조정
        col_count = len(blast_columns)
        col_width = 170 // col_count * mm  # 전체 너비를 컬럼 수로 나눔
        col_widths = [col_width] * col_count
        
        blast_table = Table(blast_data, colWidths=col_widths)
        blast_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), korean_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(blast_table)
        story.append(Spacer(1, 6*mm))
    
    # 9. 계측기 데이터 (상세)
    instrument_df = kwargs.get("instrument_df")
    if instrument_df is not None and not instrument_df.empty:
        story.append(Paragraph("계측기 현황", header_style))
        
        # 계측기 데이터 테이블
        instrument_columns = list(instrument_df.columns)
        instrument_data = [instrument_columns]  # 헤더
        
        for _, row in instrument_df.iterrows():
            instrument_data.append([str(row[col]) for col in instrument_columns])
        
        # 컬럼 수에 따라 동적으로 너비 조정
        col_count = len(instrument_columns)
        col_width = 170 // col_count * mm  # 전체 너비를 컬럼 수로 나눔
        col_widths = [col_width] * col_count
        
        instrument_table = Table(instrument_data, colWidths=col_widths)
        instrument_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), korean_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(instrument_table)
    
    # PDF 생성
    try:
        doc.build(story)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        # 한글 폰트 문제 시 폴백
        for element in story:
            if hasattr(element, 'style') and hasattr(element.style, 'fontName'):
                element.style.fontName = 'Helvetica'
        
        doc.build(story)
        output.seek(0)
        return output.getvalue()


# 이 함수는 더 이상 사용되지 않으므로 제거
# insert_data_to_excel_with_mapping 함수를 사용하세요


def insert_data_to_excel_with_mapping(template_bytes, basic_info, tables_data, cell_mapping=None, table_mapping=None, previous_data=None):
    """엑셀 템플릿에 기본정보를 특정 셀에 매핑하여 삽입합니다."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import io
        
        # 엑셀 파일 로드
        workbook = load_workbook(io.BytesIO(template_bytes))
        worksheet = workbook.active
        
        # 기본정보 매핑 (사용자 설정 또는 기본값 사용)
        if cell_mapping is None:
            # 기본 셀 매핑 설정
            cell_mapping = {
                'date': 'u2',
                'project_name': 'd4', 
                'max_temp': 'o4',
                'min_temp': 'o5',
                'precipitation': 'o6',
                'planned_progress': 'w4',
                'actual_progress': 'w5'
            }
        
        # 기본정보 삽입
        for key, cell_address in cell_mapping.items():
            if key in basic_info:
                try:
                    # 안전한 셀 값 설정
                    try:
                        # 셀 주소를 파싱하여 직접 접근
                        row_idx, col_idx = parse_cell_address(cell_address)
                        
                        # 셀에 직접 접근
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        # 병합된 셀인지 확인
                        is_merged = False
                        target_cell = cell_address
                        
                        for merged_range in worksheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                # 병합된 셀의 첫 번째 셀 주소 계산
                                target_col = chr(ord('A') + merged_range.min_col - 1)
                                target_row = merged_range.min_row
                                target_cell = f"{target_col}{target_row}"
                                break
                        
                        # 데이터 삽입 (텍스트는 그대로, 숫자는 변환)
                        if key in ['project_name', 'date']:
                            # 텍스트 데이터는 그대로 사용
                            cell_value = basic_info[key]
                        else:
                            # 숫자 데이터는 변환
                            cell_value = convert_to_number_if_possible(basic_info[key])
                        
                        if is_merged:
                            # 병합된 셀의 첫 번째 셀에 데이터 삽입
                            worksheet[target_cell] = cell_value
                        else:
                            # 일반 셀에 데이터 삽입
                            worksheet[cell_address] = cell_value
                            
                    except Exception as cell_error:
                        st.warning(f"⚠️ 셀 {cell_address} 처리 중 오류: {cell_error}")
                        continue
                        
                except Exception as e:
                    st.warning(f"⚠️ 셀 {cell_address} 처리 중 오류: {e}")
                    # 오류 발생 시 해당 셀 건너뛰기
                    continue
        
        # 테이블 추출값 삽입 (table_mapping이 있는 경우)
        if table_mapping:
            for key, cell_address in table_mapping.items():
                try:
                    # 병합된 셀 안전하게 처리
                    cell = worksheet[cell_address]
                    
                    # 병합된 셀인지 확인
                    is_merged = False
                    target_cell = cell_address
                    
                    # 병합된 셀 범위에서 해당 셀 찾기
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell_address in merged_range:
                            is_merged = True
                            target_cell = merged_range.start_cell.coordinate
                            break
                    
                    # 테이블 데이터에서 해당 값 추출하여 삽입
                    table_value = ""
                    if key in tables_data:
                        df = tables_data[key]
                        if df is not None and not df.empty:
                            table_value = str(df.iloc[0, 1]) if len(df.columns) > 1 else str(df.iloc[0, 0])
                    
                    # 데이터 삽입 (숫자 타입 강제 설정)
                    cell_value = convert_to_number_if_possible(table_value)
                    if is_merged:
                        # 병합된 셀의 첫 번째 셀에 데이터 삽입
                        worksheet[target_cell] = cell_value
                    else:
                        # 일반 셀에 데이터 삽입
                        worksheet[cell_address] = cell_value
                        
                except Exception as e:
                    st.warning(f"⚠️ 테이블 추출값 셀 {cell_address} 처리 중 오류: {e}")
                    continue
        
        # 전일 데이터 적용 (previous_data가 있는 경우)
        if previous_data:
            try:
                # 1. 시공현황 전일 데이터 적용 (T11~43 누계 → N11~43 전일까지)
                construction_data = previous_data.get("construction_data", {})
                row = 11
                for category, data in construction_data.items():
                    if row <= 43:
                        cumulative_value = data.get("누계", 0)
                        worksheet[f"N{row}"] = cumulative_value
                        row += 1
                
                # 2. 인원 전일 데이터 적용 (L66~87, Y66~87)
                personnel_data = previous_data.get("personnel_data", {})
                row = 66
                for category, data in personnel_data.items():
                    if row <= 87:
                        previous_value = data.get("전일까지", 0)
                        cumulative_value = data.get("누계", 0)
                        worksheet[f"L{row}"] = previous_value
                        worksheet[f"Y{row}"] = cumulative_value
                        row += 1
                
                # 3. 장비 전일 데이터 적용 (L91~119, Y91~119)
                equipment_data = previous_data.get("equipment_data", {})
                row = 91
                for category, data in equipment_data.items():
                    if row <= 119:
                        previous_value = data.get("전일까지", 0)
                        cumulative_value = data.get("누계", 0)
                        worksheet[f"L{row}"] = previous_value
                        worksheet[f"Y{row}"] = cumulative_value
                        row += 1
                
                st.success("✅ 전일 데이터가 자동으로 적용되었습니다!")
                
            except Exception as e:
                st.warning(f"⚠️ 전일 데이터 적용 중 오류: {e}")
        else:
            st.info("ℹ️ 전일 데이터가 없어 새로 시작합니다.")
        
        # 테이블 데이터 삽입 (전체 테이블 삽입)
        if tables_data:
            # 사용자 입력값을 기반으로 테이블 위치 계산
            table_positions = {}
            if table_mapping:
                for table_name, cell_address in table_mapping.items():
                    if cell_address:
                        # 셀 주소를 행/열로 변환
                        from openpyxl.utils import column_index_from_string
                        try:
                            # 셀 주소에서 열과 행 분리 (예: "A1" -> "A", "1")
                            import re
                            match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
                            if match:
                                col_str = match.group(1)
                                row_str = match.group(2)
                                row = int(row_str)
                                col = column_index_from_string(col_str)
                                table_positions[table_name] = {"row": row, "col": col}
                            else:
                                raise ValueError(f"잘못된 셀 주소 형식: {cell_address}")
                        except Exception as e:
                            # 기본값 사용
                            default_positions = {
                                "시공현황": {"row": 8, "col": 17},   # q8
                                "작업내용": {"row": 11, "col": 17},  # q11
                                "인원": {"row": 66, "col": 29},      # ac66
                                "장비": {"row": 106, "col": 29}      # ac106
                            }
                            if table_name in default_positions:
                                table_positions[table_name] = default_positions[table_name]
            else:
                # 기본 위치 사용
                table_positions = {
                    "시공현황": {"row": 8, "col": 17},   # q8
                    "작업내용": {"row": 11, "col": 17},  # q11
                    "인원": {"row": 66, "col": 29},      # ac66
                    "장비": {"row": 106, "col": 29}      # ac106
                }
            
            # 모든 테이블 삽입 (숫자 타입 강제 설정)
            for table_name, df in tables_data.items():
                if table_name in table_positions and df is not None and not df.empty:
                    pos = table_positions[table_name]
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                        for c_idx, value in enumerate(row):
                            # 숫자 데이터 타입 강제 설정
                            cell_value = convert_to_number_if_possible(value)
                            worksheet.cell(row=pos["row"] + r_idx, column=pos["col"] + c_idx, value=cell_value)
        
        # 수정된 엑셀 파일을 바이트로 변환
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"엑셀 데이터 매핑 삽입 중 오류: {e}")
        return None


# --- STATE INITIALIZATION ---
def initialize_session_state():
    states = {
        "kakao_work_completed": False, "basic_info_completed": False, "excel_export_completed": False,
        "processed_tables": [], "kakao_results": None,
        "final_excel_data": None, "processed_template_filename": None,
        "all_accumulated_rows": [], "reset_flag": 0,
        "prompt": DEFAULT_PROMPT, "work_date": None,
        "warning_rows_instrument": None,
        # 프롬프트 관리 관련 상태
        "current_prompt_name": "기본 프롬프트",
        "show_prompt_editor": False,
        "show_table_editor": False,
        "prompt_list": [],
        # Supabase 저장 관련 상태
        "daily_report_saved": False,
        "save_success_message": "",
        "save_error_message": "",
        "save_success_date": "",
        # 테이블 구조 관리
        "construction_rows": [
            "1. 본선터널 (1구간, 대림-신풍)",
            "1. 본선터널 (1구간, 대림-신풍) 라이닝",
            "2. 신풍정거장 - 1)정거장 라이닝",
            "2. 신풍정거장 - 1)정거장 미들 슬라브",
            "2. 신풍정거장 – 2)주출입구 수직구 라이닝",
            "2. 신풍정거장 - 2)주출입구 - (1)PCB 정거장 방면 라이닝",
            "2. 신풍정거장 - 2)주출입구 - (1)PCB 환승통로 방면 라이닝",
            "2. 신풍정거장 - 2)주출입구 - (2)PCC 라이닝",
            "2. 신풍정거장 - 2)주출입구 - (3)PCD 라이닝",
            "2. 신풍정거장 - 2)주출입구 - (4)PHA 라이닝",
            "2. 신풍정거장 - 3)특별피난계단 - 수직구 라이닝",
            "2. 신풍정거장 - 3)특별피난계단 - PHB 라이닝",
            "2. 신풍정거장 - 4)외부출입구 출입구(#3) 굴착",
            "2. 신풍정거장 - 4)외부출입구 출입구(#2) 굴착",
            "2. 신풍정거장 - 4)외부출입구 출입구(#1) 굴착",
            "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 굴착",
            "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 라이닝",
            "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 굴착",
            "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 라이닝",
            "3. 신풍 환승통로 - 2)개착 BOX 보라매 방면 구조물",
            "3. 신풍 환승통로 - 2)개착 BOX 대림 방면 굴착",
            "4. 본선터널(2구간, 신풍-도림) 굴착",
            "4. 본선터널(2구간, 신풍-도림) 라이닝",
            "5. 도림사거리정거장 - 1)정거장 터널 라이닝",
            "5. 도림사거리정거장 - 1)정거장 미들 슬라브",
            "5. 도림사거리정거장 - 2)출입구#1 수직구 라이닝",
            "5. 도림사거리정거장 - 2)출입구#1 PCA 라이닝",
            "5. 도림사거리정거장 - 2)출입구#1 PCC 라이닝",
            "5. 도림사거리정거장 - 2)출입구#1 PHA 라이닝",
            "5. 도림사거리정거장 - 3)출입구#2 수직구 라이닝",
            "5. 도림사거리정거장 - 3)출입구#2 PCA 라이닝",
            "5. 도림사거리정거장 - 3)출입구#2 PCC 라이닝",
            "5. 도림사거리정거장 - 3)출입구#2 PHB 라이닝"
        ],
        "work_content_rows": [
            "1. 본선터널 (1구간, 대림-신풍)",
            "2.신풍정거장 - 1)정거장 터널",
            "2.신풍정거장 - 2)주출입구 - (1)PCB",
            "2.신풍정거장 - 2)주출입구 - (2)PCC",
            "2.신풍정거장 - 2)주출입구 - (3)PCD",
            "2.신풍정거장 - 2)주출입구 - (4)PHA",
            "2.신풍정거장 - 3)특별피난계단",
            "2.신풍정거장 - 4)외부출입구",
            "3.신풍 환승통로 - 1)환승터널",
            "3.신풍 환승통로 - 2)개착 BOX",
            "4.본선터널(2구간, 신풍-도림)",
            "5.도림사거리정거장 - 1)정거장 터널",
            "5.도림사거리정거장 - 2)출입구#1",
            "5.도림사거리정거장 - 3)출입구#2"
        ],
        "personnel_columns": [
            "1. 본선터널 (1구간, 대림~신풍)",
            "2.신풍정거장 - 1)정거장 터널",
            "2.신풍정거장 - 2)주출입구 - (1)PCB",
            "2.신풍정거장 - 2)주출입구 - (2)PCC",
            "2.신풍정거장 - 2)주출입구 - (3)PCD",
            "2.신풍정거장 - 2)주출입구 - (4)PHA",
            "2.신풍정거장 - 3)특별피난계단",
            "2.신풍정거장 - 4)외부출입구",
            "3.신풍 환승통로 - 1)환승터널",
            "3.신풍 환승통로 - 2)개착 BOX",
            "4.본선터널(2구간, 신풍~도림)",
            "5.도림사거리정거장 - 1)정거장 터널",
            "5.도림사거리정거장 - 2)출입구#1",
            "5.도림사거리정거장 - 3)출입구#2"
        ],
        "personnel_rows": [
            "직영반장", "연수생", "장비운전원", "전기주임", "화약주임", "터널공", "목공", "철근공", 
            "라이닝폼공", "오폐수처리공", "카리프트공", "BP공", "가시설공/해체공", "동바리공", 
            "신호수", "부단수공", "슬러리월공", "CIP공", "미장공", "시설물공", "경계석공", "조경공", 
            "배관공", "도색공", "방수공", "장비/작업지킴이", "보통인부", "포장공", "용접공", "타설공", 
            "보링공/앙카공", "비계공", "도장공", "석면공", "주입공/그라우팅공"
        ],
        "equipment_rows": [
            "B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)",
            "덤프트럭(5T)", "덤프트럭(15T)", "덤프트럭(25T)", "앵글크레인(100T)", "앵글크레인(80T)", "앵글크레인(35T)", "앵글크레인(25T)",
            "카고크레인(25T)", "카고크레인(5T)", "콤프", "점보드릴", "페이로더", "숏트머신", "차징카", "살수차", "하이드로크레인",
            "믹서트럭", "화물차(5T)", "펌프카", "스카이", "콘크리트피니셔", "전주오거", "로더(바브켓)", "유제살포기(비우다)",
            "지게차", "싸인카", "BC커터기", "바이브로해머", "롤러(2.5T)", "롤러(1T)", "롤러(0.7T)", "몰리", "항타기", 
            "크레인", "콤비로라", "공압드릴", "유압드릴", "기타"
        ]
    }
    for key, value in states.items():
        if key not in st.session_state:
            st.session_state[key] = value
    
    # 앱 시작 시 저장된 프롬프트 목록 로드
    if supabase_client and not st.session_state.prompt_list:
        st.session_state.prompt_list = get_all_prompts_from_supabase()

initialize_session_state()


# --- SIDEBAR ---
with st.sidebar:
    st.markdown("<h1>🏗️ 공사일보 자동화</h1>", unsafe_allow_html=True)
    
    if st.button("모든 작업 초기화", key="reset_all"):
        for key in list(st.session_state.keys()): del st.session_state[key]
        st.rerun()

    st.markdown('<div class="glass-card">', unsafe_allow_html=True)

    steps = [
        ("SNS일작업보고 입력", "kakao_work_completed", "📝"),
        ("기본정보 입력", "basic_info_completed", "📋"),
        ("작업일보 생성", "excel_export_completed", "📊")
    ]
    
    current_step_index = 0
    if st.session_state.basic_info_completed: current_step_index = 1
    if st.session_state.excel_export_completed: current_step_index = 2

    for i, (text, state_key, icon) in enumerate(steps):
        is_completed = st.session_state.get(state_key, False)
        is_active = (i == current_step_index)
        status_class = "completed" if is_completed else "active" if is_active else ""
        step_icon = "✅" if is_completed else icon
        
        st.markdown(f'<div class="step {status_class}"><div class="step-icon">{step_icon}</div><div class="step-text"><strong>{i+1}.</strong> {text}</div></div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# --- MAIN CONTENT ---
st.markdown("""
<div style="text-align: center; padding: 20px 0;">
    <h1 style="margin: 0; font-size: 2.5rem; font-weight: bold; color: #1E3A8A;">🏗️ 공사일보 자동화 시스템</h1>
</div>
""", unsafe_allow_html=True)

# --- STEP 1: SNS WORK REPORT INPUT ---
with st.container():
    st.markdown('<h3><span style="font-size: 1.5em;">📝</span> 1. SNS일작업보고 입력</h3>', unsafe_allow_html=True)
    
    # 프롬프트 관리 섹션
    with st.expander("⚙️ 프롬프트 관리", expanded=False):
        # 1. 저장된 프롬프트 선택 + 편집
        st.markdown("**1. 저장된 프롬프트 선택**")
        col1, col2 = st.columns([4, 1])
        
        with col1:
            if st.session_state.prompt_list:
                prompt_options = ["기본 프롬프트"] + [p["name"] for p in st.session_state.prompt_list]
                selected_prompt = st.selectbox(
                    "저장된 프롬프트 선택",
                    options=prompt_options,
                    index=prompt_options.index(st.session_state.current_prompt_name) if st.session_state.current_prompt_name in prompt_options else 0
                )
                
                if selected_prompt != st.session_state.current_prompt_name:
                    st.session_state.current_prompt_name = selected_prompt
                    if selected_prompt == "기본 프롬프트":
                        st.session_state.prompt = DEFAULT_PROMPT
                    else:
                        loaded_prompt = load_prompt_from_supabase(selected_prompt)
                        if loaded_prompt:
                            st.session_state.prompt = loaded_prompt["content"]
                    st.rerun()
        
        with col2:
            if st.button("✏️ 편집", key="edit_prompt", use_container_width=True):
                st.session_state.show_prompt_editor = True
                st.session_state.show_table_editor = True
                st.rerun()
        
        # 2. 프롬프트 편집
        if st.session_state.show_prompt_editor:
            st.markdown("---")
            st.markdown("**2. 프롬프트 편집**")
            
            # 프롬프트 이름과 설명
            prompt_name = st.text_input(
                "이름",
                value=st.session_state.current_prompt_name if st.session_state.current_prompt_name != "기본 프롬프트" else "",
                placeholder="새 프롬프트 이름을 입력하세요"
            )
            
            prompt_description = st.text_input(
                "설명 (선택사항)",
                placeholder="프롬프트에 대한 설명을 입력하세요"
            )
            
            # 텍스트와 테이블을 2분할로 배치
            text_col, table_col = st.columns(2)
            
            with text_col:
                # 프롬프트 내용 (텍스트) - 실시간 업데이트
                edited_prompt = st.text_area(
                    "내용 (텍스트)",
                    value=st.session_state.prompt,
                    height=600,
                    help="AI가 데이터 분석에 사용할 지시문을 작성하세요. 테이블 구조를 변경하면 자동으로 업데이트됩니다.",
                    key="prompt_text_area"
                )
                
                # 디버깅: 현재 프롬프트 상태 표시
                if st.checkbox("🔍 프롬프트 상태 확인", key="debug_prompt"):
                    st.info(f"현재 프롬프트 길이: {len(st.session_state.prompt)} 문자")
                    st.code(st.session_state.prompt[:200] + "..." if len(st.session_state.prompt) > 200 else st.session_state.prompt)
                
                # 텍스트에서 테이블 구조 추출 버튼
                if st.button("🔄 텍스트에서 테이블 구조 추출", key="extract_from_text"):
                    try:
                        # 텍스트에서 테이블 구조를 추출하는 로직
                        extracted_tables = extract_table_structure_from_prompt(edited_prompt)
                        if extracted_tables:
                            # 추출된 구조로 세션 상태 업데이트
                            if 'construction' in extracted_tables:
                                st.session_state.construction_rows = extracted_tables['construction']
                            if 'work_content' in extracted_tables:
                                st.session_state.work_content_rows = extracted_tables['work_content']
                            if 'personnel_columns' in extracted_tables:
                                st.session_state.personnel_columns = extracted_tables['personnel_columns']
                            if 'personnel_rows' in extracted_tables:
                                st.session_state.personnel_rows = extracted_tables['personnel_rows']
                            if 'equipment' in extracted_tables:
                                st.session_state.equipment_rows = extracted_tables['equipment']
                            
                            st.success("✅ 텍스트에서 테이블 구조를 추출했습니다!")
                            st.rerun()
                        else:
                            st.warning("⚠️ 텍스트에서 테이블 구조를 추출할 수 없습니다.")
                    except Exception as e:
                        st.error(f"❌ 테이블 구조 추출 중 오류: {e}")
            
            with table_col:
                # 테이블 구조 편집
                st.markdown("**테이블 구조 편집**")
                
                # 탭으로 각 테이블 구분
                tab1, tab2, tab3, tab4 = st.tabs(["🏗️ 시공현황", "📝 작업내용", "👥 인원", "🚛 장비"])
                
                with tab1:
                    # 시공현황 행 편집
                    construction_df = pd.DataFrame({"구분": st.session_state.construction_rows})
                    edited_construction = st.data_editor(
                        construction_df,
                        use_container_width=True,
                        num_rows="dynamic",
                        height=600,
                        column_config={
                            "구분": st.column_config.TextColumn("구분", help="시공현황 항목명")
                        },
                        key="construction_editor"
                    )
                    
                    # 시공현황이 변경되면 자동으로 프롬프트 업데이트
                    new_construction_rows = edited_construction["구분"].tolist()
                    if new_construction_rows != st.session_state.construction_rows:
                        st.session_state.construction_rows = new_construction_rows
                        # 자동 프롬프트 업데이트
                        st.session_state.prompt = generate_prompt_from_tables()
                        st.info("✅ 시공현황 테이블 변경으로 프롬프트가 자동 업데이트되었습니다.")
                        st.rerun()
                
                with tab2:
                    # 작업내용 행 편집
                    work_content_df = pd.DataFrame({"구분": st.session_state.work_content_rows})
                    edited_work_content = st.data_editor(
                        work_content_df,
                        use_container_width=True,
                        num_rows="dynamic",
                        height=600,
                        column_config={
                            "구분": st.column_config.TextColumn("구분", help="작업내용 항목명")
                        },
                        key="work_content_editor"
                    )
                    
                    # 작업내용이 변경되면 자동으로 프롬프트 업데이트
                    new_work_content_rows = edited_work_content["구분"].tolist()
                    if new_work_content_rows != st.session_state.work_content_rows:
                        st.session_state.work_content_rows = new_work_content_rows
                        # 자동 프롬프트 업데이트
                        st.session_state.prompt = generate_prompt_from_tables()
                        st.info("✅ 작업내용 테이블 변경으로 프롬프트가 자동 업데이트되었습니다.")
                        st.rerun()
                
                with tab3:
                    # 2개 컬럼으로 나누어 표시
                    col_left, col_right = st.columns(2)
                    
                    with col_left:
                        st.markdown("##### 📍 작업 위치 (열)")
                        personnel_columns_df = pd.DataFrame({"작업위치": st.session_state.personnel_columns})
                        edited_personnel_columns = st.data_editor(
                            personnel_columns_df,
                            use_container_width=True,
                            num_rows="dynamic",
                            height=600,
                            column_config={
                                "작업위치": st.column_config.TextColumn("작업위치", help="인원 테이블의 열 항목")
                            },
                            key="personnel_columns_editor"
                        )
                        # 인원 열이 변경되면 자동으로 프롬프트 업데이트
                        new_personnel_columns = edited_personnel_columns["작업위치"].tolist()
                        if new_personnel_columns != st.session_state.personnel_columns:
                            st.session_state.personnel_columns = new_personnel_columns
                            # 자동 프롬프트 업데이트
                            st.session_state.prompt = generate_prompt_from_tables()
                            st.info("✅ 인원 테이블 열 변경으로 프롬프트가 자동 업데이트되었습니다.")
                            st.rerun()
                    
                    with col_right:
                        st.markdown("##### 👥 직종 (행)")
                        personnel_rows_df = pd.DataFrame({"직종": st.session_state.personnel_rows})
                        edited_personnel_rows = st.data_editor(
                            personnel_rows_df,
                            use_container_width=True,
                            num_rows="dynamic",
                            height=600,
                            column_config={
                                "직종": st.column_config.TextColumn("직종", help="인원 테이블의 행 항목")
                            },
                            key="personnel_rows_editor"
                        )
                        
                        # 인원 행이 변경되면 자동으로 프롬프트 업데이트
                        new_personnel_rows = edited_personnel_rows["직종"].tolist()
                        if new_personnel_rows != st.session_state.personnel_rows:
                            st.session_state.personnel_rows = new_personnel_rows
                            # 자동 프롬프트 업데이트
                            st.session_state.prompt = generate_prompt_from_tables()
                            st.info("✅ 인원 테이블 행 변경으로 프롬프트가 자동 업데이트되었습니다.")
                            st.rerun()
                
                with tab4:
                    # 장비 행 편집
                    equipment_df = pd.DataFrame({"구분": st.session_state.equipment_rows})
                    edited_equipment = st.data_editor(
                        equipment_df,
                        use_container_width=True,
                        num_rows="dynamic",
                        height=600,
                        column_config={
                            "구분": st.column_config.TextColumn("구분", help="장비 항목명")
                        },
                        key="equipment_editor"
                    )
                    
                    # 장비가 변경되면 자동으로 프롬프트 업데이트
                    new_equipment_rows = edited_equipment["구분"].tolist()
                    if new_equipment_rows != st.session_state.equipment_rows:
                        st.session_state.equipment_rows = new_equipment_rows
                        # 자동 프롬프트 업데이트
                        st.session_state.prompt = generate_prompt_from_tables()
                        st.info("✅ 장비 테이블 변경으로 프롬프트가 자동 업데이트되었습니다.")
                        st.rerun()
            
            # 상태 메시지를 위한 1분할 섹션
            st.markdown("---")
            
            # 편집기 액션 버튼 - 나란히 배치
            edit_col1, edit_col2, edit_col3, edit_col4 = st.columns(4)
            
            with edit_col1:
                if st.button("💾 저장", key="save_prompt", use_container_width=True):
                    if prompt_name.strip():
                        if save_prompt_to_supabase(prompt_name.strip(), edited_prompt, prompt_description):
                            st.session_state.prompt = edited_prompt
                            st.session_state.current_prompt_name = prompt_name.strip()
                            st.session_state.show_prompt_editor = False
                            st.session_state.prompt_list = get_all_prompts_from_supabase()
                            st.rerun()
                    else:
                        st.error("프롬프트 이름을 입력해주세요.")
            
            with edit_col2:
                if st.button("🔄 적용", key="apply_prompt", use_container_width=True):
                    st.session_state.prompt = edited_prompt
                    st.success("프롬프트가 현재 세션에 적용되었습니다.")
            
            with edit_col3:
                if st.button("🔙 기본값", key="reset_to_default", use_container_width=True):
                    st.session_state.prompt = DEFAULT_PROMPT
                    st.session_state.current_prompt_name = "기본 프롬프트"
                    st.rerun()
            
            with edit_col4:
                if st.button("❌ 취소", key="cancel_edit", use_container_width=True):
                    st.session_state.show_prompt_editor = False
                    st.rerun()
        

    
    if not st.session_state.kakao_work_completed:
        kakao_text = st.text_area("카카오톡 작업보고", placeholder="이곳에 카카오톡 작업보고 내용을 붙여넣으세요...", height=200, label_visibility="collapsed")
        
        # AI 데이터 추출 버튼만 배치
        if st.button("✨AI 데이터 추출", key="structure_button", use_container_width=True):
            if kakao_text:
                # 진행 상황을 단계별로 표시
                progress_placeholder = st.empty()
                status_placeholder = st.empty()
                
                try:
                    # 1단계: 프롬프트 준비
                    progress_placeholder.progress(0, text="📝 프롬프트 준비 중...")
                    prompt = st.session_state.prompt + "\n" + kakao_text
                    
                    # 2단계: AI 모델 호출
                    progress_placeholder.progress(30, text="🤖 AI 모델에 요청 중...")
                    status_placeholder.info("AI가 데이터를 추출하고 있습니다. 잠시만 기다려주세요...")
                    
                    response_text = safe_generate_content(prompt)
                    
                    if response_text:
                        # 3단계: 응답 처리
                        progress_placeholder.progress(70, text="📊 응답 데이터 처리 중...")
                        
                        # AI 응답에서 작업 날짜 추출
                        work_date = extract_work_date_from_response(response_text)
                        
                        # 4단계: 완료
                        progress_placeholder.progress(100, text="✅ 완료!")
                        status_placeholder.success("🎉 AI 데이터 추출이 완료되었습니다!")
                        
                        st.session_state.kakao_results = response_text
                        st.session_state.work_date = work_date  # 추출된 작업 날짜 저장
                        st.session_state.kakao_work_completed = True
                        
                        st.info(f"📅 추출된 작업 날짜: {work_date}")
                        st.toast("✅ 1단계 완료: SNS 작업보고 데이터 구조화 성공!")
                        st.success("✅ 1단계 완료: SNS 작업보고가 성공적으로 처리되었습니다.")
                        
                        # 진행 상황 표시 제거
                        progress_placeholder.empty()
                        status_placeholder.empty()
                        
                        # 페이지 새로고침하여 처리된 데이터 보기 표시
                        st.rerun()
                    else:
                        progress_placeholder.empty()
                        status_placeholder.empty()
                        st.error("❌ AI 응답 생성에 실패했습니다.")
                except Exception as e: 
                    progress_placeholder.empty()
                    status_placeholder.empty()
                    st.error(f"❌ AI 데이터 추출 중 오류: {e}")
                    st.error("💡 네트워크 연결을 확인하거나 잠시 후 다시 시도해주세요.")
            else: 
                st.warning("보고 내용을 입력해주세요.")
    else:
        st.success("✅ 1단계 완료: SNS 작업보고가 성공적으로 처리되었습니다.")
        
        # 처리된 데이터 보기
        with st.expander("📊 처리된 데이터 보기", expanded=True):
            # QA-Checklist 표시
            qa_checklist = extract_qa_checklist_from_response(st.session_state.kakao_results)
            if qa_checklist:
                st.subheader("🔍 QA-Checklist (자동 검증 결과)")
                st.markdown(qa_checklist)
                st.markdown("---")
            
            # AI 응답에서 각 테이블을 개별적으로 추출
            response_text = st.session_state.kakao_results
            
            # 각 테이블 섹션을 찾아서 개별적으로 처리
            table_sections = {
                "시공현황": None,
                "작업내용": None,
                "인원": None,
                "장비": None
            }
            
            # AI 응답 구조 분석 및 테이블 추출
            # 1. 먼저 TSV 블록을 찾기
            tsv_blocks = re.findall(r'```(?:tsv)?\n(.*?)```', response_text, re.DOTALL | re.IGNORECASE)
            
            # 2. TSV 블록이 없으면 섹션별로 추출 시도
            if not tsv_blocks:
                # 시공현황 섹션 찾기
                construction_match = re.search(r'## 1\. 시공현황.*?(?=## 2\.|$)', response_text, re.DOTALL | re.IGNORECASE)
                if construction_match:
                    construction_text = construction_match.group(0)
                    # 시공현황에서 테이블 형태 데이터 추출
                    lines = construction_text.split('\n')
                    table_data = []
                    for line in lines:
                        if line.strip() and not line.startswith('##') and not line.startswith('#'):
                            # 숫자가 포함된 라인을 테이블 데이터로 간주
                            if re.search(r'\d+', line):
                                parts = line.split()
                                if len(parts) >= 2:
                                    table_data.append('\t'.join(parts))
                    if table_data:
                        table_sections["시공현황"] = '\n'.join(table_data)
                
                # 작업내용 섹션 찾기
                work_match = re.search(r'## 2\. 작업내용.*?(?=## 3\.|$)', response_text, re.DOTALL | re.IGNORECASE)
                if work_match:
                    work_text = work_match.group(0)
                    lines = work_text.split('\n')
                    table_data = []
                    for line in lines:
                        if line.strip() and not line.startswith('##') and not line.startswith('#'):
                            if re.search(r'\d+', line):
                                parts = line.split()
                                if len(parts) >= 2:
                                    table_data.append('\t'.join(parts))
                    if table_data:
                        table_sections["작업내용"] = '\n'.join(table_data)
                
                # 인원 섹션 찾기
                personnel_match = re.search(r'## 3\. 인원.*?(?=## 4\.|$)', response_text, re.DOTALL | re.IGNORECASE)
                if personnel_match:
                    personnel_text = personnel_match.group(0)
                    lines = personnel_text.split('\n')
                    table_data = []
                    for line in lines:
                        if line.strip() and not line.startswith('##') and not line.startswith('#'):
                            if re.search(r'\d+', line):
                                parts = line.split()
                                if len(parts) >= 2:
                                    table_data.append('\t'.join(parts))
                    if table_data:
                        table_sections["인원"] = '\n'.join(table_data)
                
                # 장비 섹션 찾기
                equipment_match = re.search(r'## 4\. 장비.*?(?=##|$)', response_text, re.DOTALL | re.IGNORECASE)
                if equipment_match:
                    equipment_text = equipment_match.group(0)
                    lines = equipment_text.split('\n')
                    table_data = []
                    for line in lines:
                        if line.strip() and not line.startswith('##') and not line.startswith('#'):
                            if re.search(r'\d+', line):
                                parts = line.split()
                                if len(parts) >= 2:
                                    table_data.append('\t'.join(parts))
                    if table_data:
                        table_sections["장비"] = '\n'.join(table_data)
            else:
                # TSV 블록이 있으면 기존 방식으로 처리
                table_names = ["시공현황", "작업내용", "인원", "장비"]
                for i, tsv_data in enumerate(tsv_blocks):
                    if i < len(table_names):
                        table_sections[table_names[i]] = tsv_data.strip()
            
            # 디버깅: 찾은 데이터 표시
            found_tables = sum(1 for data in table_sections.values() if data)
            if found_tables > 0:
                st.info(f"🔍 {found_tables}개의 테이블 데이터를 찾았습니다.")
            else:
                st.warning("⚠️ 테이블 데이터를 찾을 수 없습니다. AI 응답을 확인해주세요.")
                st.code(response_text[:1000] + "..." if len(response_text) > 1000 else response_text)
            
            processed_tables = []
            for table_name, tsv_data in table_sections.items():
                if tsv_data:
                    df = parse_tsv_to_dataframe(fix_tsv_field_count(tsv_data))
                    if df is not None:
                        st.subheader(table_name)
                        st.dataframe(df)
                        processed_tables.append(df)
                    else:
                        st.warning(f"⚠️ {table_name} 테이블 파싱 실패")
                else:
                    st.warning(f"⚠️ {table_name} 테이블을 찾을 수 없습니다")
            
            st.session_state.processed_tables = processed_tables
            
            # Supabase 저장 기능 (항상 표시)
            st.markdown("---")
            st.markdown("### 💾 데이터 저장")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                if st.button("💾 Supabase에 저장", key="save_step1", use_container_width=True):
                    # 데이터를 딕셔너리로 변환하여 저장
                    report_data = {}
                    for table_name, tsv_data in table_sections.items():
                        if tsv_data:
                            df = parse_tsv_to_dataframe(fix_tsv_field_count(tsv_data))
                            if df is not None:
                                report_data[table_name] = df.to_dict('records')
                    
                    if report_data:
                        # Supabase에 저장 (추출된 작업 날짜 사용)
                        work_date = st.session_state.get('work_date', datetime.now().strftime('%Y-%m-%d'))
                        if save_step1_to_supabase(report_data, work_date):
                            st.session_state.daily_report_saved = True
                            st.success("✅ 1단계 데이터가 4개 테이블에 저장되었습니다!")
                            st.info("💡 시공현황, 작업내용, 인원, 장비 데이터가 각각 별도 테이블에 저장되었습니다.")
                        else:
                            st.error("❌ 1단계 데이터 저장에 실패했습니다.")
                    else:
                        st.warning("⚠️ 저장할 데이터가 없습니다.")
            
            with col2:
                if st.button("📅 날짜별 데이터 조회", key="load_step1", use_container_width=True):
                    selected_date = st.date_input("조회할 날짜 선택", value=datetime.now(), key="load_date_step1")
                    loaded_data = load_from_supabase("daily_report", selected_date.strftime('%Y-%m-%d'))
                    if loaded_data:
                        st.json(loaded_data)
                    else:
                        st.info("해당 날짜의 데이터가 없습니다.")

# --- STEP 2: BASIC INFO INPUT ---
with st.container():
    st.markdown('<h3><span style="font-size: 1.5em;">📋</span> 2. 기본정보 입력</h3>', unsafe_allow_html=True)
    
    # 기본정보 입력 섹션 - 첫 번째 행: 작성일자와 공사명
    col1, col2 = st.columns(2)
    
    with col1:
        # 추출된 작업 날짜가 있으면 사용, 없으면 현재 날짜 사용
        default_date = None
        if st.session_state.get('work_date'):
            try:
                default_date = datetime.strptime(st.session_state.work_date, '%Y-%m-%d').date()
            except:
                default_date = datetime.now().date()
        else:
            default_date = datetime.now().date()
        
        report_date = st.date_input("작성일자", value=default_date, key="excel_report_date")
    
    with col2:
        project_name = st.text_input("공사명", value="신안산선 복선전철 민간투자사업 4-1공구", key="excel_project_name")
    
    # 두 번째 행: 날씨정보와 공정률
    col3, col4 = st.columns(2)
    
    with col3:
        # 날씨정보 섹션
        st.markdown("**🌤️ 날씨정보**")
        
        weather_col1, weather_col2, weather_col3 = st.columns(3)
        
        # 저장된 날씨 데이터가 있으면 사용
        weather_auto_fill = st.session_state.get('weather_auto_fill_data', {})
        default_min_temp = weather_auto_fill.get('min_temp', 18.2) if weather_auto_fill else 18.2
        default_max_temp = weather_auto_fill.get('max_temp', 25.5) if weather_auto_fill else 25.5
        default_precipitation = weather_auto_fill.get('precipitation', 0.0) if weather_auto_fill else 0.0
        
        with weather_col1:
            min_temp = st.number_input("최저기온 (°C)", value=default_min_temp, key="excel_min_temp", format="%.1f")
        with weather_col2:
            max_temp = st.number_input("최고기온 (°C)", value=default_max_temp, key="excel_max_temp", format="%.1f")
        with weather_col3:
            precipitation = st.number_input("강수량 (mm)", value=default_precipitation, key="excel_precipitation", format="%.1f")
        
        # 날씨 상태 표시
        if weather_auto_fill:
            st.info(f"🌤️ 현재 날씨: 최고 {weather_auto_fill.get('max_temp', 0):.1f}°C / 최저 {weather_auto_fill.get('min_temp', 0):.1f}°C / 강수량 {weather_auto_fill.get('precipitation', 0):.1f}mm")
        
        # AI 자동채우기 버튼 (제일 왼쪽, 작은 버튼)
        if st.button("🪄AI 자동채우기", key="weather_auto_fill_button", help="AI 자동채우기"):
            try:
                weather_data = get_weather_data()
                if weather_data:
                    st.session_state.weather_auto_fill_data = weather_data
                    st.success(f"✅ 날씨 데이터 가져오기 성공: {weather_data['max_temp']}°C / {weather_data['min_temp']}°C")
                    st.rerun()
                else:
                    st.error(f"❌ 날씨 데이터 가져오기 실패")
            except Exception as e:
                st.error(f"❌ 날씨 데이터 가져오기 실패")
    
    with col4:
        # 공정률 섹션
        st.markdown("**📊 공정률**")
        progress_col1, progress_col2 = st.columns(2)
        
        with progress_col1:
            planned_progress = st.number_input("계획 (%)", value=50, key="excel_planned_progress")
        with progress_col2:
            actual_progress = st.number_input("실적 (%)", value=48.5, key="excel_actual_progress")

# --- STEP 3: WORK REPORT GENERATION ---
with st.container():
    st.markdown('<h3><span style="font-size: 1.5em;">📊</span> 3. 작업일보 생성</h3>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)

    # 작업일보 생성 기능 추가 (항상 표시)
    
    # 간단한 템플릿 선택
    col1, col2 = st.columns([3, 1])
    
    with col1:
        # 템플릿 옵션 (최신 업로드 파일이 기본값)
        template_options = ["새로운 양식으로 생성"]
        
        # Supabase에서 저장된 템플릿 목록 가져오기 (최신 순으로 정렬)
        if supabase_client:
            try:
                saved_templates = get_all_templates()
                if saved_templates:
                    # 최신 템플릿을 첫 번째로 설정
                    template_options.extend([t["template_name"] for t in saved_templates])
            except:
                pass
        
        selected_template_option = st.selectbox(
            "템플릿 선택",
            options=template_options,
            index=0
        )
    
    with col2:
        # 템플릿 업로드 버튼
        if st.button("📤 템플릿 업로드", key="upload_template", use_container_width=True):
            st.session_state.show_template_upload = True
            st.rerun()
    
    # 간단한 템플릿 업로드 섹션
    if st.session_state.get('show_template_upload', False):
        with st.expander("📤 템플릿 업로드", expanded=True):
            uploaded_template = st.file_uploader(
                "엑셀 템플릿 파일 선택",
                type=['xlsx', 'xls'],
                key="template_uploader"
            )
            
            if uploaded_template:
                template_name = st.text_input("템플릿 이름", value="새 템플릿")
                
                if st.button("💾 템플릿 저장", key="save_template", use_container_width=True):
                    template_bytes = uploaded_template.read()
                    success, message = save_template_to_supabase(template_bytes, template_name, "업로드된 템플릿")
                    if success:
                        st.success(f"✅ {message}")
                        st.rerun()
                    else:
                        st.error(f"❌ {message}")
            
            if st.button("❌ 닫기", key="close_template_upload"):
                st.session_state.show_template_upload = False
                st.rerun()
    
    # 템플릿 파일 처리
    template_bytes = None
    
    if selected_template_option != "새로운 양식으로 생성":
        # Supabase에서 선택된 템플릿 로드
        template_bytes = get_template_from_supabase(selected_template_option)
        if template_bytes:
            st.success(f"✅ 템플릿 '{selected_template_option}' 로드 완료")
        else:
            st.error(f"❌ 템플릿 '{selected_template_option}' 로드 실패")
        
        # 셀 매핑 설정 (템플릿이 있는 경우)
        with st.expander("⚙️ 셀 매핑 설정", expanded=True):
            # 템플릿이 로드되지 않은 경우 안내 메시지
            if not template_bytes:
                st.warning("⚠️ 템플릿을 먼저 선택하거나 업로드해주세요.")
            
            # 콤팩트한 3열 레이아웃
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("**기본 정보**")
                st.session_state.date_cell = st.text_input("작성일자", value=st.session_state.get('date_cell', 'u2'), disabled=not template_bytes)
                st.session_state.project_cell = st.text_input("공사명", value=st.session_state.get('project_cell', 'd4'), disabled=not template_bytes)
                st.session_state.max_temp_cell = st.text_input("최고기온", value=st.session_state.get('max_temp_cell', 'o4'), disabled=not template_bytes)
                st.session_state.min_temp_cell = st.text_input("최저기온", value=st.session_state.get('min_temp_cell', 'o5'), disabled=not template_bytes)
                st.session_state.precipitation_cell = st.text_input("강수량", value=st.session_state.get('precipitation_cell', 'o6'), disabled=not template_bytes)
            
            with col2:
                st.markdown("**공정률**")
                st.session_state.planned_progress_cell = st.text_input("계획", value=st.session_state.get('planned_progress_cell', 'w4'), disabled=not template_bytes)
                st.session_state.actual_progress_cell = st.text_input("실적", value=st.session_state.get('actual_progress_cell', 'w5'), disabled=not template_bytes)
                
                st.markdown("**테이블 시작 위치**")
                st.session_state.table_construction_cell = st.text_input("시공현황", value=st.session_state.get('table_construction_cell', 'q8'), disabled=not template_bytes)
                st.session_state.table_work_cell = st.text_input("작업내용", value=st.session_state.get('table_work_cell', 'q11'), disabled=not template_bytes)
                st.session_state.table_personnel_cell = st.text_input("인원", value=st.session_state.get('table_personnel_cell', 'ac66'), disabled=not template_bytes)
                st.session_state.table_equipment_cell = st.text_input("장비", value=st.session_state.get('table_equipment_cell', 'ac106'), disabled=not template_bytes)
            
            with col3:
                st.markdown("**저장/불러오기**")
                if st.button("💾 저장", key="save_mapping", disabled=not template_bytes, use_container_width=True):
                    mapping_data = {
                        'date_cell': st.session_state.date_cell,
                        'project_cell': st.session_state.project_cell,
                        'max_temp_cell': st.session_state.max_temp_cell,
                        'min_temp_cell': st.session_state.min_temp_cell,
                        'precipitation_cell': st.session_state.precipitation_cell,
                        'planned_progress_cell': st.session_state.planned_progress_cell,
                        'actual_progress_cell': st.session_state.actual_progress_cell,
                        'table_construction_cell': st.session_state.table_construction_cell,
                        'table_work_cell': st.session_state.table_work_cell,
                        'table_personnel_cell': st.session_state.table_personnel_cell,
                        'table_equipment_cell': st.session_state.table_equipment_cell
                    }
                    
                    success, message = save_cell_mapping_to_supabase(mapping_data, "default")
                    if success:
                        st.session_state.mapping_save_success = True
                        st.session_state.mapping_save_message = f"✅ {message}"
                        st.toast("💾 셀 매핑 설정이 저장되었습니다!", icon="✅")
                        st.rerun()
                    else:
                        st.session_state.mapping_save_success = False
                        st.session_state.mapping_save_message = f"❌ {message}"
                        st.toast("❌ 저장에 실패했습니다.", icon="❌")
                        st.rerun()
                
                # 저장 결과 메시지 표시
                if hasattr(st.session_state, 'mapping_save_success'):
                    if st.session_state.mapping_save_success:
                        st.success(st.session_state.mapping_save_message)
                        # 메시지 표시 후 상태 초기화
                        del st.session_state.mapping_save_success
                        del st.session_state.mapping_save_message
                    else:
                        st.error(st.session_state.mapping_save_message)
                        # 메시지 표시 후 상태 초기화
                        del st.session_state.mapping_save_success
                        del st.session_state.mapping_save_message
                
                if st.button("📥 불러오기", key="load_mapping", disabled=not template_bytes, use_container_width=True):
                    loaded_mapping = get_cell_mapping_from_supabase("default")
                    if loaded_mapping:
                        for key, value in loaded_mapping.items():
                            if key in st.session_state:
                                st.session_state[key] = value
                        st.session_state.mapping_load_success = True
                        st.session_state.mapping_load_message = "✅ 매핑 설정을 불러왔습니다."
                        st.toast("📥 매핑 설정을 불러왔습니다!", icon="✅")
                        st.rerun()
                    else:
                        st.session_state.mapping_load_success = False
                        st.session_state.mapping_load_message = "⚠️ 저장된 매핑 설정이 없습니다."
                        st.toast("⚠️ 저장된 매핑 설정이 없습니다.", icon="⚠️")
                        st.rerun()
                
                # 불러오기 결과 메시지 표시
                if hasattr(st.session_state, 'mapping_load_success'):
                    if st.session_state.mapping_load_success:
                        st.success(st.session_state.mapping_load_message)
                        # 메시지 표시 후 상태 초기화
                        del st.session_state.mapping_load_success
                        del st.session_state.mapping_load_message
                    else:
                        st.warning(st.session_state.mapping_load_message)
                        # 메시지 표시 후 상태 초기화
                        del st.session_state.mapping_load_success
                        del st.session_state.mapping_load_message
        
        # 공사일보 생성 버튼
        st.markdown("---")
        
        # 공사일보 생성 버튼 (전체 너비)
        if st.button("📊 공사일보 생성", key="create_report", use_container_width=True):
            # 템플릿이 있는 경우 기본값 설정 (검증 없이)
            if template_bytes:
                # 기본값 설정
                default_mappings = {
                    'date_cell': 'u2',
                    'project_cell': 'd4', 
                    'max_temp_cell': 'o4',
                    'min_temp_cell': 'o5',
                    'precipitation_cell': 'o6',
                    'planned_progress_cell': 'w4',
                    'actual_progress_cell': 'w5',
                    'table_construction_cell': 'q8',
                    'table_work_cell': 'q11',
                    'table_personnel_cell': 'ac66',
                    'table_equipment_cell': 'ac106'
                }
                
                # 세션 상태에 기본값 설정
                for key, default_value in default_mappings.items():
                    if not st.session_state.get(key):
                        st.session_state[key] = default_value
            
            with st.spinner('📊 공사일보를 생성하고 있습니다...'):
                try:
                    # 기본 정보 준비 (1단계 결과 사용)
                    # 1단계에서 추출된 날짜 사용
                    work_date = st.session_state.get('work_date', datetime.now().strftime('%Y-%m-%d'))
                    
                    # 날씨 데이터: 2단계 입력값 우선, 없으면 API에서 가져오기
                    weather_data = {}
                    
                    # 2단계에서 사용자가 입력한 날씨 데이터 확인
                    if st.session_state.get('excel_max_temp') is not None:
                        weather_data['max_temp'] = st.session_state.get('excel_max_temp')
                    if st.session_state.get('excel_min_temp') is not None:
                        weather_data['min_temp'] = st.session_state.get('excel_min_temp')
                    if st.session_state.get('excel_precipitation') is not None:
                        weather_data['precipitation'] = st.session_state.get('excel_precipitation')
                    
                    # 입력된 값이 없으면 API에서 가져오기
                    if not weather_data:
                        weather_data = get_weather_data()
                    
                    planned_progress = st.session_state.get('excel_planned_progress', 48.0)
                    actual_progress = st.session_state.get('excel_actual_progress', 48.5)
                    progress_diff = actual_progress - planned_progress
                    
                    basic_info = {
                        'date': work_date,
                        'project_name': st.session_state.get('excel_project_name', '서울지하철 2호선 신풍~도림 구간 건설공사'),
                        'max_temp': weather_data.get('max_temp', 25.5),
                        'min_temp': weather_data.get('min_temp', 18.2),
                        'precipitation': weather_data.get('precipitation', 0.0),
                        'planned_progress': planned_progress,
                        'actual_progress': actual_progress,
                        'progress_diff': progress_diff
                    }
                    
                    # 테이블 데이터 준비 (SNS 작업보고가 있는 경우)
                    tables_data = {}
                    if st.session_state.kakao_work_completed and st.session_state.get('kakao_results'):
                        tables = st.session_state.kakao_results.split("```")
                        # 날씨정보는 제외하고 실제 테이블만 처리
                        table_names = ["시공현황", "작업내용", "인원", "장비"]
                        real_tables = [t.strip() for t in tables if "\t" in t.strip()]
                        
                        # 날씨정보를 제외하고 실제 테이블만 처리
                        table_index = 0
                        for i, tsv_data in enumerate(real_tables):
                            try:
                                # TSV 데이터 정제
                                cleaned_tsv = re.sub(r'^tsv\n', '', tsv_data, flags=re.IGNORECASE)
                                fixed_tsv = fix_tsv_field_count(cleaned_tsv)
                                df = parse_tsv_to_dataframe(fixed_tsv)
                                
                                if df is not None and table_index < len(table_names):
                                    tables_data[table_names[table_index]] = df
                                    table_index += 1
                                    
                            except Exception as e:
                                continue
                    else:
                        # SNS 작업보고가 없는 경우 기본 테이블 생성
                        import pandas as pd
                        st.info("ℹ️ SNS 작업보고가 없어 기본 테이블로 공사일보를 생성합니다.")
                        
                        # 기본 테이블 데이터 생성
                        default_construction = pd.DataFrame({
                            '구분': ['본선터널(1구간)', '본선터널(2구간)', '신풍정거장', '도림정거장'],
                            '누계': ['0%', '0%', '0%', '0%']
                        })
                        default_work = pd.DataFrame({
                            '구분': ['본선터널(1구간)', '본선터널(2구간)', '신풍정거장', '도림정거장'],
                            '금일작업': ['준비중', '준비중', '준비중', '준비중']
                        })
                        default_personnel = pd.DataFrame({
                            '구분': ['직영반장', '터널공', '목공', '철근공'],
                            '인원': ['0명', '0명', '0명', '0명']
                        })
                        default_equipment = pd.DataFrame({
                            '구분': ['B/H(1.0LC)', '덤프트럭(5T)', '앵글크레인(25T)', '믹서트럭'],
                            '대수': ['0대', '0대', '0대', '0대']
                        })
                        
                        tables_data = {
                            "시공현황": default_construction,
                            "작업내용": default_work,
                            "인원": default_personnel,
                            "장비": default_equipment
                        }
                        

                    
                    # 엑셀 파일 생성
                    try:
                        if template_bytes:
                            # 템플릿 사용 - 셀 매핑 설정 전달
                            cell_mapping = {
                                'date': st.session_state.get('date_cell', 'u2'),
                                'project_name': st.session_state.get('project_cell', 'd4'),
                                'max_temp': st.session_state.get('max_temp_cell', 'o4'),
                                'min_temp': st.session_state.get('min_temp_cell', 'o5'),
                                'precipitation': st.session_state.get('precipitation_cell', 'o6'),
                                'planned_progress': st.session_state.get('planned_progress_cell', 'w4'),
                                'actual_progress': st.session_state.get('actual_progress_cell', 'w5'),
                                'progress_diff': st.session_state.get('progress_diff_cell', 'w6')
                            }
                            
                            table_mapping = {
                                '시공현황': st.session_state.get('table_construction_cell', 'ac10'),
                                '작업내용': st.session_state.get('table_work_cell', 'ac48'),
                                '인원': st.session_state.get('table_personnel_cell', 'ac66'),
                                '장비': st.session_state.get('table_equipment_cell', 'ac106')
                            }
                            
                            # 전일 데이터 가져오기
                            previous_data = get_previous_day_data(basic_info['date'])
                            
                            excel_bytes = insert_data_to_excel_with_mapping(
                                template_bytes, 
                                basic_info, 
                                tables_data,
                                cell_mapping=cell_mapping,
                                table_mapping=table_mapping,
                                previous_data=previous_data
                            )
                        else:
                            # 새로운 양식으로 생성
                            excel_bytes = create_excel_report(
                                basic_info=basic_info,
                                tables_data=tables_data
                            )
                        
                        if excel_bytes:
                            st.success("✅ 공사일보 생성 완료!")
                            
                            # 엑셀 다운로드 버튼
                            st.download_button(
                                label="📥 공사일보 다운로드",
                                data=excel_bytes,
                                file_name=f"공사일보_{basic_info['date']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            # 저장 결과 메시지 표시 (화면 새로고침 후에도 유지)
                            if hasattr(st.session_state, 'daily_report_saved'):
                                if st.session_state.daily_report_saved:
                                    st.success(st.session_state.save_success_message)
                                    st.info(f"📅 저장된 날짜: {st.session_state.save_success_date}")
                                    # 메시지 표시 후 상태 초기화
                                    del st.session_state.daily_report_saved
                                    del st.session_state.save_success_message
                                    del st.session_state.save_success_date
                                else:
                                    st.error(st.session_state.save_error_message)
                                    # 메시지 표시 후 상태 초기화
                                    del st.session_state.daily_report_saved
                                    del st.session_state.save_error_message
                            
                            # Supabase 저장 버튼 추가
                            col1, col2 = st.columns(2)
                            with col1:
                                # 저장 버튼 상태에 따른 텍스트 변경
                                save_button_text = "💾 Supabase에 저장"
                                if hasattr(st.session_state, 'daily_report_saved') and st.session_state.daily_report_saved:
                                    save_button_text = "✅ 저장 완료"
                                
                                save_button_clicked = st.button(save_button_text, key="save_to_supabase", use_container_width=True)
                                if save_button_clicked:
                                    # 저장 진행 상태 표시
                                    with st.status("💾 Supabase 저장 중...", expanded=True) as status:
                                        st.write("🔍 저장 프로세스를 시작합니다...")
                                        
                                        # Supabase 연결 테스트
                                        connection_success, connection_message = test_supabase_connection()
                                        if not connection_success:
                                            st.error(f"❌ Supabase 연결 실패: {connection_message}")
                                            status.update(label="❌ 저장 실패 - 연결 오류", state="error")
                                        else:
                                            st.success("✅ Supabase 연결 성공!")
                                            
                                            # 테이블 존재 여부 확인
                                            table_exists, table_message = check_daily_report_data_table()
                                            if not table_exists:
                                                st.warning(f"⚠️ daily_report_data 테이블이 존재하지 않습니다: {table_message}")
                                                st.write("🔍 테이블을 자동으로 생성합니다...")
                                                
                                                # 테이블 생성 시도
                                                create_success, create_message = create_daily_report_data_table()
                                                if not create_success:
                                                    st.error(f"❌ 테이블 생성 실패: {create_message}")
                                                    st.info("💡 Supabase에서 수동으로 테이블을 생성해주세요.")
                                                    status.update(label="❌ 저장 실패 - 테이블 생성 오류", state="error")
                                                else:
                                                    st.success("✅ 테이블 생성 완료!")
                                            else:
                                                st.success("✅ 테이블이 존재합니다.")
                                            
                                            # 저장 시도
                                            try:
                                                # 저장할 데이터 준비 (엑셀 파일에서 추출된 셀 데이터만)
                                                st.write("🔍 3단계 엑셀 파일에서 추출된 셀 데이터를 저장합니다.")
                                                report_data = {
                                                    "시공현황": tables_data.get("시공현황"),
                                                    "작업내용": tables_data.get("작업내용"),
                                                    "인원": tables_data.get("인원"),
                                                    "장비": tables_data.get("장비"),
                                                    "기본정보": basic_info,
                                                    "excel_bytes": excel_bytes  # 엑셀 파일 바이트 추가
                                                }
                                                
                                                st.write(f"🔍 저장할 데이터 준비 완료: {list(report_data.keys())}")
                                                
                                                work_date = basic_info['date']
                                                st.write(f"🔍 저장할 날짜: {work_date}")
                                                
                                                save_result = save_to_supabase("daily_report", report_data, work_date)
                                                
                                                if save_result:
                                                    st.session_state.daily_report_saved = True
                                                    st.session_state.save_success_message = "🎉 작업일보가 Supabase에 성공적으로 저장되었습니다!"
                                                    st.session_state.save_success_date = work_date
                                                    status.update(label="✅ 저장 완료!", state="complete")
                                                    st.success("🎉 작업일보가 Supabase에 성공적으로 저장되었습니다!")
                                                    st.balloons()
                                                    st.toast("💾 Supabase 저장 완료!", icon="✅")
                                                else:
                                                    st.session_state.daily_report_saved = False
                                                    st.session_state.save_error_message = "❌ Supabase 저장에 실패했습니다."
                                                    status.update(label="❌ 저장 실패", state="error")
                                                    st.error("❌ Supabase 저장에 실패했습니다.")
                                                    st.info("💡 네트워크 연결이나 Supabase 설정을 확인해주세요.")
                                                    st.toast("❌ 저장 실패", icon="❌")
                                                    
                                            except Exception as save_error:
                                                status.update(label="❌ 저장 중 오류 발생", state="error")
                                                st.error(f"❌ 저장 중 오류: {save_error}")
                                                import traceback
                                                st.error(f"❌ 상세 오류: {traceback.format_exc()}")
                                                st.info("💡 문제가 지속되면 관리자에게 문의해주세요.")
                                        

                            
                            with col2:
                                if st.button("📅 저장된 데이터 조회", key="load_saved_data", use_container_width=True):
                                    selected_date = st.date_input("조회할 날짜 선택", value=datetime.strptime(basic_info['date'], '%Y-%m-%d').date(), key="load_date_step3")
                                    loaded_data = load_from_supabase("daily_report", selected_date.strftime('%Y-%m-%d'))
                                    if loaded_data:
                                        st.json(loaded_data)
                                    else:
                                        st.info("해당 날짜의 저장된 데이터가 없습니다.")
                        else:
                            st.error("❌ 공사일보 생성 실패: 파일 생성에 실패했습니다.")
                    except Exception as excel_error:
                        st.error(f"❌ 엑셀 파일 생성 중 오류: {excel_error}")
                        st.info("💡 템플릿 파일이나 데이터 형식을 확인해주세요.")
                except Exception as e:
                    st.error(f"❌ 공사일보 생성 중 오류: {e}")
                    st.info("💡 기본 정보를 확인해주세요.")
        
        # PDF 보고서 생성 기능 제거됨





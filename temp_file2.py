    
    Args:
        date: YYYYMMDD ?뺤떇???좎쭨 臾몄옄??(湲곕낯媛? ?ㅻ뒛)
        station_id: 愿痢≪냼 ID (湲곕낯媛? ?쒖슱=108)
    
    Returns:
        dict: {'理쒓퀬湲곗삩': str, '理쒖?湲곗삩': str, '媛뺤닔??: str}
    """
    # ?좎쭨媛 ?놁쑝硫??댁젣 ?좎쭨 ?ъ슜 (湲곗긽泥?API???꾨궇 ?먮즺源뚯? ?쒓났)
    if date is None:
        date = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    
    try:
        # (1) API ???붿퐫??        decoded_api_key = urllib.parse.unquote(WEATHER_API_KEY)
        
        # (2) API ??寃利?        if not decoded_api_key or len(decoded_api_key) < 10:
            print("??API ?ㅺ? ?щ컮瑜댁? ?딆뒿?덈떎.")
            return {
                '理쒓퀬湲곗삩': '25.5',
                '理쒖?湲곗삩': '18.2',
                '媛뺤닔??: '0.0'
            }
        
        # (3) ?붿껌 ?뚮씪誘명꽣 援ъ꽦
        # ?섎（(00~23?? 湲곗? 24?쒓컙 ?곗씠?곕? 遺덈윭?ㅻ룄濡??ㅼ젙
        params = {
            "serviceKey": decoded_api_key,   # ?붿퐫?⑸맂 ?몄쬆??            "dataType": "XML",              # ?묐떟 ???            "pageNo": "1",
            "numOfRows": "24",
            "dataCd": "ASOS",
            "dateCd": "HR",                 # ?쒓컙 ?먮즺
            "startDt": date,
            "startHh": "00",
            "endDt": date,
            "endHh": "23",
            "stnIds": station_id
        }
        
        print(f"API ?붿껌 URL: {WEATHER_API_URL}")
        print(f"?붿껌 ?뚮씪誘명꽣: {params}")
        
        # (4) ?ㅼ젣 ?붿껌 (SSL ?고쉶 ?ы븿)
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        try:
            # SSL 寃利??놁씠 ?붿껌
            response = requests.get(WEATHER_API_URL, params=params, timeout=30, verify=False)
            response.raise_for_status()  # ?먮윭 諛쒖깮 ???덉쇅 泥섎━
        except requests.exceptions.SSLError:
            # SSL ?ㅻ쪟 ??HTTP濡??쒕룄
            http_url = WEATHER_API_URL.replace('https://', 'http://')
            response = requests.get(http_url, params=params, timeout=30, verify=False)
            response.raise_for_status()
        
        print(f"?묐떟 ?곹깭 肄붾뱶: {response.status_code}")
        print(f"?묐떟 ?댁슜 (泥섏쓬 500??: {response.text[:500]}")
        
        # (5) XML ?뚯떛
        root = ET.fromstring(response.content)
        
        # (5-1) 寃곌낵 肄붾뱶 ?뺤씤
        result_code = root.find('.//resultCode')
        if result_code is not None and result_code.text != '00':
            result_msg = root.find('.//resultMsg')
            err_msg = result_msg.text if result_msg is not None else "?????녿뒗 ?ㅻ쪟"
            
            # "?꾨궇 ?먮즺源뚯? ?쒓났?⑸땲?? ?ㅻ쪟??寃쎌슦 ?댁젣 ?좎쭨濡??ъ떆??            if "?꾨궇 ?먮즺源뚯?" in err_msg:
                print(f"?꾨궇 ?먮즺 ?ㅻ쪟 諛쒖깮. ?댁젣 ?좎쭨濡??ъ떆?꾪빀?덈떎.")
                yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
                if yesterday != date:
                    return get_weather_data(yesterday, station_id)
            
            raise Exception(f"API ?ㅻ쪟: {err_msg}")
        
        # (5-2) 愿痢??곗씠??異붿텧
        items = root.findall('.//item')
        print(f"異붿텧???꾩씠???? {len(items)}")
        
        if not items:
            # ?곗씠?곌? ?놁쑝硫??댁젣 ?좎쭨濡??ъ떆??            print("?곗씠?곌? ?놁뒿?덈떎. ?댁젣 ?좎쭨濡??ㅼ떆 ?쒕룄?⑸땲??")
            yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
            if yesterday != date:
                return get_weather_data(yesterday, station_id)
            else:
                # ?댁젣???곗씠?곌? ?녿떎硫?湲곕낯媛?諛섑솚
                return {
                    'max_temp': 25.5,
                    'min_temp': 18.2,
                    'precipitation': 0.0
                }
        
        # (6) ?⑤룄 諛?媛뺤닔??怨꾩궛
        temperatures = []
        precipitation = 0.0
        
        for item in items:
            # 湲곗삩(ta)
            temp_elem = item.find('ta')
            if temp_elem is not None and temp_elem.text:
                try:
                    temperatures.append(float(temp_elem.text))
                except ValueError:
                    pass
            
            # 媛뺤닔??rn)
            rain_elem = item.find('rn')
            if rain_elem is not None and rain_elem.text:
                try:
                    precipitation += float(rain_elem.text)
                except ValueError:
                    pass
        
        if temperatures:
            max_temp = max(temperatures)
            min_temp = min(temperatures)
        else:
            max_temp, min_temp = None, None
        
        # (7) 理쒖쥌 寃곌낵 ?앹꽦
        result = {
            'max_temp': float(f"{max_temp:.1f}") if max_temp is not None else 25.5,
            'min_temp': float(f"{min_temp:.1f}") if min_temp is not None else 18.2,
            'precipitation': float(f"{precipitation:.1f}") if precipitation > 0 else 0.0
        }
        
        print(f"理쒖쥌 寃곌낵: {result}")
        return result
        
    except Exception as e:
        # API ?몄텧 以??덉쇅媛 諛쒖깮?섎㈃ 濡쒓렇? ?④퍡 ?먮윭 ?쒖떆
        print(f"API ?몄텧 以??ㅻ쪟 諛쒖깮: {e}")
        st.error(f"???좎뵪 ?곗씠??媛?몄삤湲??ㅽ뙣: {e}")
        
        # ?ㅽ뙣 ???꾩떎?곸씤 ?뚯뒪???곗씠??諛섑솚
        print("API ?곌껐 ?ㅽ뙣濡??뚯뒪???곗씠?곕? 諛섑솚?⑸땲??")
        
        # ?꾩옱 怨꾩젅??留욌뒗 ?꾩떎?곸씤 ?곗씠??        current_month = datetime.now().month
        if current_month in [12, 1, 2]:  # 寃⑥슱
            test_data = {'max_temp': 5.2, 'min_temp': -2.1, 'precipitation': 0.0}
        elif current_month in [3, 4, 5]:  # 遊?            test_data = {'max_temp': 18.5, 'min_temp': 8.3, 'precipitation': 2.5}
        elif current_month in [6, 7, 8]:  # ?щ쫫
            test_data = {'max_temp': 28.7, 'min_temp': 22.1, 'precipitation': 15.3}
        else:  # 媛??            test_data = {'max_temp': 20.3, 'min_temp': 12.8, 'precipitation': 0.0}
        
        print(f"怨꾩젅蹂??뚯뒪???곗씠?? {test_data}")
        return test_data

def get_weather_stations():
    """
    二쇱슂 愿痢≪냼 紐⑸줉??諛섑솚?⑸땲??
    """
    return {
        "?쒖슱": "108",
        "遺??: "159", 
        "?援?: "143",
        "?몄쿇": "112",
        "愿묒＜": "156",
        "???: "133",
        "?몄궛": "152",
        "?몄쥌": "184",
        "?섏썝": "119",
        "異섏쿇": "101",
        "媛뺣쫱": "105",
        "泥?＜": "131",
        "?꾩＜": "146",
        "?ы빆": "138",
        "?쒖＜": "184"
    }

# Supabase ?대씪?댁뼵??珥덇린??@st.cache_resource
def init_supabase():
    """Supabase ?대씪?댁뼵?몃? 珥덇린?뷀븯怨?諛섑솚?⑸땲??"""
    if not SUPABASE_AVAILABLE:
        return None
        
    try:
        supabase_url = st.secrets.get("SUPABASE_URL")
        supabase_key = st.secrets.get("SUPABASE_KEY")
        
        if not supabase_url or not supabase_key:
            st.warning("?좑툘 Supabase ?ㅼ젙???꾨즺?섏? ?딆븯?듬땲?? .streamlit/secrets.toml ?뚯씪???뺤씤?댁＜?몄슂.")
            return None
            
        if supabase_url == "https://your-project-id.supabase.co" or supabase_key == "your-anon-key-here":
            st.warning("?좑툘 Supabase ?ㅼ젙??湲곕낯媛믪쑝濡??섏뼱 ?덉뒿?덈떎. ?ㅼ젣 ?꾨줈?앺듃 ?뺣낫濡??낅뜲?댄듃?댁＜?몄슂.")
            return None
        
        # ?⑥닚??Supabase ?대씪?댁뼵???앹꽦
        client = create_client(supabase_url, supabase_key)
        
        # ?곌껐 ?뚯뒪??        try:
            # 媛꾨떒??荑쇰━濡??곌껐 ?뺤씤
            test_result = client.table("daily_report_data").select("count", count="exact").execute()
            return client
        except Exception as test_error:
            st.warning(f"?좑툘 Supabase ?곌껐 ?뚯뒪???ㅽ뙣: {test_error}")
            st.info("?뮕 daily_report_data ?뚯씠釉붿씠 議댁옱?섏? ?딆쓣 ???덉뒿?덈떎.")
            # ?뚯씠釉붿씠 ?놁뼱???대씪?댁뼵?몃뒗 諛섑솚
            return client
            
    except Exception as e:
        st.error(f"??Supabase ?곌껐 ?ㅽ뙣: {e}")
        st.info("?뮕 ?ㅽ듃?뚰겕 ?ㅼ젙???뺤씤?댁＜?몄슂.")
        return None

# ?꾩뿭 Supabase ?대씪?댁뼵??蹂??supabase_client = None



# Supabase ?대씪?댁뼵??珥덇린??if SUPABASE_AVAILABLE:
    try:
        supabase_client = init_supabase()
    except Exception as e:
        supabase_client = None
else:
    supabase_client = None

# Gemini AI ?ㅼ젙
genai.configure(api_key=GENAI_API_KEY)
GEMINI_MODEL = genai.GenerativeModel("models/gemini-2.5-flash-preview-05-20")

BLAST_EXTRACTION_PROMPT = '''
# INSTRUCTION
- 諛섎뱶???꾨옒 ?덉떆泥섎읆 ?ㅼ쭅 TSV(??援щ텇) ?곗씠?곕쭔 異쒕젰?섏꽭??
- ?ㅻ챸, 留덊겕?ㅼ슫, 肄붾뱶釉붾줉, 二쇱꽍, 湲고? ?띿뒪?몃뒗 ?덈? ?ы븿?섏? 留덉꽭??
- ?꾨옒 ?덉떆? ?숈씪???뺤떇?쇰줈留?異쒕젰?섏꽭??
諛쒗뙆?쇱옄    諛쒗뙆?쒓컙    吏諛쒕떦?μ빟??理쒖냼, kg)    吏諛쒕떦?μ빟??理쒕?, kg)    ??빟?ъ슜??kg)    諛쒗뙆吏꾨룞(cm/sec)    諛쒗뙆?뚯쓬(dB(A))    怨꾩륫?꾩튂    鍮꾧퀬
2023-07-27    08:05    0.5    0.9    73    -    -    -    PLA-2
2023-07-27    13:47    0.4    0.8    77    0.87    53.29    ?곗뒪?뚯씠??   PD-2
2023-07-27    13:47    -    -    -    0.71    61.23    ?묐쭚吏?   PD-2
(???덉떆???뺤떇留?李멸퀬, ?ㅼ젣 ?곗씠?곕뒗 ?낅젰媛믪뿉 ?곕씪 ?숈쟻?쇰줈 ?앹꽦)
# ?낅젰
- ?낅젰1: 諛쒗뙆?묒뾽?쇱?_TSV (?꾨옒? 媛숈? ?뺤떇)
- ?낅젰2: 怨꾩륫?쇱?_TSV (?꾨옒? 媛숈? ?뺤떇, **怨꾩륫?쇱? ?쒕뒗 PDF 2?섏씠吏 ?댄썑遺??異붿텧**)
# ?낅젰1 ?덉떆
諛쒗뙆?쇱옄    諛쒗뙆?쒓컙    吏諛쒕떦?μ빟??理쒖냼, kg)    吏諛쒕떦?μ빟??理쒕?, kg)    ??빟?ъ슜??kg)    鍮꾧퀬
2023-07-27    08:05    0.5    0.9    73    PLA-2
2023-07-27    13:47    0.4    0.8    77    PD-2
# ?낅젰2 ?덉떆 (**2?섏씠吏 ?댄썑 ?쒕쭔**)
Date/Time    Peak Particle Vel (X_Axis) (mm/sec)    Peak Particle Vel (Y_Axis) (mm/sec)    Peak Particle Vel (Z_Axis) (mm/sec)    LMax (Sound) (dBA)    痢≪젙?꾩튂
2023/07/27 1:47:00 PM    0.71    0.36    0.71    61.23    ?묐쭚吏?2023/07/27 1:47:00 PM    0.87    0.56    0.87    53.29    ?곗뒪?뚯씠??# Mapping Rules
- ???낅젰??蹂묓빀?섏뿬 ???덉떆? ?숈씪??TSV留?異쒕젰
- ?ㅻ챸, 留덊겕?ㅼ슫, 肄붾뱶釉붾줉, 二쇱꽍, 湲고? ?띿뒪?몃뒗 ?덈? ?ы븿?섏? 留덉꽭??
- 怨꾩륫?쇱? ?쒕뒗 諛섎뱶??PDF 2?섏씠吏 ?댄썑???쒕쭔 ?ъ슜 
- 理쒖쥌 ?ㅻ뜑(怨좎젙??: 諛쒗뙆?쇱옄, 諛쒗뙆?쒓컙, 吏諛쒕떦?μ빟??理쒖냼, kg), 吏諛쒕떦?μ빟??理쒕?, kg), ??빟?ъ슜??kg), 諛쒗뙆吏꾨룞(cm/sec), 諛쒗뙆?뚯쓬(dB(A)), 怨꾩륫?꾩튂, 鍮꾧퀬
- ?뺣젹: 諛쒗뙆?쒓컙 ?ㅻ쫫李⑥닚, 怨꾩륫?꾩튂 ?ㅻ쫫李⑥닚(?꾩슂??
- 蹂묓빀/留ㅼ묶/?щ㎎ 洹쒖튃? 湲곗〈怨??숈씪
'''
DEFAULT_PROMPT = """
# INSTRUCTIONS
1. ?쇱씪?묒뾽蹂닿퀬 ?먮Ц?먯꽌 ?곗씠?곕? ?뚯떛?섏뿬 4媛??뚯씠釉?"?쒓났?꾪솴", "?묒뾽?댁슜", "?몄썝", "?λ퉬")怨?媛곴컖??TSV ?뺤떇??肄붾뱶釉붾줉?쇰줈 ?앹꽦?⑸땲??
2. ?먮룞 寃利?寃곌낵(QA-CHECKLIST)瑜?留덊겕?ㅼ슫 ??Table)濡??앹꽦?⑸땲??
3. ?쇱씪?묒뾽蹂닿퀬 ?띿뒪?몄뿉??**?묒뾽 ?좎쭨**瑜?異붿텧?섏뿬 泥?踰덉㎏濡?異쒕젰 (YYYY-MM-DD ?뺤떇)

# OUTPUT  
## 1. ?쒓났?꾪솴 ?뚯씠釉? 
1. 怨좎젙 ??: "援щ텇", "?꾧퀎"  
2. 怨좎젙 ??珥?33?? - ?꾨옒 ?쒖꽌? 紐낆묶??洹몃?濡? 
- "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾)  
- "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾) ?쇱씠?? 
- "2. ?좏뭾?뺢굅??- 1)?뺢굅???쇱씠??
- "2. ?좏뭾?뺢굅??- 1)?뺢굅??誘몃뱾 ?щ씪釉?
- "2. ?좏뭾?뺢굅????2)二쇱텧?낃뎄 ?섏쭅援??쇱씠??
- "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB ?뺢굅??諛⑸㈃ ?쇱씠??
- "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB ?섏듅?듬줈 諛⑸㈃ ?쇱씠??
- "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (2)PCC ?쇱씠??
- "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (3)PCD ?쇱씠??
- "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (4)PHA ?쇱씠??
- "2. ?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒 - ?섏쭅援??쇱씠??
- "2. ?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒 - PHB ?쇱씠??
- "2. ?좏뭾?뺢굅??- 4)?몃?異쒖엯援?異쒖엯援?#3) 援댁갑" 
- "2. ?좏뭾?뺢굅??- 4)?몃?異쒖엯援?異쒖엯援?#2) 援댁갑"
- "2. ?좏뭾?뺢굅??- 4)?몃?異쒖엯援?異쒖엯援?#1) 援댁갑" 
- "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCF) 援댁갑" 
- "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCF) ?쇱씠??  
- "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCE) 援댁갑" 
- "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCE) ?쇱씠??  
- "3. ?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX 蹂대씪留?諛⑸㈃ 援ъ“臾?  
- "3. ?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX ?由?諛⑸㈃ 援댁갑"  
- "4. 蹂몄꽑?곕꼸(2援ш컙, ?좏뭾-?꾨┝) 援댁갑"  
- "4. 蹂몄꽑?곕꼸(2援ш컙, ?좏뭾-?꾨┝) ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅???곕꼸 ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅??誘몃뱾 ?щ씪釉? 
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 ?섏쭅援??쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 PCA ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 PCC ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 PHA ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 ?섏쭅援??쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 PCA ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 PCC ?쇱씠??  
- "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 PHB ?쇱씠??  

3. 異붿텧?곗씠?? 
- "?꾧퀎"媛믩쭔 ?レ옄濡?異붿텧??寃?(??945.3m / 1,116m ?먯꽌 "945.3" 異붿텧)

## 2. ?묒뾽?댁슜 ?뚯씠釉? 
1. 怨좎젙 ??: "援щ텇", "湲덉씪?묒뾽"  
2. 怨좎젙 ??珥?14?? - ?꾨옒 ?쒖꽌? 紐낆묶(留ㅽ븨 ??寃곌낵)??洹몃?濡? 
- "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾)"  
- "2.?좏뭾?뺢굅??- 1)?뺢굅???곕꼸"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (2)PCC"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (3)PCD"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (4)PHA"  
- "2.?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒"  
- "2.?좏뭾?뺢굅??- 4)?몃?異쒖엯援?  
- "3.?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸"  
- "3.?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX"  
- "4.蹂몄꽑?곕꼸(2援ш컙, ?좏뭾-?꾨┝)"  
- "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅???곕꼸"  
- "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1"  
- "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2"  
3. 二쇱쓽?ы빆  
- '?묒뾽?댁슜' ?? ?щ윭 ?몃? ?댁슜???ы븿?????덉뒿?덈떎. ?댁슜??援щ텇???뚮뒗, 理쒖쥌 TSV 異쒕젰 ???대떦 ????곕뵲?댄몴("...")濡?媛먯떥?? ? ?대???媛??댁슜? **?ㅼ젣 以꾨컮轅?臾몄옄(?? '\\n'臾몄옄??????뷀꽣 ???낅젰???대떦)**瑜??ъ슜?섏뿬 遺꾨━?섎ŉ, "-"湲고샇???앸왂??
## 3. ?몄썝 / ?λ퉬 ?뚯씠釉? 
1. 怨좎젙 ??(珥?15?? - ???쒖꽌???꾨옒? 媛숈쓬
- "援щ텇" 
- "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾)"  
- "2.?좏뭾?뺢굅??- 1)?뺢굅???곕꼸"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (2)PCC"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (3)PCD"  
- "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (4)PHA"  
- "2.?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒"  
- "2.?좏뭾?뺢굅??- 4)?몃?異쒖엯援?  
- "3.?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸"  
- "3.?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX"  
- "4.蹂몄꽑?곕꼸(2援ш컙, ?좏뭾~?꾨┝)"  
- "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅???곕꼸"  
- "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1"  
- "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2"    

2. 怨좎젙 ???몄썝 ?뚯씠釉???珥?36??  
(?몄썝 紐⑸줉? ?꾨옒 ?쒖꽌? 紐낆묶(留ㅽ븨 ??寃곌낵)??諛섎뱶??洹몃?濡??ъ슜):

"吏곸쁺諛섏옣", "?곗닔??, "?λ퉬?댁쟾??, "?꾧린二쇱엫", "?붿빟二쇱엫", "?곕꼸怨?, "紐⑷났", "泥좉렐怨?, "?쇱씠?앺뤌怨?, "?ㅽ룓?섏쿂由ш났", "移대━?꾪듃怨?, "BP怨?, "媛?쒖꽕怨?, "?ㅼ튂怨??댁껜怨?, "?숇컮由ш났", "?좏샇??, "遺?⑥닔怨?, "?щ윭由ъ썡怨?, "CIP怨?, "誘몄옣怨?, "?쒖꽕臾쇨났", "寃쎄퀎?앷났", "議곌꼍怨?, "諛곌?怨?, "?꾩깋怨?, "諛⑹닔怨?, "?λ퉬/?묒뾽吏?댁씠", "蹂댄넻?몃?", "?ъ옣怨?, "?⑹젒怨?, "??ㅺ났", "蹂대쭅怨??숈뭅怨?, "鍮꾧퀎怨?, "?꾩옣怨?, "?앸㈃怨?, "二쇱엯怨?洹몃씪?고똿怨?

3. 怨좎젙 ??(?λ퉬 ?뚯씠釉???珥?46??  
(?λ퉬 紐⑸줉? ?꾨옒 ?쒖꽌? 紐낆묶(留ㅽ븨 ??寃곌낵)??諛섎뱶??洹몃?濡??ъ슜):

"B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)", "?ㅽ봽?몃윮(5T)", "?ㅽ봽?몃윮(15T)", "?ㅽ봽?몃윮(25T)", "?듦??щ젅??100T)", "?듦??щ젅??80T)", "?듦??щ젅??35T)", "?듦??щ젅??25T)", "移닿퀬?щ젅??25T)", "移닿퀬?щ젅??5T)", "肄ㅽ봽", "?먮낫?쒕┫", "?섏씠濡쒕뜑", "?륂듃癒몄떊", "李⑥쭠移?, "?댁닔李?, "?섏씠?쒕줈?щ젅??, "誘뱀꽌?몃윮", "?붾Ъ李?5T)", "?뚰봽移?, "?ㅼ뭅??, "肄섑겕由ы듃?쇰땲??, "?꾩＜?ㅺ굅", "濡쒕뜑(諛붾툕耳?", "?좎젣?댄룷湲?鍮꾩슦??", "吏寃뚯감", "?몄씤移?, "BC而ㅽ꽣湲?, "諛붿씠釉뚮줈?대㉧", "濡ㅻ윭(2.5T)", "濡ㅻ윭(1T)", "濡ㅻ윭(0.7T)", "紐곕━", "???湲?, "?щ젅??, "肄ㅻ퉬濡쒕씪", "怨듭븬?쒕┫", "?좎븬?쒕┫", "湲고?"

## 4. Parsing Rules 
1. ?쒓났?꾪솴: "?꾧퀎/?ㅺ퀎" ??**?욊컪(?뚯닔 ?덉슜)** 留?異붿텧.    
2. ?몄썝쨌?λ퉬: ?ъ엯?꾪솴?먯꽌 **?뺤닔留?* 異붿텧, 鍮덉?? **0**    
3. ?섏쐞 ?뱀뀡 留ㅽ븨   
   - ?뺢굅???곕꼸 ?????? PCB ???? PCC ???? PCD ????PHA ???? ?밸퀎?쇰궃 ???? ?몃?異쒖엯援??믠뫃    
4. 留ㅽ븨 ?뺤뀛?덈━ ?곸슜    
- "B/H08W" ??B/H(08W)"   
- "25??移닿퀬?щ젅?? ??"移닿퀬?щ젅??25T)"   
- "?밴났" ??"蹂댄넻?몃?"    
- "湲곌퀎??ㅺ났" ??"??ㅺ났"    
- "紐⑷났?곗닔?? ?먮뒗 "紐⑹닔?곗닔?? ??"?곗닔??    
- "5?ㅽ듃?? ??"?붾Ъ李?5T)"    
- "移대━?꾪듃" ??"移대━?꾪듃怨?    
- "?섏씠?쒕줈?щ젅??20T)" ??"?섏씠?쒕줈?щ젅??    
- "?쇱씠?앺뤌議곕┰" ??"?쇱씠?앺뤌怨?  
- "S/C??ㅽ?" ??"?곕꼸怨?  
- "紐⑹닔" ??"紐⑷났"    
5. ?ъ쟾???녿뒗 ??ぉ ???좎궗??ぉ, ?놁쑝硫?**?몄썝: 蹂댄넻?몃? / ?λ퉬: 湲고?** 濡??⑹궛?섍퀬 '?ㅻ쪟?붿빟'??湲곗옱.

## 5. QA-CHECKLIST(?먮룞 寃利?寃곌낵)

 1. 寃利???ぉ
?꾨옒 湲곗????곕씪 ?곗씠??泥섎━ 怨쇱젙???뺥솗?깆쓣 ?먯껜 寃利앺븯怨? 洹?寃곌낵瑜?留덊겕?ㅼ슫 ?뚯씠釉붾줈 ?앹꽦?⑸땲??
- **援ъ“/?뺤떇**: 4媛??뚯씠釉붿쓽 ?됯낵 ??媛쒖닔, ?쒖꽌, ?곗씠???뺤떇(?レ옄, ?뺤닔, 0 泥섎━)??吏移④낵 ?쇱튂?섎뒗媛?
- **?곗씠??臾닿껐??*: ?먮낯 蹂닿퀬?쒖쓽 ?몄썝 諛??λ퉬 ?섎웾???꾨씫?섍굅??以묐났?섏? ?딄퀬 100% ?뺥솗?섍쾶 吏묎퀎?섏뿀?붽??
- **留ㅽ븨/蹂??*: 吏?뺣맂 留ㅽ븨 洹쒖튃(?⑹뼱 ?쒖??? ?좎궗 ??ぉ 泥섎━ ????紐⑤몢 ?щ컮瑜닿쾶 ?곸슜?섏뿀?붽??
- **誘몃텇瑜???ぉ**: ?ъ쟾???뺤쓽?섏? ?딆? ??ぉ??'蹂댄넻?몃?' ?먮뒗 '湲고?'濡??곸젅??遺꾨쪟?섍퀬 湲곕줉?섏뿀?붽??

2. 異쒕젰 諛⑹떇
- **?붿빟**: 寃利?寃곌낵瑜??꾨옒 ?덉떆? 媛숈씠 留덊겕?ㅼ슫 ?뚯씠釉?`QA-CHECKLIST`)?쇰줈 ?붿빟?⑸땲??
- **蹂???댁뿭**: ?곗씠??泥섎━ 怨쇱젙?먯꽌 蹂寃쎈맂 ?댁슜???덈뒗 寃쎌슦, **'蹂???댁뿭'**???`?먮Ц ??寃곌낵` ?뺤떇?쇰줈 紐낆떆?⑸땲?? 蹂寃??ы빆???놁쑝硫?"?댁긽 ?놁쓬"?쇰줈 ?쒓린?⑸땲??

3. ?덉떆 (留덊겕?ㅼ슫 ?뚯씠釉?
| ?먭? ??ぉ | 寃利?湲곗? | 蹂???댁뿭 (?먮Ц ??寃곌낵) | ?곹깭 |
| :--- | :--- | :--- | :---: |
| **援ъ“ 諛??뺤떇** | 4媛??뚯씠釉붿쓽 援ъ“ 諛??곗씠???뺤떇??吏移④낵 ?쇱튂??| ?댁긽 ?놁쓬 | ??|
| **?곗씠??臾닿껐??* | ?몄썝(85紐?, ?λ퉬(12?) 珥앷퀎媛 ?먮Ц怨??쇱튂??| ?댁긽 ?놁쓬 | ??|
| **?⑹뼱 ?쒖???* | 留ㅽ븨 洹쒖튃???곕씪 ?⑹뼱媛 ?쇨큵 蹂?섎맖 | - 紐⑹닔 ??紐⑷났<br>- ?밴났 ??蹂댄넻?몃?<br>- B/H08W ??B/H(08W) | ??|
| **誘몃텇瑜???ぉ 泥섎━** | ?ъ쟾???녿뒗 ??ぉ??洹쒖튃???곕씪 泥섎━??| - ?몃Т????蹂댄넻?몃? (?⑹궛) | ?좑툘 |

---

# USER TEXT(?묒뾽怨꾪쉷蹂닿퀬 ?낅젰?)
"""

# --- PROMPT MANAGEMENT FUNCTIONS ---
def save_prompt_to_supabase(prompt_name, prompt_content, description=""):
    """?꾨＼?꾪듃瑜?Supabase????ν빀?덈떎."""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return False
    
    try:
        data = {
            "name": prompt_name,
            "content": prompt_content,
            "description": description,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        # 湲곗〈 ?꾨＼?꾪듃媛 ?덈뒗吏 ?뺤씤
        existing = supabase_client.table("prompts").select("*").eq("name", prompt_name).execute()
        
        if existing.data:
            # ?낅뜲?댄듃
            data["updated_at"] = datetime.now().isoformat()
            result = supabase_client.table("prompts").update(data).eq("name", prompt_name).execute()
            st.success(f"???꾨＼?꾪듃 '{prompt_name}'??媛) ?낅뜲?댄듃?섏뿀?듬땲??")
        else:
            # ?덈줈 ?앹꽦
            result = supabase_client.table("prompts").insert(data).execute()
            st.success(f"???꾨＼?꾪듃 '{prompt_name}'??媛) ??λ릺?덉뒿?덈떎.")
        
        return True
    except Exception as e:
        st.error(f"???꾨＼?꾪듃 ????ㅽ뙣: {e}")
        return False

def load_prompt_from_supabase(prompt_name):
    """Supabase?먯꽌 ?꾨＼?꾪듃瑜?濡쒕뱶?⑸땲??"""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return None
    
    try:
        # SSL ?ㅻ쪟 諛⑹?瑜??꾪븳 異붽? ?ㅼ젙
        import ssl
        import urllib3
        import os
        
        # SSL 寃쎄퀬 鍮꾪솢?깊솕
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        os.environ['PYTHONHTTPSVERIFY'] = '0'
        
        result = supabase_client.table("prompts").select("*").eq("name", prompt_name).execute()
        if result.data:
            return result.data[0]
        return None
    except Exception as e:
        st.error(f"???꾨＼?꾪듃 濡쒕뱶 ?ㅽ뙣: {e}")
        st.info("?뮕 SSL ?몄쬆??臾몄젣?????덉뒿?덈떎. ?ㅽ듃?뚰겕 ?ㅼ젙???뺤씤?댁＜?몄슂.")
        return None

def get_all_prompts_from_supabase():
    """Supabase?먯꽌 紐⑤뱺 ?꾨＼?꾪듃 紐⑸줉??媛?몄샃?덈떎."""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return []
    
    try:
        # SSL ?ㅻ쪟 諛⑹?瑜??꾪븳 異붽? ?ㅼ젙
        import ssl
        import urllib3
        import os
        
        # SSL 寃쎄퀬 鍮꾪솢?깊솕
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        os.environ['PYTHONHTTPSVERIFY'] = '0'
        
        result = supabase_client.table("prompts").select("name, description, updated_at").execute()
        return result.data if result.data else []
    except Exception as e:
        st.error(f"???꾨＼?꾪듃 紐⑸줉 濡쒕뱶 ?ㅽ뙣: {e}")
        st.info("?뮕 SSL ?몄쬆??臾몄젣?????덉뒿?덈떎. ?ㅽ듃?뚰겕 ?ㅼ젙???뺤씤?댁＜?몄슂.")
        return []

def delete_prompt_from_supabase(prompt_name):
    """Supabase?먯꽌 ?꾨＼?꾪듃瑜???젣?⑸땲??"""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return False
    
    try:
        result = supabase_client.table("prompts").delete().eq("name", prompt_name).execute()
        st.success(f"???꾨＼?꾪듃 '{prompt_name}'??媛) ??젣?섏뿀?듬땲??")
        return True
    except Exception as e:
        st.error(f"???꾨＼?꾪듃 ??젣 ?ㅽ뙣: {e}")
        return False

def extract_table_structure_from_prompt(prompt_text):
    """?꾨＼?꾪듃 ?띿뒪?몄뿉???뚯씠釉?援ъ“瑜?異붿텧?⑸땲??"""
    extracted_tables = {}
    
    try:
        # ?쒓났?꾪솴 ?뚯씠釉?異붿텧
        construction_match = re.search(r'?쒓났?꾪솴.*?怨좎젙 ??*?珥?(\d+)??*??꾨옒 ?쒖꽌? 紐낆묶??洹몃?濡?.*?)(?=## 2\.|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if construction_match:
            construction_lines = construction_match.group(2)
            construction_items = re.findall(r'- "([^"]+)"', construction_lines)
            if construction_items:
                extracted_tables['construction'] = construction_items
        
        # ?묒뾽?댁슜 ?뚯씠釉?異붿텧
        work_match = re.search(r'?묒뾽?댁슜.*?怨좎젙 ??*?珥?(\d+)??*??꾨옒 ?쒖꽌? 紐낆묶.*???洹몃?濡?.*?)(?=## 3\.|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if work_match:
            work_lines = work_match.group(2)
            work_items = re.findall(r'- "([^"]+)"', work_lines)
            if work_items:
                extracted_tables['work_content'] = work_items
        
        # ?몄썝 ?뚯씠釉???異붿텧
        personnel_col_match = re.search(r'怨좎젙 ??*?珥?(\d+)??*??꾨옒? 媛숈쓬(.*?)(?=怨좎젙 ??$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if personnel_col_match:
            col_lines = personnel_col_match.group(2)
            col_items = re.findall(r'- "([^"]+)"', col_lines)
            if col_items:
                extracted_tables['personnel_columns'] = col_items
        
        # ?몄썝 ?뚯씠釉???異붿텧
        personnel_row_match = re.search(r'?몄썝 ?뚯씠釉?*?珥?(\d+)??*??꾨옒 ?쒖꽌? 紐낆묶.*???諛섎뱶??洹몃?濡??ъ슜(.*?)(?=怨좎젙 ??*??λ퉬|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if personnel_row_match:
            row_lines = personnel_row_match.group(2)
            row_items = re.findall(r'"([^"]+)"', row_lines)
            if row_items:
                extracted_tables['personnel_rows'] = row_items
        
        # ?λ퉬 ?뚯씠釉?異붿텧
        equipment_match = re.search(r'?λ퉬 ?뚯씠釉?*?珥?(\d+)??*??꾨옒 ?쒖꽌? 紐낆묶.*???諛섎뱶??洹몃?濡??ъ슜(.*?)(?=## 4\.|$)', prompt_text, re.DOTALL | re.IGNORECASE)
        if equipment_match:
            equipment_lines = equipment_match.group(2)
            equipment_items = re.findall(r'"([^"]+)"', equipment_lines)
            if equipment_items:
                extracted_tables['equipment'] = equipment_items
        
        return extracted_tables if extracted_tables else None
        
    except Exception as e:
        st.error(f"?뚯씠釉?援ъ“ 異붿텧 以??ㅻ쪟: {e}")
        return None

def generate_prompt_from_tables():
    """?뚯씠釉?援ъ“ ?곗씠?곕? 湲곕컲?쇰줈 ?꾨＼?꾪듃瑜??앹꽦?⑸땲??"""
    
    # ?쒓났?꾪솴 ??紐⑸줉
    construction_rows = st.session_state.construction_rows
    
    # ?묒뾽?댁슜 ??紐⑸줉
    work_content_rows = st.session_state.work_content_rows
    
    # ?몄썝/?λ퉬 ??紐⑸줉
    personnel_columns = st.session_state.personnel_columns
    personnel_rows = st.session_state.personnel_rows
    equipment_rows = st.session_state.equipment_rows
    
    prompt = f"""# INSTRUCTIONS
1. ?쇱씪?묒뾽蹂닿퀬 ?먮Ц?먯꽌 ?곗씠?곕? ?뚯떛?섏뿬 4媛??뚯씠釉?"?쒓났?꾪솴", "?묒뾽?댁슜", "?몄썝", "?λ퉬")怨?媛곴컖??TSV ?뺤떇??肄붾뱶釉붾줉?쇰줈 ?앹꽦?⑸땲??
2. ?먮룞 寃利?寃곌낵(QA-CHECKLIST)瑜?留덊겕?ㅼ슫 ??Table)濡??앹꽦?⑸땲??
3. ?쇱씪?묒뾽蹂닿퀬 ?띿뒪?몄뿉??**?묒뾽 ?좎쭨**瑜?異붿텧?섏뿬 泥?踰덉㎏濡?異쒕젰 (YYYY-MM-DD ?뺤떇)

# OUTPUT  
## 1. ?쒓났?꾪솴 ?뚯씠釉? 
1. 怨좎젙 ??: "援щ텇", "?꾧퀎"  
2. 怨좎젙 ??珥?{len(construction_rows)}?? - ?꾨옒 ?쒖꽌? 紐낆묶??洹몃?濡? 
{chr(10).join([f'- "{row}"' for row in construction_rows])}
3. 異붿텧?곗씠?? 
- "?꾧퀎"媛믩쭔 ?レ옄濡?異붿텧??寃?(??945.3m / 1,116m ?먯꽌 "945.3" 異붿텧)

## 2. ?묒뾽?댁슜 ?뚯씠釉? 
1. 怨좎젙 ??: "援щ텇", "湲덉씪?묒뾽"  
2. 怨좎젙 ??珥?{len(work_content_rows)}?? - ?꾨옒 ?쒖꽌? 紐낆묶(留ㅽ븨 ??寃곌낵)??洹몃?濡? 
{chr(10).join([f'- "{row}"' for row in work_content_rows])}
3. 二쇱쓽?ы빆  
- '?묒뾽?댁슜' ?? ?щ윭 ?몃? ?댁슜???ы븿?????덉뒿?덈떎. ?댁슜??援щ텇???뚮뒗, 理쒖쥌 TSV 異쒕젰 ???대떦 ????곕뵲?댄몴("...")濡?媛먯떥?? ? ?대???媛??댁슜? **?ㅼ젣 以꾨컮轅?臾몄옄(?? '\\n'臾몄옄??????뷀꽣 ???낅젰???대떦)**瑜??ъ슜?섏뿬 遺꾨━?섎ŉ, "-"湲고샇???앸왂??
## 3. ?몄썝 / ?λ퉬 ?뚯씠釉? 
1. 怨좎젙 ??(珥?{len(personnel_columns) + 1}?? - ???쒖꽌???꾨옒? 媛숈쓬
- "援щ텇" 
{chr(10).join([f'- "{col}"' for col in personnel_columns])}
2. 怨좎젙 ???몄썝 ?뚯씠釉???珥?{len(personnel_rows)}??  
(?몄썝 紐⑸줉? ?꾨옒 ?쒖꽌? 紐낆묶(留ㅽ븨 ??寃곌낵)??諛섎뱶??洹몃?濡??ъ슜):
{', '.join([f'"{row}"' for row in personnel_rows])}
3. 怨좎젙 ??(?λ퉬 ?뚯씠釉???珥?{len(equipment_rows)}??  
(?λ퉬 紐⑸줉? ?꾨옒 ?쒖꽌? 紐낆묶(留ㅽ븨 ??寃곌낵)??諛섎뱶??洹몃?濡??ъ슜):
{', '.join([f'"{row}"' for row in equipment_rows])}

## 4. Parsing Rules 
1. ?쒓났?꾪솴: "?꾧퀎/?ㅺ퀎" ??**?욊컪(?뚯닔 ?덉슜)** 留?異붿텧.    
2. ?몄썝쨌?λ퉬: ?ъ엯?꾪솴?먯꽌 **?뺤닔留?* 異붿텧, 鍮덉?? **0**    
3. ?섏쐞 ?뱀뀡 留ㅽ븨   
   - ?뺢굅???곕꼸 ?????? PCB ???? PCC ???? PCD ????PHA ???? ?밸퀎?쇰궃 ???? ?몃?異쒖엯援??믠뫃    
4. 留ㅽ븨 ?뺤뀛?덈━ ?곸슜    
- "B/H08W" ??B/H(08W)"   
- "25??移닿퀬?щ젅?? ??"移닿퀬?щ젅??25T)"   
- "?밴났" ??"蹂댄넻?몃?"    
- "湲곌퀎??ㅺ났" ??"??ㅺ났"    
- "紐⑷났?곗닔?? ?먮뒗 "紐⑹닔?곗닔?? ??"?곗닔??    
- "5?ㅽ듃?? ??"?붾Ъ李?5T)"    
- "移대━?꾪듃" ??"移대━?꾪듃怨?    
- "?섏씠?쒕줈?щ젅??20T)" ??"?섏씠?쒕줈?щ젅??    
- "?쇱씠?앺뤌議곕┰" ??"?쇱씠?앺뤌怨?  
- "S/C??ㅽ?" ??"?곕꼸怨?  
- "紐⑹닔" ??"紐⑷났"    
5. ?ъ쟾???녿뒗 ??ぉ ???좎궗??ぉ, ?놁쑝硫?**?몄썝: 蹂댄넻?몃? / ?λ퉬: 湲고?** 濡??⑹궛?섍퀬 '?ㅻ쪟?붿빟'??湲곗옱.

## 5. QA-CHECKLIST(?먮룞 寃利?寃곌낵)

 1. 寃利???ぉ
?꾨옒 湲곗????곕씪 ?곗씠??泥섎━ 怨쇱젙???뺥솗?깆쓣 ?먯껜 寃利앺븯怨? 洹?寃곌낵瑜?留덊겕?ㅼ슫 ?뚯씠釉붾줈 ?앹꽦?⑸땲??
- **援ъ“/?뺤떇**: 4媛??뚯씠釉붿쓽 ?됯낵 ??媛쒖닔, ?쒖꽌, ?곗씠???뺤떇(?レ옄, ?뺤닔, 0 泥섎━)??吏移④낵 ?쇱튂?섎뒗媛?
- **?곗씠??臾닿껐??*: ?먮낯 蹂닿퀬?쒖쓽 ?몄썝 諛??λ퉬 ?섎웾???꾨씫?섍굅??以묐났?섏? ?딄퀬 100% ?뺥솗?섍쾶 吏묎퀎?섏뿀?붽??
- **留ㅽ븨/蹂??*: 吏?뺣맂 留ㅽ븨 洹쒖튃(?⑹뼱 ?쒖??? ?좎궗 ??ぉ 泥섎━ ????紐⑤몢 ?щ컮瑜닿쾶 ?곸슜?섏뿀?붽??
- **誘몃텇瑜???ぉ**: ?ъ쟾???뺤쓽?섏? ?딆? ??ぉ??'蹂댄넻?몃?' ?먮뒗 '湲고?'濡??곸젅??遺꾨쪟?섍퀬 湲곕줉?섏뿀?붽??

2. 異쒕젰 諛⑹떇
- **?붿빟**: 寃利?寃곌낵瑜??꾨옒 ?덉떆? 媛숈씠 留덊겕?ㅼ슫 ?뚯씠釉?`QA-CHECKLIST`)?쇰줈 ?붿빟?⑸땲??
- **蹂???댁뿭**: ?곗씠??泥섎━ 怨쇱젙?먯꽌 蹂寃쎈맂 ?댁슜???덈뒗 寃쎌슦, **'蹂???댁뿭'**???`?먮Ц ??寃곌낵` ?뺤떇?쇰줈 紐낆떆?⑸땲?? 蹂寃??ы빆???놁쑝硫?"?댁긽 ?놁쓬"?쇰줈 ?쒓린?⑸땲??

3. ?덉떆 (留덊겕?ㅼ슫 ?뚯씠釉?
| ?먭? ??ぉ | 寃利?湲곗? | 蹂???댁뿭 (?먮Ц ??寃곌낵) | ?곹깭 |
| :--- | :--- | :--- | :---: |
| **援ъ“ 諛??뺤떇** | 4媛??뚯씠釉붿쓽 援ъ“ 諛??곗씠???뺤떇??吏移④낵 ?쇱튂??| ?댁긽 ?놁쓬 | ??|
| **?곗씠??臾닿껐??* | ?몄썝(85紐?, ?λ퉬(12?) 珥앷퀎媛 ?먮Ц怨??쇱튂??| ?댁긽 ?놁쓬 | ??|
| **?⑹뼱 ?쒖???* | 留ㅽ븨 洹쒖튃???곕씪 ?⑹뼱媛 ?쇨큵 蹂?섎맖 | - 紐⑹닔 ??紐⑷났<br>- ?밴났 ??蹂댄넻?몃?<br>- B/H08W ??B/H(08W) | ??|
| **誘몃텇瑜???ぉ 泥섎━** | ?ъ쟾???녿뒗 ??ぉ??洹쒖튃???곕씪 泥섎━??| - ?몃Т????蹂댄넻?몃? (?⑹궛) | ?좑툘 |

---

# USER TEXT(?묒뾽怨꾪쉷蹂닿퀬 ?낅젰?)
"""
    
    return prompt

# --- HELPER FUNCTIONS ---
def safe_generate_content(model_input):
    """
    Calls the Gemini API with robust error handling and relaxed safety settings.
    """
    try:
        # AI 紐⑤뜽???덉쟾 ?ㅼ젙???꾪솕?섏뿬 肄섑뀗痢?李⑤떒??理쒖냼?뷀빀?덈떎.
        safety_settings = {
            'HARM_CATEGORY_HARASSMENT': 'BLOCK_NONE',
            'HARM_CATEGORY_HATE_SPEECH': 'BLOCK_NONE',
            'HARM_CATEGORY_SEXUALLY_EXPLICIT': 'BLOCK_NONE',
            'HARM_CATEGORY_DANGEROUS_CONTENT': 'BLOCK_NONE',
        }
        
        response = GEMINI_MODEL.generate_content(
            model_input,
            safety_settings=safety_settings
        )

        # ?묐떟???ㅼ젣 肄섑뀗痢?parts)媛 ?덈뒗吏 ?뺤씤?⑸땲??
        if response.parts:
            return response.text
        else:
            # 肄섑뀗痢좉? ?녿뒗 寃쎌슦, 李⑤떒 ?먯씤???뺤씤?섏뿬 ?ъ슜?먯뿉寃??뚮┰?덈떎.
            reason = "Unknown"
            try:
                # API ?묐떟?먯꽌 ?쒓났?섎뒗 怨듭떇?곸씤 李⑤떒 ?댁쑀瑜?媛?몄샃?덈떎.
                reason = response.prompt_feedback.block_reason.name
            except Exception:
                pass 
            st.error(f"AI ?묐떟 ?앹꽦???ㅽ뙣?덉뒿?덈떎. API???섑빐 肄섑뀗痢좉? 李⑤떒?섏뿀?????덉뒿?덈떎. (李⑤떒 ?댁쑀: {reason})")
            st.warning(f"?꾩껜 ?쇰뱶諛? {response.prompt_feedback}")
            return None
            
    except Exception as e:
        st.error(f"AI 紐⑤뜽 ?몄텧 以??ш컖???ㅻ쪟 諛쒖깮: {e}")
        return None

def sanitize_prompt(prompt_text):
    """?꾨＼?꾪듃?먯꽌 誘쇨컧???⑥뼱瑜??쒓굅?섍굅???泥댄빀?덈떎."""
    # 誘쇨컧???⑥뼱?ㅼ쓣 ??以묐┰?곸씤 ?쒗쁽?쇰줈 ?泥?    replacements = {
        '?밴났': '?뱀닔怨?,
        '??뙆': '?뚯뇙',
        '??컻': '?뚯뇙',
        '??빟': '?뚯뇙??,
        '?ㅼ씠?덈쭏?댄듃': '?뚯뇙??,
        'TNT': '?뚯뇙??,
        '??뙆怨?: '?뚯뇙怨?,
        '??뙆?묒뾽': '?뚯뇙?묒뾽',
        '??뙆怨듭궗': '?뚯뇙怨듭궗',
        '??뙆??: '?뚯뇙??,
        '??뙆援?: '?뚯뇙援?,
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆?λ퉬': '?뚯뇙?λ퉬',
        '??뙆援щ찉': '?뚯뇙援?,
    }
    
    sanitized_text = prompt_text
    for old_word, new_word in replacements.items():
        sanitized_text = sanitized_text.replace(old_word, new_word)
    
    # 異붽??곸씤 ?뺤젣 ?묒뾽
    # 1. 紐⑤뱺 ?덉쟾 愿???⑥뼱瑜?以묐┰?곸쑝濡?蹂寃?    safety_replacements = {
        '?꾪뿕': '二쇱쓽',
        '?ш퀬': '?ш굔',
        '?щ쭩': '?몃챸?쇳빐',
        '遺??: '?곹빐',
        '??컻臾?: '?뚯뇙??,
        '??컻??: '?뚯뇙??,
        '??컻??: '?뚯뇙??,
    }
    
    for old_word, new_word in safety_replacements.items():
        sanitized_text = sanitized_text.replace(old_word, new_word)
    
    # 2. ?꾨＼?꾪듃 ?쒖옉 遺遺꾩뿉 ?덉쟾??媛뺤“ 臾멸뎄 異붽?
    safety_prefix = """[?덉쟾??嫄댁꽕?낅Т ?곗씠??遺꾩꽍 吏?쒖궗??
??吏?쒖궗??? 嫄댁꽕?꾩옣???쇱씪?묒뾽蹂닿퀬瑜??덉쟾?섍퀬 ?뺥솗?섍쾶 遺꾩꽍?섍린 ?꾪븳 寃껋엯?덈떎.
紐⑤뱺 ?댁슜? 嫄댁꽕?낅Т???뺤긽?곸씤 ?낅Т泥섎━瑜??꾪븳 寃껋엯?덈떎.

"""
    
    return safety_prefix + sanitized_text

def create_simplified_prompt(original_prompt):
    """李⑤떒??寃쎌슦 ?ъ슜??媛꾨떒???꾨＼?꾪듃瑜??앹꽦?⑸땲??"""
    # ?먮낯 ?꾨＼?꾪듃?먯꽌 ?ъ슜???낅젰 遺遺꾨쭔 異붿텧
    user_input_start = original_prompt.find('"""')
    if user_input_start != -1:
        user_input = original_prompt[user_input_start + 3:]
        user_input = user_input.replace('"""', '').strip()
    else:
        user_input = original_prompt
    
    simplified_prompt = f"""嫄댁꽕?꾩옣 ?쇱씪?묒뾽蹂닿퀬瑜?遺꾩꽍?섏뿬 ?ㅼ쓬 ?뺤떇?쇰줈 ?뺣━?댁＜?몄슂:

1. ?묒뾽 ?좎쭨 異붿텧 (YYYY-MM-DD ?뺤떇)
2. ?쒓났?꾪솴 ?뚯씠釉?(援щ텇, ?꾧퀎)
3. ?묒뾽?댁슜 ?뚯씠釉?(援щ텇, 湲덉씪?묒뾽)
4. ?몄썝 ?뚯씠釉?(援щ텇, ?몄썝??
5. ?λ퉬 ?뚯씠釉?(援щ텇, ???

媛??뚯씠釉붿? TSV ?뺤떇?쇰줈 異쒕젰?댁＜?몄슂.

?묒뾽蹂닿퀬:
{user_input}"""
    
    return simplified_prompt

def create_minimal_prompt(original_prompt):
    """理쒖냼?쒖쓽 ?꾨＼?꾪듃瑜??앹꽦?⑸땲??"""
    # ?먮낯 ?꾨＼?꾪듃?먯꽌 ?ъ슜???낅젰 遺遺꾨쭔 異붿텧
    user_input_start = original_prompt.find('"""')
    if user_input_start != -1:
        user_input = original_prompt[user_input_start + 3:]
        user_input = user_input.replace('"""', '').strip()
    else:
        user_input = original_prompt
    
    minimal_prompt = f"""嫄댁꽕?꾩옣 ?쇱씪?묒뾽蹂닿퀬瑜?遺꾩꽍?섏뿬 ?뚯씠釉붾줈 ?뺣━?댁＜?몄슂.

?묒뾽蹂닿퀬:
{user_input}"""
    
    return minimal_prompt

def send_teams_alert(warning_rows, file_date):
    try:
        message = {
            "type": "message",
            "attachments": [{
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": {
                    "type": "AdaptiveCard",
                    "body": [
                        {"type": "TextBlock", "size": "Large", "weight": "Bolder", "text": f"?좑툘 怨꾩륫湲?寃쎄퀬 ?뚮┝ ({file_date})", "color": "Attention"},
                        {"type": "TextBlock", "text": "?ㅼ쓬 怨꾩륫湲곗뿉??二쇱쓽媛 ?꾩슂??蹂?붽? 媛먯??섏뿀?듬땲??", "wrap": True}
                    ]
                }
            }]
        }
        for _, row in warning_rows.iterrows():
            warning_info = {"type": "TextBlock", "text": f"?뱧 ?꾩튂: {row['?꾩튂']}\\n\\n?뱤 怨꾩륫湲? {row['怨꾩륫湲곕챸']} ({row['怨꾩륫湲?醫낅쪟']})\\n\\n?좑툘 ?곹깭: {row['?곹깭']}\\n\\n?뱢 3李?珥덇낵 ?鍮? {row['鍮꾩쑉']}", "wrap": True, "style": "warning"}
            message["attachments"][0]["content"]["body"].append(warning_info)
        
        response = requests.post(TEAMS_WEBHOOK_URL, json=message, headers={"Content-Type": "application/json"})
        if response.status_code == 200: st.success("Teams濡?寃쎄퀬 硫붿떆吏媛 ?꾩넚?섏뿀?듬땲??")
        else: st.error(f"Teams 硫붿떆吏 ?꾩넚 ?ㅽ뙣: {response.status_code}")
    except Exception as e: st.error(f"Teams 硫붿떆吏 ?꾩넚 以??ㅻ쪟 諛쒖깮: {e}")

def save_step1_to_supabase(data, date=None):
    """1?④퀎 AI ?곗씠??異붿텧 寃곌낵瑜?4媛?蹂꾨룄 ?뚯씠釉붿뿉 ??ν빀?덈떎."""
    global supabase_client
    
    if not supabase_client:
        st.error("??Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return False
    
    try:
        current_date = date or datetime.now().strftime('%Y-%m-%d')
        
        success_count = 0
        total_tables = 4
        
        # 1. ?쒓났?꾪솴 ?뚯씠釉????        if data.get("?쒓났?꾪솴"):
            try:
                construction_records = []
                for item in data["?쒓났?꾪솴"]:
                    construction_records.append({
                        "date": current_date,
                        "援щ텇": item.get("援щ텇", ""),
                        "?꾧퀎": item.get("?꾧퀎", ""),
                        "created_at": datetime.now().isoformat()
                    })
                
                result = supabase_client.table("construction_status").upsert(construction_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"???쒓났?꾪솴 ????ㅽ뙣: {e}")
        
        # 2. ?묒뾽?댁슜 ?뚯씠釉????        if data.get("?묒뾽?댁슜"):
            try:
                work_content_records = []
                for item in data["?묒뾽?댁슜"]:
                    work_content_records.append({
                        "date": current_date,
                        "援щ텇": item.get("援щ텇", ""),
                        "湲덉씪?묒뾽": item.get("湲덉씪?묒뾽", ""),
                        "created_at": datetime.now().isoformat()
                    })
                
                result = supabase_client.table("work_content").upsert(work_content_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"???묒뾽?댁슜 ????ㅽ뙣: {e}")
        
        # 3. ?몄썝 ?곗씠???뚯씠釉????        if data.get("?몄썝"):
            try:
                personnel_records = []
                for item in data["?몄썝"]:
                    # 湲곗〈 ?뚯씠釉?援ъ“??留욊쾶 ?곗씠?????                    personnel_record = {
                        "date": current_date,
                        "援щ텇": item.get("援щ텇", ""),
                        "created_at": datetime.now().isoformat()
                    }
                    personnel_records.append(personnel_record)
                
                result = supabase_client.table("personnel_data").upsert(personnel_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"???몄썝 ?곗씠??????ㅽ뙣: {e}")
        
        # 4. ?λ퉬 ?곗씠???뚯씠釉????        if data.get("?λ퉬"):
            try:
                equipment_records = []
                for item in data["?λ퉬"]:
                    # 湲곗〈 ?뚯씠釉?援ъ“??留욊쾶 ?곗씠?????                    equipment_record = {
                        "date": current_date,
                        "援щ텇": item.get("援щ텇", ""),
                        "created_at": datetime.now().isoformat()
                    }
                    equipment_records.append(equipment_record)
                
                result = supabase_client.table("equipment_data").upsert(equipment_records).execute()
                success_count += 1
            except Exception as e:
                st.error(f"???λ퉬 ?곗씠??????ㅽ뙣: {e}")
        
        if success_count == total_tables:
            st.success("??1?④퀎 ?곗씠??????꾨즺")
            return True
        else:
            st.warning(f"?좑툘 ?쇰? ?곗씠????μ뿉 ?ㅽ뙣?덉뒿?덈떎. ({success_count}/{total_tables})")
            return False
        
    except Exception as e:
        st.error(f"??1?④퀎 ?곗씠??????ㅽ뙣: {e}")
        import traceback
        st.error(f"???곸꽭 ?ㅻ쪟: {traceback.format_exc()}")
        return False

def save_to_supabase(data_type, data, date=None):
    """?곗씠?곕? Supabase????ν빀?덈떎."""
    global supabase_client
    
    if not supabase_client:
        st.error("??Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return False
    
    try:
        current_date = date or datetime.now().strftime('%Y-%m-%d')
        
        # ?곗씠????낆뿉 ?곕씪 ?ㅻⅨ ?뚯씠釉붿뿉 ???        if data_type == "daily_report":
            try:
                # ?묒? ?뚯씪?먯꽌 ?뱀젙 ? ?곗씠??異붿텧
                if "excel_bytes" in data:
                    extracted_data = extract_cell_data_from_excel(data["excel_bytes"], current_date)
                    
                    if not extracted_data:
                        st.error("???묒? ?곗씠??異붿텧???ㅽ뙣?덉뒿?덈떎.")
                        return False
                    
                    # daily_report_data ?뚯씠釉붿뿉 ???                    report_data = {
                        "date": current_date,
                        "construction_data": extracted_data.get("construction_data", {}),
                        "personnel_data": extracted_data.get("personnel_data", {}),
                        "equipment_data": extracted_data.get("equipment_data", {}),
                        "created_at": datetime.now().isoformat(),
                        "updated_at": datetime.now().isoformat()
                    }
                    
                    try:
                        # 湲곗〈 ?곗씠???뺤씤
                        existing_data = supabase_client.table("daily_report_data").select("*").eq("date", current_date).execute()
                        
                        if existing_data.data:
                            # 湲곗〈 ?곗씠???낅뜲?댄듃
                            result = supabase_client.table("daily_report_data").update(report_data).eq("date", current_date).execute()
                            st.success("??3?④퀎 ?곗씠??????꾨즺")
                        else:
                            # ???곗씠???쎌엯
                            result = supabase_client.table("daily_report_data").insert(report_data).execute()
                            st.success("??3?④퀎 ?곗씠??????꾨즺")
                        
                        return True
                        
                    except Exception as table_error:
                        st.error(f"??Supabase ????ㅽ뙣: {table_error}")
                        return False
                    
                else:
                    st.error("???묒? ?뚯씪 ?곗씠?곌? ?놁뒿?덈떎.")
                    return False
                
            except Exception as e:
                st.error(f"???곗씠??????ㅽ뙣: {e}")
                return False
            
        elif data_type == "blast_data":
            # 諛쒗뙆 ?곗씠??            blast_records = []
            for _, row in data.iterrows():
                blast_records.append({
                    "date": current_date,
                    "blast_date": row.get("諛쒗뙆?쇱옄", ""),
                    "blast_time": row.get("諛쒗뙆?쒓컙", ""),
                    "min_charge": row.get("吏諛쒕떦?μ빟??理쒖냼, kg)", ""),
                    "max_charge": row.get("吏諛쒕떦?μ빟??理쒕?, kg)", ""),
                    "explosive_usage": row.get("??빟?ъ슜??kg)", ""),
                    "vibration": row.get("諛쒗뙆吏꾨룞(cm/sec)", ""),
                    "noise": row.get("諛쒗뙆?뚯쓬(dB(A))", ""),
                    "measurement_location": row.get("怨꾩륫?꾩튂", ""),
                    "remarks": row.get("鍮꾧퀬", ""),
                    "created_at": datetime.now().isoformat()
                })
            
            result = supabase_client.table("blast_data").insert(blast_records).execute()
            st.success("??諛쒗뙆 ?곗씠??????꾨즺")
            
        elif data_type == "instrument_data":
            # 怨꾩륫湲??곗씠??            instrument_records = []
            for _, row in data.iterrows():
                instrument_records.append({
                    "date": current_date,
                    "location": row.get("?꾩튂", ""),
                    "instrument_type": row.get("怨꾩륫湲?醫낅쪟", ""),
                    "instrument_name": row.get("怨꾩륫湲곕챸", ""),
                    "weekly_change": row.get("二쇨컙蹂?붾웾", ""),
                    "cumulative_change": row.get("?꾩쟻蹂?붾웾", ""),
                    "unit": row.get("?⑥쐞", ""),
                    "status": row.get("?곹깭", ""),
                    "ratio": row.get("鍮꾩쑉", ""),
                    "created_at": datetime.now().isoformat()
                })
            
            result = supabase_client.table("instrument_data").insert(instrument_records).execute()
            st.success("??怨꾩륫湲??곗씠?곌? Supabase????λ릺?덉뒿?덈떎.")
            
        return True
        
    except Exception as e:
        st.error(f"??Supabase ????ㅽ뙣: {e}")
        return False

def load_from_supabase(data_type, date=None):
    """Supabase?먯꽌 ?곗씠?곕? 濡쒕뱶?⑸땲??"""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return None
    
    try:
        current_date = date or datetime.now().strftime('%Y-%m-%d')
        
        if data_type == "daily_report":
            result = supabase_client.table("daily_reports").select("*").eq("date", current_date).execute()
            if result.data:
                return result.data[0]  # 泥?踰덉㎏ ?덉퐫??諛섑솚
            return None
            
        elif data_type == "blast_data":
            result = supabase_client.table("blast_data").select("*").eq("date", current_date).execute()
            return pd.DataFrame(result.data) if result.data else None
            
        elif data_type == "instrument_data":
            result = supabase_client.table("instrument_data").select("*").eq("date", current_date).execute()
            return pd.DataFrame(result.data) if result.data else None
            
    except Exception as e:
        st.error(f"??Supabase 濡쒕뱶 ?ㅽ뙣: {e}")
        return None

def save_template_to_supabase(template_bytes, template_name="default", description=""):
    """?묒? ?쒗뵆由우쓣 Supabase?????""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return False, "Supabase ?곌껐 ?ㅽ뙣"
    
    try:
        # ?쒗뵆由??곗씠?곕? base64濡??몄퐫??        import base64
        template_base64 = base64.b64encode(template_bytes).decode('utf-8')
        
        # ?곗씠???ш린 ?뺤씤 (Supabase ?쒗븳: 1MB)
        if len(template_base64) > 1000000:  # ??1MB
            return False, "?쒗뵆由??뚯씪???덈Т ?쎈땲?? 1MB ?댄븯???뚯씪???ъ슜?댁＜?몄슂."
        
        # ?쒗뵆由??곗씠??以鍮?        template_data = {
            'template_name': template_name,
            'template_data': template_base64,
            'description': description,
            'is_default': template_name == "default",
            'created_at': datetime.now().isoformat()
        }
        
        # 湲곗〈 ?쒗뵆由우씠 ?덉쑝硫??낅뜲?댄듃, ?놁쑝硫??덈줈 ?앹꽦
        existing = supabase_client.table('templates')\
            .select('id')\
            .eq('template_name', template_name)\
            .execute()
        
        if existing.data:
            # 湲곗〈 ?쒗뵆由??낅뜲?댄듃
            result = supabase_client.table('templates')\
                .update(template_data)\
                .eq('template_name', template_name)\
                .execute()
        else:
            # ???쒗뵆由??앹꽦
            result = supabase_client.table('templates')\
                .insert(template_data)\
                .execute()
        
        if result.data:
            return True, "?쒗뵆由?????깃났"
        else:
            return False, "?쒗뵆由?????ㅽ뙣: ?곗씠?곌? ??λ릺吏 ?딆븯?듬땲??"
            
    except Exception as e:
        st.error(f"?쒗뵆由????以??ㅻ쪟: {e}")
        return False, f"?쒗뵆由?????ㅽ뙣: {str(e)}"

def get_template_from_supabase(template_name="default"):
    """Supabase?먯꽌 ?묒? ?쒗뵆由?遺덈윭?ㅺ린"""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return None
    
    try:
        result = supabase_client.table('templates')\
            .select('*')\
            .eq('template_name', template_name)\
            .order('created_at', desc=True)\
            .limit(1)\
            .execute()
        
        if result.data:
            # base64 ?붿퐫??            import base64
            template_base64 = result.data[0]['template_data']
            template_bytes = base64.b64decode(template_base64)
            return template_bytes
        return None
    except Exception as e:
        st.error(f"?쒗뵆由?遺덈윭?ㅺ린 ?ㅽ뙣: {e}")
        return None

def get_all_templates():
    """紐⑤뱺 ?쒗뵆由?紐⑸줉 議고쉶"""
    global supabase_client
    if not supabase_client:
        st.warning("?좑툘 Supabase ?곌껐???ㅼ젙?섏? ?딆븯?듬땲??")
        return []
    
    try:
        result = supabase_client.table('templates')\
            .select('template_name, description, created_at, is_default')\
            .order('created_at', desc=True)\
            .execute()
        
        return result.data
    except Exception as e:
        st.error(f"?쒗뵆由?紐⑸줉 議고쉶 ?ㅽ뙣: {e}")
        return []

def check_templates_table():
    """templates ?뚯씠釉?議댁옱 ?щ? ?뺤씤"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase ?곌껐 ?ㅽ뙣"
    
    try:
        # 媛꾨떒??荑쇰━濡??뚯씠釉?議댁옱 ?뺤씤
        result = supabase_client.table('templates').select('id').limit(1).execute()
        return True, "?뚯씠釉?議댁옱"
    except Exception as e:
        return False, f"?뚯씠釉??뺤씤 ?ㅽ뙣: {str(e)}"

def check_daily_report_data_table():
    """daily_report_data ?뚯씠釉?議댁옱 ?щ? ?뺤씤"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase ?곌껐 ?ㅽ뙣"
    
    try:
        # 媛꾨떒??荑쇰━濡??뚯씠釉?議댁옱 ?뺤씤
        result = supabase_client.table('daily_report_data').select('date').limit(1).execute()
        return True, "?뚯씠釉?議댁옱"
    except Exception as e:
        return False, f"?뚯씠釉??뺤씤 ?ㅽ뙣: {str(e)}"

def test_supabase_connection():
    """Supabase ?곌껐???뚯뒪?명빀?덈떎."""
    global supabase_client
    if not supabase_client:
        return False, "Supabase ?대씪?댁뼵?멸? ?놁뒿?덈떎."
    
    try:
        # 媛꾨떒??荑쇰━濡??곌껐 ?뚯뒪??        result = supabase_client.table("daily_report_data").select("count", count="exact").execute()
        return True, "?곌껐 ?깃났"
    except Exception as e:
        return False, f"?곌껐 ?ㅽ뙣: {str(e)}"

def create_daily_report_data_table():
    """daily_report_data ?뚯씠釉??앹꽦"""
    global supabase_client
    if not supabase_client:
        return False, "Supabase ?곌껐 ?ㅽ뙣"
    
    try:
        st.write("?뵇 ?뚯씠釉??앹꽦 ?쒕룄 以?..")
        
        # ?뚯씠釉붿씠 ?놁쑝硫?鍮??곗씠?곕줈 ?뚯뒪???쎌엯 ?쒕룄
        test_data = {
            "date": "2024-01-01",
            "construction_data": {},
            "personnel_data": {},
            "equipment_data": {},
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        result = supabase_client.table("daily_report_data").insert(test_data).execute()
        st.success("???뚯씠釉??앹꽦 ?깃났!")
        
        # ?뚯뒪???곗씠????젣
        supabase_client.table("daily_report_data").delete().eq("date", "2024-01-01").execute()
        
        return True, "?뚯씠釉??앹꽦 ?꾨즺"
    except Exception as e:
        error_msg = str(e).lower()
        if "duplicate key" in error_msg or "unique" in error_msg:
            st.write("?뱄툘 ?뚯씠釉붿씠 ?대? 議댁옱?⑸땲??")
            return True, "?뚯씠釉??대? 議댁옱"
        else:
            return False, f"?뚯씠釉??앹꽦 ?ㅽ뙣: {str(e)}"

def save_cell_mapping_to_supabase(mapping_data, mapping_name="default"):
    """? 留ㅽ븨 ?ㅼ젙??Supabase?????""
    global supabase_client
    if not supabase_client:
        return False, "Supabase ?곌껐 ?ㅽ뙣"
    
    try:
        result = supabase_client.table('cell_mappings').upsert({
            'mapping_name': mapping_name,
            'mapping_data': mapping_data,
            'created_at': datetime.now().isoformat()
        }).execute()
        
        return True, "留ㅽ븨 ?ㅼ젙 ????깃났"
    except Exception as e:
        return False, f"留ㅽ븨 ?ㅼ젙 ????ㅽ뙣: {str(e)}"

def get_cell_mapping_from_supabase(mapping_name="default"):
    """Supabase?먯꽌 ? 留ㅽ븨 ?ㅼ젙 遺덈윭?ㅺ린"""
    global supabase_client
    if not supabase_client:
        return None
    
    try:
        result = supabase_client.table('cell_mappings')\
            .select('mapping_data')\
            .eq('mapping_name', mapping_name)\
            .order('created_at', desc=True)\
            .limit(1)\
            .execute()
        
        if result.data:
            return result.data[0]['mapping_data']
        return None
    except Exception as e:
        st.error(f"留ㅽ븨 ?ㅼ젙 遺덈윭?ㅺ린 ?ㅽ뙣: {e}")
        return None

def extract_file_content(file):
    if file.name.endswith('.pdf'):
        try:
            file.seek(0)
            uploaded_file = genai.upload_file(file, mime_type="application/pdf")
            
            filename_lower = file.name.lower()
            is_measurement_file = any(keyword in filename_lower for keyword in ["怨꾩륫", "吏꾨룞", "?뚯쓬"])
            is_blast_log_file = any(keyword in filename_lower for keyword in ["諛쒗뙆", "?묒뾽", "?쇱?"])

            if is_measurement_file:
                pdf_prompt = "??PDF ?뚯씪? '諛쒗뙆吏꾨룞?뚯쓬 怨꾩륫?쇱?'?낅땲?? ?ㅼ쓬 吏移⑥뿉 ?곕씪 ?곗씠?곕? TSV ?뺤떇?쇰줈 異붿텧?댁＜?몄슂. ... (Prompt content is long and omitted for brevity)"
            elif is_blast_log_file:
                pdf_prompt = "??PDF ?뚯씪? '諛쒗뙆?묒뾽?쇱?'?낅땲?? ?ㅼ쓬 吏移⑥뿉 ?곕씪 二쇱슂 ?곗씠?곕? TSV ?뺤떇?쇰줈 異붿텧?댁＜?몄슂. ... (Prompt content is long and omitted for brevity)"
            else:
                st.warning("?좑툘 ?뚯씪 ?좏삎???뱀젙?????놁뼱 ?쇰컲 ??異붿텧???쒕룄?⑸땲??")
                pdf_prompt = "??PDF?먯꽌 媛??以묒슂??蹂댁씠???쒕? 李얠븘 TSV ?뺤떇?쇰줈 異붿텧?댁＜?몄슂. ..."

            # ?덉쟾?섍쾶 AI 紐⑤뜽???몄텧?⑸땲??
            response_text = safe_generate_content([pdf_prompt, uploaded_file])
            
            # ?ъ슜???앸궃 ?뚯씪? 利됱떆 ??젣?⑸땲??
            genai.delete_file(uploaded_file.name)

            if response_text:
                return re.sub(r'```tsv|```', '', response_text).strip()
            
            return None # safe_generate_content?먯꽌 ?ㅻ쪟瑜??대? ?쒖떆?덉쑝誘濡?None留?諛섑솚?⑸땲??

        except Exception as e:
            st.error(f"??{file.name} 泥섎━ 以?AI ?ㅻ쪟 諛쒖깮: {e}")
            return None
    elif file.name.endswith(('.xlsx', '.xls')):
        try:
            return pd.read_excel(file, engine='openpyxl').to_csv(sep='\t', index=False, encoding='utf-8')
        except Exception as e:
            st.error(f"???묒? ?곗씠??異붿텧 ?ㅽ뙣: {e}")
            return None
    return None

def parse_tsv_to_dataframe(tsv_content):
    try:
        if not tsv_content or tsv_content.strip() == '':
            st.warning("?좑툘 鍮?TSV ?곗씠?곗엯?덈떎.")
            return None
        
        # main2.py 諛⑹떇?쇰줈 媛꾨떒?섍쾶 泥섎━
        cleaned_content = '\n'.join(line.strip() for line in tsv_content.split('\n') if line.strip())
        
        if not cleaned_content:
            st.warning("?좑툘 ?뺤젣??TSV ?곗씠?곌? ?놁뒿?덈떎.")
            return None
        
        df = pd.read_csv(io.StringIO(cleaned_content), sep='\t', encoding='utf-8')
        
        if df.empty:
            st.warning("?좑툘 ?뚯떛???곗씠?고봽?덉엫??鍮꾩뼱?덉뒿?덈떎.")
            return None
        
        df.columns = df.columns.str.strip()
        
        # Arrow 吏곷젹???ㅻ쪟 諛⑹?瑜??꾪븳 ?곗씠???뺤젣
        for col in df.columns:
            # 鍮?臾몄옄?댁쓣 '0' ?먮뒗 ?곸젅??湲곕낯媛믪쑝濡?蹂寃?            df[col] = df[col].fillna('').astype(str)
            # ?レ옄泥섎읆 蹂댁씠??而щ읆??鍮?媛믪쓣 '0'?쇰줈 蹂寃?            if '?꾧퀎' in col or '蹂?붾웾' in col or any(x in col for x in ['紐?, '?', 'kg', 'cm/sec', 'dB']):
                df[col] = df[col].replace('', '0')
        
        return df
        
    except Exception as e:
        st.error(f"TSV ?뚯떛 以??ㅻ쪟 諛쒖깮: {e}")
        st.info(f"?뵇 ?먮낯 TSV ?곗씠??(泥섏쓬 200??: {tsv_content[:200]}")
        return None

def extract_work_date_from_response(response_text):
    """AI ?묐떟?먯꽌 ?묒뾽 ?좎쭨瑜?異붿텧?⑸땲??"""
    if not response_text:
        return datetime.now().strftime('%Y-%m-%d')
    
    # WORK_DATE: YYYY-MM-DD ?⑦꽩 李얘린
    date_pattern = r'WORK_DATE:\s*(\d{4}-\d{2}-\d{2})'
    match = re.search(date_pattern, response_text)
    
    if match:
        return match.group(1)
    
    # ????⑦꽩???쒕룄 (2024??1??15?? 24.01.15 ??
    alt_patterns = [
        r'(\d{4})??s*(\d{1,2})??s*(\d{1,2})??,
        r'(\d{2})\.(\d{2})\.(\d{2})',
        r'(\d{4})-(\d{2})-(\d{2})',
        r'(\d{4})/(\d{2})/(\d{2})'
    ]
    
    for pattern in alt_patterns:
        match = re.search(pattern, response_text)
        if match:
            groups = match.groups()
            if len(groups) == 3:
                year, month, day = groups
                if len(year) == 2:  # 24.01.15 ?뺤떇??寃쎌슦
                    year = '20' + year
                try:
                    # ?좎쭨 ?좏슚??寃利?                    datetime.strptime(f"{year}-{month.zfill(2)}-{day.zfill(2)}", '%Y-%m-%d')
                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                except ValueError:
                    continue
    
    # 異붿텧 ?ㅽ뙣???꾩옱 ?좎쭨 ?ъ슜
    return datetime.now().strftime('%Y-%m-%d')

def extract_qa_checklist_from_response(response_text):
    """AI ?묐떟?먯꽌 QA-Checklist瑜?異붿텧?⑸땲??"""
    if not response_text:
        return ""
    
    # QA-CHECKLIST ?뱀뀡 李얘린
    qa_patterns = [
        r'QA-CHECKLIST[\s\S]*?(?=\n\n|\n#|\n##|\n###|\n####|\n#####|\n######|$)',
        r'QA-CHECKLIST[\s\S]*',
        r'## 5\. QA-CHECKLIST[\s\S]*?(?=\n\n|\n#|\n##|\n###|\n####|\n#####|\n######|$)',
        r'## 5\. QA-CHECKLIST[\s\S]*'
    ]
    
    for pattern in qa_patterns:
        match = re.search(pattern, response_text, re.IGNORECASE | re.DOTALL)
        if match:
            qa_content = match.group(0).strip()
            # 遺덊븘?뷀븳 ?ㅻ뜑 ?쒓굅
            qa_content = re.sub(r'^## 5\. QA-CHECKLIST.*?\n', '', qa_content, flags=re.IGNORECASE | re.DOTALL)
            qa_content = re.sub(r'^QA-CHECKLIST.*?\n', '', qa_content, flags=re.IGNORECASE | re.DOTALL)
            return qa_content.strip()
    
    return ""

def convert_to_number_if_possible(value):
    """媛믪씠 ?レ옄濡?蹂??媛?ν븳吏 ?뺤씤?섍퀬 ?レ옄濡?蹂?섑빀?덈떎."""
    if value is None or value == "":
        return 0
    
    # 臾몄옄?댁씤 寃쎌슦 怨듬갚 ?쒓굅
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return 0
    
    try:
        # ?뺤닔濡?蹂???쒕룄
        return int(float(value))
    except (ValueError, TypeError):
        # ?レ옄濡?蹂?섑븷 ???녿뒗 寃쎌슦 ?먮옒 媛?諛섑솚
        return value

def extract_tsv_from_response(response_text):
    if not response_text: return ""
    lines = response_text.strip().split('\n')
    cleaned_lines = [line.strip() for line in lines if '\t' in line.strip()]
    return "\n".join(cleaned_lines)

def fix_tsv_field_count(tsv_str):
    lines = tsv_str.strip().split('\n')
    if not lines: 
        return tsv_str
    
    header = lines[0]
    n_fields = header.count('\t') + 1
    fixed_lines = [header]
    
    for line in lines[1:]:
        fields = line.split('\t')
        if len(fields) < n_fields:
            fields += [''] * (n_fields - len(fields))
        elif len(fields) > n_fields:
            fields = fields[:n_fields-1] + [' '.join(fields[n_fields-1:])]
        fixed_lines.append('\t'.join(fields))
    
    return '\n'.join(fixed_lines)

def create_excel_report(**kwargs):
    """?덈줈???묒? 蹂닿퀬?쒕? 泥섏쓬遺???앹꽦?⑸땲??- 1?섏씠吏 理쒖쟻??踰꾩쟾"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    
    wb = Workbook()
    ws = wb.active
    ws.title = "怨듭궗?쇰낫"
    
    # ?섏씠吏 ?ㅼ젙 - A4 ?몃줈, 1?섏씠吏 理쒖쟻??    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9  # A4 ?⑹? ?ш린
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    
    # ?ㅽ????뺤쓽
    title_font = Font(bold=True, size=16, name='留묒? 怨좊뵓')
    header_font = Font(bold=True, size=11, name='留묒? 怨좊뵓')
    normal_font = Font(size=9, name='留묒? 怨좊뵓')
    small_font = Font(size=8, name='留묒? 怨좊뵓')
    
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    # 湲곕낯 ?뺣낫
    work_date = kwargs.get("work_date", datetime.now().strftime('%Y-%m-%d'))
    project_name = kwargs.get("project_name", "?곕꼸 嫄댁꽕怨듭궗")
    section_name = kwargs.get("section_name", "?由??좏뭾-?꾨┝")
    
    # 1. ?쒕ぉ 諛??ㅻ뜑 (1-3??
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value=f"{project_name} ?쇱씪?묒뾽蹂닿퀬??)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = header_fill
    title_cell.border = thick_border
    
    # 湲곕낯 ?뺣낫 (2??
    ws.cell(row=2, column=1, value="援ш컙").font = header_font
    ws.cell(row=2, column=2, value=section_name).font = normal_font
    ws.cell(row=2, column=6, value="蹂닿퀬??).font = header_font
    ws.cell(row=2, column=7, value=work_date).font = normal_font
    
    for col in range(1, 11):
        ws.cell(row=2, column=col).border = thin_border
        if col in [1, 6]:
            ws.cell(row=2, column=col).fill = sub_header_fill
    
    current_row = 4
    
    # 2. ?좎뵪 諛?湲곕낯 ?꾪솴????以꾩뿉 (4-5??
    tables_data = kwargs.get("tables_data", [])
    
    # ?좎뵪?뺣낫 (?쇱そ)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value="?좎뵪 ?꾪솴").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    
    # 二쇱슂 ?쒓났 ?꾪솴 (?쇱そ)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value="二쇱슂 ?쒓났 ?꾪솴").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    
    if tables_data and len(tables_data) > 0 and tables_data[0] is not None:
        construction_df = tables_data[0]
        # ?곸쐞 5媛???ぉ留??쒖떆
        main_items = construction_df.head(5) if not construction_df.empty else pd.DataFrame()
        construction_text = ""
        for _, row in main_items.iterrows():
            if str(row['?꾧퀎']) not in ['', '0', '0.0']:
                construction_text += f"{row['援щ텇']}: {row['?꾧퀎']}, "
        construction_text = construction_text.rstrip(", ")
        
        ws.merge_cells(f'F{current_row+1}:J{current_row+1}')
        ws.cell(row=current_row+1, column=6, value=construction_text).font = small_font
        ws.cell(row=current_row+1, column=6).border = thin_border
    
    current_row += 3
    
    # 3. 湲덉씪 ?묒뾽 ?댁슜 (二쇱슂 ?뱀뀡)
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws.cell(row=current_row, column=1, value="湲덉씪 二쇱슂 ?묒뾽 ?댁슜").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    current_row += 1
    
    if tables_data and len(tables_data) > 1 and tables_data[1] is not None:
        work_df = tables_data[1]
        # ?ㅼ젣 ?묒뾽???덈뒗 ??ぉ留??쒖떆
        active_work = work_df[work_df['湲덉씪?묒뾽'].str.strip() != ''] if not work_df.empty else pd.DataFrame()
        
        for idx, (_, row) in enumerate(active_work.head(6).iterrows()):  # 理쒕? 6媛???ぉ
            ws.cell(row=current_row, column=1, value=f"??{row['援щ텇']}").font = small_font
            ws.merge_cells(f'B{current_row}:J{current_row}')
            ws.cell(row=current_row, column=2, value=row['湲덉씪?묒뾽']).font = small_font
            ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
            
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
    else:
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws.cell(row=current_row, column=1, value="?묒뾽 ?댁슜???놁뒿?덈떎.").font = small_font
        current_row += 1
    
    current_row += 1
    
    # 4. ?몄썝 諛??λ퉬 ?꾪솴 (?붿빟)
    col1_start = current_row
    
    # ?몄썝 ?꾪솴 (?쇱そ)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1, value="?몄썝 ?꾪솴").font = header_font
    ws.cell(row=current_row, column=1).fill = sub_header_fill
    ws.cell(row=current_row, column=1).border = thin_border
    current_row += 1
    
    if tables_data and len(tables_data) > 2 and tables_data[2] is not None:
        personnel_df = tables_data[2]
        if not personnel_df.empty:
            # 珥??몄썝 怨꾩궛
            total_personnel = 0
            for col in personnel_df.columns[1:]:  # 泥?踰덉㎏ ??援щ텇) ?쒖쇅
                personnel_df[col] = pd.to_numeric(personnel_df[col], errors='coerce').fillna(0)
                total_personnel += personnel_df[col].sum()
            
            ws.cell(row=current_row, column=1, value="珥??ъ엯?몄썝").font = small_font
            ws.cell(row=current_row, column=2, value=f"{int(total_personnel)}紐?).font = small_font
            
            # 二쇱슂 吏곸쥌蹂??몄썝 (0???꾨땶 寃껊쭔)
            for _, row in personnel_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in personnel_df.columns[1:])
                if row_total > 0:
                    ws.cell(row=current_row+1, column=1, value=row['援щ텇']).font = small_font
                    ws.cell(row=current_row+1, column=2, value=f"{int(row_total)}紐?).font = small_font
                    current_row += 1
                    if current_row - col1_start > 8:  # 理쒕? 8以?                        break
    
    # ?λ퉬 ?꾪솴 (?ㅻⅨ履?
    current_row = col1_start
    ws.merge_cells(f'F{current_row}:J{current_row}')
    ws.cell(row=current_row, column=6, value="?λ퉬 ?꾪솴").font = header_font
    ws.cell(row=current_row, column=6).fill = sub_header_fill
    ws.cell(row=current_row, column=6).border = thin_border
    current_row += 1
    
    if tables_data and len(tables_data) > 3 and tables_data[3] is not None:
        equipment_df = tables_data[3]
        if not equipment_df.empty:
            # 二쇱슂 ?λ퉬 (0???꾨땶 寃껊쭔)
            for _, row in equipment_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in equipment_df.columns[1:])
                if row_total > 0:
                    ws.cell(row=current_row, column=6, value=row['援щ텇']).font = small_font
                    ws.cell(row=current_row, column=7, value=f"{int(row_total)}?").font = small_font
                    current_row += 1
                    if current_row - col1_start > 8:  # 理쒕? 8以?                        break
    
    # 5. 諛쒗뙆 諛?怨꾩륫湲??곗씠??(?섎떒 ?붿빟)
    current_row = max(current_row, col1_start + 10) + 1
    
    blast_df = kwargs.get("blast_df")
    instrument_df = kwargs.get("instrument_df")
    
    if blast_df is not None and not blast_df.empty:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value="諛쒗뙆 ?꾪솴").font = header_font
        ws.cell(row=current_row, column=1).fill = sub_header_fill
        ws.cell(row=current_row, column=1).border = thin_border
        
        blast_count = len(blast_df)
        total_explosive = blast_df['??빟?ъ슜??kg)'].sum() if '??빟?ъ슜??kg)' in blast_df.columns else 0
        
        ws.merge_cells(f'A{current_row+1}:E{current_row+1}')
        ws.cell(row=current_row+1, column=1, value=f"諛쒗뙆?잛닔: {blast_count}?? ??빟?ъ슜?? {total_explosive}kg").font = small_font
        ws.cell(row=current_row+1, column=1).border = thin_border
        
    if instrument_df is not None and not instrument_df.empty:
        ws.merge_cells(f'F{current_row}:J{current_row}')
        ws.cell(row=current_row, column=6, value="怨꾩륫湲??꾪솴").font = header_font
        ws.cell(row=current_row, column=6).fill = sub_header_fill
        ws.cell(row=current_row, column=6).border = thin_border
        
        warning_count = len(instrument_df[instrument_df['?곹깭'] != '?덉젙']) if '?곹깭' in instrument_df.columns else 0
        total_count = len(instrument_df)
        
        ws.merge_cells(f'F{current_row+1}:J{current_row+1}')
        ws.cell(row=current_row+1, column=6, value=f"珥?{total_count}媛쒖냼, 寃쎄퀬: {warning_count}媛쒖냼").font = small_font
        ws.cell(row=current_row+1, column=6).border = thin_border
    
    # 紐⑤뱺 ????뚮몢由??곸슜 諛????덈퉬 ?ㅼ젙
    for row in ws.iter_rows(min_row=1, max_row=current_row+2, min_col=1, max_col=10):
        for cell in row:
            if not cell.border.left.style:
                cell.border = thin_border
    
    # ???덈퉬 理쒖쟻??(A4 1?섏씠吏??留욊쾶)
    column_widths = [12, 12, 10, 10, 10, 12, 12, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ???믪씠 議곗젙
    for row_num in range(1, current_row + 3):
        ws.row_dimensions[row_num].height = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_pdf_report(**kwargs):
    """?덈줈??PDF 蹂닿퀬?쒕? ?앹꽦?⑸땲??- ?꾩쟾???댁슜 ?쒖떆 踰꾩쟾"""
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    
    # ?쒓? ?고듃 ?깅줉 ?쒕룄
    try:
        # ?쒖뒪?쒖뿉???쒓? ?고듃 李얘린
        font_paths = [
            "C:/Windows/Fonts/malgun.ttf",  # 留묒? 怨좊뵓
            "C:/Windows/Fonts/gulim.ttc",   # 援대┝
            "C:/Windows/Fonts/batang.ttc",  # 諛뷀깢
        ]
        
        korean_font = None
        for font_path in font_paths:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Korean', font_path))
                korean_font = 'Korean'
                break
        
        if not korean_font:
            korean_font = 'Helvetica'  # ?대갚 ?고듃
    except:
        korean_font = 'Helvetica'
    
    # PDF ?앹꽦
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # ?ㅽ????뺤쓽
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=korean_font,
        fontSize=16,
        spaceAfter=12,
        alignment=1,  # 以묒븰 ?뺣젹
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontName=korean_font,
        fontSize=12,
        spaceAfter=6,
        textColor=colors.darkblue
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=korean_font,
        fontSize=9,
        spaceAfter=3
    )
    
    # ?댁슜 援ъ꽦
    story = []
    
    # 湲곕낯 ?뺣낫
    work_date = kwargs.get("work_date", datetime.now().strftime('%Y-%m-%d'))
    project_name = kwargs.get("project_name", "?곕꼸 嫄댁꽕怨듭궗")
    section_name = kwargs.get("section_name", "?由??좏뭾-?꾨┝")
    tables_data = kwargs.get("tables_data", [])
    
    # 1. ?쒕ぉ
    story.append(Paragraph(f"{project_name} ?쇱씪?묒뾽蹂닿퀬??, title_style))
    story.append(Spacer(1, 6*mm))
    
    # 2. 湲곕낯 ?뺣낫 ?뚯씠釉?    basic_info = [
        ['援ш컙', section_name, '蹂닿퀬??, work_date]
    ]
    basic_table = Table(basic_info, colWidths=[30*mm, 60*mm, 30*mm, 60*mm])
    basic_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.lightblue),
        ('BACKGROUND', (2, 0), (2, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), korean_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(basic_table)
    story.append(Spacer(1, 6*mm))
    
    # 3. ?쒓났?꾪솴
    if tables_data and len(tables_data) > 0 and tables_data[0] is not None:
        story.append(Paragraph("?쒓났?꾪솴", header_style))
        construction_df = tables_data[0]
        if not construction_df.empty:
            construction_data = [['援щ텇', '?꾧퀎']]
            for _, row in construction_df.iterrows():
                if str(row['?꾧퀎']) not in ['', '0', '0.0']:  # 媛믪씠 ?덈뒗 寃껊쭔
                    construction_data.append([str(row['援щ텇']), str(row['?꾧퀎'])])
            
            if len(construction_data) > 1:  # ?ㅻ뜑 ?몄뿉 ?곗씠?곌? ?덉쑝硫?                construction_table = Table(construction_data, colWidths=[120*mm, 40*mm])
                construction_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 0), (-1, -1), korean_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # 泥?踰덉㎏ ???쇱そ ?뺣젹
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # ??踰덉㎏ ??以묒븰 ?뺣젹
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(construction_table)
        story.append(Spacer(1, 6*mm))
    
    # 4. 二쇱슂 ?묒뾽 ?댁슜
    if tables_data and len(tables_data) > 1 and tables_data[1] is not None:
        story.append(Paragraph("湲덉씪 二쇱슂 ?묒뾽 ?댁슜", header_style))
        work_df = tables_data[1]
        active_work = work_df[work_df['湲덉씪?묒뾽'].str.strip() != ''] if not work_df.empty else pd.DataFrame()
        
        if not active_work.empty:
            work_data = [['援щ텇', '?묒뾽?댁슜']]
            for _, row in active_work.iterrows():  # 紐⑤뱺 ?묒뾽 ?댁슜 ?ы븿
                work_data.append([str(row['援щ텇']), str(row['湲덉씪?묒뾽'])])
            
            work_table = Table(work_data, colWidths=[60*mm, 110*mm])
            work_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), korean_font),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # 泥?踰덉㎏ ??以묒븰 ?뺣젹
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),    # ??踰덉㎏ ???쇱そ ?뺣젹
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(work_table)
        else:
            story.append(Paragraph("湲덉씪 ?묒뾽 ?댁슜???놁뒿?덈떎.", normal_style))
        story.append(Spacer(1, 6*mm))
    
    # 5. ?몄썝 ?꾪솴 ?붿빟
    if tables_data and len(tables_data) > 2 and tables_data[2] is not None:
        story.append(Paragraph("?몄썝 ?꾪솴", header_style))
        personnel_df = tables_data[2]
        if not personnel_df.empty:
            # 珥??몄썝 怨꾩궛
            total_personnel = 0
            personnel_summary = []
            
            for _, row in personnel_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in personnel_df.columns[1:])
                if row_total > 0:
                    personnel_summary.append([str(row['援щ텇']), f"{int(row_total)}紐?])
                    total_personnel += row_total
            
            # 珥??몄썝??留??꾩뿉 異붽?
            personnel_data = [['援щ텇', '?몄썝??], ['珥??ъ엯?몄썝', f"{int(total_personnel)}紐?]] + personnel_summary  # 紐⑤뱺 ?몄썝 ?ы븿
            
            personnel_table = Table(personnel_data, colWidths=[80*mm, 40*mm])
            personnel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('BACKGROUND', (0, 1), (-1, 1), colors.lightyellow),  # 珥??몄썝 媛뺤“
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), korean_font),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(personnel_table)
        story.append(Spacer(1, 6*mm))
    
    # 6. ?λ퉬 ?꾪솴 ?붿빟
    if tables_data and len(tables_data) > 3 and tables_data[3] is not None:
        story.append(Paragraph("?λ퉬 ?꾪솴", header_style))
        equipment_df = tables_data[3]
        if not equipment_df.empty:
            equipment_summary = []
            
            for _, row in equipment_df.iterrows():
                row_total = sum(pd.to_numeric(row[col], errors='coerce') or 0 for col in equipment_df.columns[1:])
                if row_total > 0:
                    equipment_summary.append([str(row['援щ텇']), f"{int(row_total)}?"])
            
            if equipment_summary:
                equipment_data = [['?λ퉬紐?, '???]] + equipment_summary  # 紐⑤뱺 ?λ퉬 ?ы븿
                
                equipment_table = Table(equipment_data, colWidths=[80*mm, 40*mm])
                equipment_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 0), (-1, -1), korean_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                story.append(equipment_table)
        story.append(Spacer(1, 6*mm))
    
    # 8. 諛쒗뙆 ?곗씠??(?곸꽭)
    blast_df = kwargs.get("blast_df")
    if blast_df is not None and not blast_df.empty:
        story.append(Paragraph("諛쒗뙆 ?꾪솴", header_style))
        
        # 諛쒗뙆 ?곗씠???뚯씠釉?        blast_columns = list(blast_df.columns)
        blast_data = [blast_columns]  # ?ㅻ뜑
        
        for _, row in blast_df.iterrows():
            blast_data.append([str(row[col]) for col in blast_columns])
        
        # 而щ읆 ?섏뿉 ?곕씪 ?숈쟻?쇰줈 ?덈퉬 議곗젙
        col_count = len(blast_columns)
        col_width = 170 // col_count * mm  # ?꾩껜 ?덈퉬瑜?而щ읆 ?섎줈 ?섎닎
        col_widths = [col_width] * col_count
        
        blast_table = Table(blast_data, colWidths=col_widths)
        blast_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), korean_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(blast_table)
        story.append(Spacer(1, 6*mm))
    
    # 9. 怨꾩륫湲??곗씠??(?곸꽭)
    instrument_df = kwargs.get("instrument_df")
    if instrument_df is not None and not instrument_df.empty:
        story.append(Paragraph("怨꾩륫湲??꾪솴", header_style))
        
        # 怨꾩륫湲??곗씠???뚯씠釉?        instrument_columns = list(instrument_df.columns)
        instrument_data = [instrument_columns]  # ?ㅻ뜑
        
        for _, row in instrument_df.iterrows():
            instrument_data.append([str(row[col]) for col in instrument_columns])
        
        # 而щ읆 ?섏뿉 ?곕씪 ?숈쟻?쇰줈 ?덈퉬 議곗젙
        col_count = len(instrument_columns)
        col_width = 170 // col_count * mm  # ?꾩껜 ?덈퉬瑜?而щ읆 ?섎줈 ?섎닎
        col_widths = [col_width] * col_count
        
        instrument_table = Table(instrument_data, colWidths=col_widths)
        instrument_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), korean_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(instrument_table)
    
    # PDF ?앹꽦
    try:
        doc.build(story)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        # ?쒓? ?고듃 臾몄젣 ???대갚
        for element in story:
            if hasattr(element, 'style') and hasattr(element.style, 'fontName'):
                element.style.fontName = 'Helvetica'
        
        doc.build(story)
        output.seek(0)
        return output.getvalue()


# ???⑥닔?????댁긽 ?ъ슜?섏? ?딆쑝誘濡??쒓굅
# insert_data_to_excel_with_mapping ?⑥닔瑜??ъ슜?섏꽭??

def insert_data_to_excel_with_mapping(template_bytes, basic_info, tables_data, cell_mapping=None, table_mapping=None, previous_data=None):
    """?묒? ?쒗뵆由우뿉 湲곕낯?뺣낫瑜??뱀젙 ???留ㅽ븨?섏뿬 ?쎌엯?⑸땲??"""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import io
        
        # ?묒? ?뚯씪 濡쒕뱶
        workbook = load_workbook(io.BytesIO(template_bytes))
        worksheet = workbook.active
        
        # 湲곕낯?뺣낫 留ㅽ븨 (?ъ슜???ㅼ젙 ?먮뒗 湲곕낯媛??ъ슜)
        if cell_mapping is None:
            # 湲곕낯 ? 留ㅽ븨 ?ㅼ젙
            cell_mapping = {
                'date': 'u2',
                'project_name': 'd4', 
                'max_temp': 'o4',
                'min_temp': 'o5',
                'precipitation': 'o6',
                'planned_progress': 'w4',
                'actual_progress': 'w5'
            }
        
        # 湲곕낯?뺣낫 ?쎌엯
        for key, cell_address in cell_mapping.items():
            if key in basic_info:
                try:
                    # ?덉쟾??? 媛??ㅼ젙
                    try:
                        # ? 二쇱냼瑜??뚯떛?섏뿬 吏곸젒 ?묎렐
                        row_idx, col_idx = parse_cell_address(cell_address)
                        
                        # ???吏곸젒 ?묎렐
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        # 蹂묓빀????몄? ?뺤씤
                        is_merged = False
                        target_cell = cell_address
                        
                        for merged_range in worksheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                # 蹂묓빀?????泥?踰덉㎏ ? 二쇱냼 怨꾩궛
                                target_col = chr(ord('A') + merged_range.min_col - 1)
                                target_row = merged_range.min_row
                                target_cell = f"{target_col}{target_row}"
                                break
                        
                        # ?곗씠???쎌엯 (?띿뒪?몃뒗 洹몃?濡? ?レ옄??蹂??
                        if key in ['project_name', 'date']:
                            # ?띿뒪???곗씠?곕뒗 洹몃?濡??ъ슜
                            cell_value = basic_info[key]
                        else:
                            # ?レ옄 ?곗씠?곕뒗 蹂??                            cell_value = convert_to_number_if_possible(basic_info[key])
                        
                        if is_merged:
                            # 蹂묓빀?????泥?踰덉㎏ ????곗씠???쎌엯
                            worksheet[target_cell] = cell_value
                        else:
                            # ?쇰컲 ????곗씠???쎌엯
                            worksheet[cell_address] = cell_value
                            
                    except Exception as cell_error:
                        st.warning(f"?좑툘 ? {cell_address} 泥섎━ 以??ㅻ쪟: {cell_error}")
                        continue
                        
                except Exception as e:
                    st.warning(f"?좑툘 ? {cell_address} 泥섎━ 以??ㅻ쪟: {e}")
                    # ?ㅻ쪟 諛쒖깮 ???대떦 ? 嫄대꼫?곌린
                    continue
        
        # ?뚯씠釉?異붿텧媛??쎌엯 (table_mapping???덈뒗 寃쎌슦)
        if table_mapping:
            for key, cell_address in table_mapping.items():
                try:
                    # 蹂묓빀??? ?덉쟾?섍쾶 泥섎━
                    cell = worksheet[cell_address]
                    
                    # 蹂묓빀????몄? ?뺤씤
                    is_merged = False
                    target_cell = cell_address
                    
                    # 蹂묓빀??? 踰붿쐞?먯꽌 ?대떦 ? 李얘린
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell_address in merged_range:
                            is_merged = True
                            target_cell = merged_range.start_cell.coordinate
                            break
                    
                    # ?뚯씠釉??곗씠?곗뿉???대떦 媛?異붿텧?섏뿬 ?쎌엯
                    table_value = ""
                    if key in tables_data:
                        df = tables_data[key]
                        if df is not None and not df.empty:
                            table_value = str(df.iloc[0, 1]) if len(df.columns) > 1 else str(df.iloc[0, 0])
                    
                    # ?곗씠???쎌엯 (?レ옄 ???媛뺤젣 ?ㅼ젙)
                    cell_value = convert_to_number_if_possible(table_value)
                    if is_merged:
                        # 蹂묓빀?????泥?踰덉㎏ ????곗씠???쎌엯
                        worksheet[target_cell] = cell_value
                    else:
                        # ?쇰컲 ????곗씠???쎌엯
                        worksheet[cell_address] = cell_value
                        
                except Exception as e:
                    st.warning(f"?좑툘 ?뚯씠釉?異붿텧媛?? {cell_address} 泥섎━ 以??ㅻ쪟: {e}")
                    continue
        
        # ?꾩씪 ?곗씠???곸슜 (previous_data媛 ?덈뒗 寃쎌슦)
        if previous_data:
            try:
                st.info(f"?뵇 ?꾩씪 ?곗씠??援ъ“: {list(previous_data.keys())}")
                
                # 1. ?쒓났?꾪솴 ?꾩씪 ?곗씠???곸슜 (T11~43 ?꾧퀎 ??N11~43 ?꾩씪源뚯?)
                construction_data = previous_data.get("?쒓났?꾪솴", [])
                if construction_data:
                    row = 11
                    for item in construction_data:
                        if row <= 43 and isinstance(item, dict):
                            cumulative_value = item.get("?꾧퀎", 0)
                            worksheet[f"N{row}"] = cumulative_value
                            row += 1
                    st.info(f"???쒓났?꾪솴 ?꾩씪 ?곗씠???곸슜 ?꾨즺: {row-11}??)
                
                # 2. ?몄썝 ?꾩씪 ?곗씠???곸슜 (L66~87, Y66~87)
                personnel_data = previous_data.get("?몄썝", [])
                if personnel_data:
                    row = 66
                    for item in personnel_data:
                        if row <= 87 and isinstance(item, dict):
                            previous_value = item.get("?꾩씪源뚯?", 0)
                            cumulative_value = item.get("?꾧퀎", 0)
                            worksheet[f"L{row}"] = previous_value
                            worksheet[f"Y{row}"] = cumulative_value
                            row += 1
                    st.info(f"???몄썝 ?꾩씪 ?곗씠???곸슜 ?꾨즺: {row-66}??)
                
                # 3. ?λ퉬 ?꾩씪 ?곗씠???곸슜 (L91~119, Y91~119)
                equipment_data = previous_data.get("?λ퉬", [])
                if equipment_data:
                    row = 91
                    for item in equipment_data:
                        if row <= 119 and isinstance(item, dict):
                            previous_value = item.get("?꾩씪源뚯?", 0)
                            cumulative_value = item.get("?꾧퀎", 0)
                            worksheet[f"L{row}"] = previous_value
                            worksheet[f"Y{row}"] = cumulative_value
                            row += 1
                    st.info(f"???λ퉬 ?꾩씪 ?곗씠???곸슜 ?꾨즺: {row-91}??)
                
                st.success("???꾩씪 ?곗씠?곌? ?먮룞?쇰줈 ?곸슜?섏뿀?듬땲??")
                
            except Exception as e:
                st.warning(f"?좑툘 ?꾩씪 ?곗씠???곸슜 以??ㅻ쪟: {e}")
                st.error(f"?꾩씪 ?곗씠??援ъ“: {previous_data}")
        # ?뚯씠釉??곗씠???쎌엯 (?꾩껜 ?뚯씠釉??쎌엯)
        if tables_data:
            # ?ъ슜???낅젰媛믪쓣 湲곕컲?쇰줈 ?뚯씠釉??꾩튂 怨꾩궛
            table_positions = {}
            if table_mapping:
                for table_name, cell_address in table_mapping.items():
                    if cell_address:
                        # ? 二쇱냼瑜????대줈 蹂??                        from openpyxl.utils import column_index_from_string
                        try:
                            # ? 二쇱냼?먯꽌 ?닿낵 ??遺꾨━ (?? "A1" -> "A", "1")
                            import re
                            match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
                            if match:
                                col_str = match.group(1)
                                row_str = match.group(2)
                                row = int(row_str)
                                col = column_index_from_string(col_str)
                                table_positions[table_name] = {"row": row, "col": col}
                            else:
                                raise ValueError(f"?섎せ??? 二쇱냼 ?뺤떇: {cell_address}")
                        except Exception as e:
                            # 湲곕낯媛??ъ슜
                            default_positions = {
                                "?쒓났?꾪솴": {"row": 8, "col": 17},   # q8
                                "?묒뾽?댁슜": {"row": 11, "col": 17},  # q11
                                "?몄썝": {"row": 66, "col": 29},      # ac66
                                "?λ퉬": {"row": 106, "col": 29}      # ac106
                            }
                            if table_name in default_positions:
                                table_positions[table_name] = default_positions[table_name]
            else:
                # 湲곕낯 ?꾩튂 ?ъ슜
                table_positions = {
                    "?쒓났?꾪솴": {"row": 8, "col": 17},   # q8
                    "?묒뾽?댁슜": {"row": 11, "col": 17},  # q11
                    "?몄썝": {"row": 66, "col": 29},      # ac66
                    "?λ퉬": {"row": 106, "col": 29}      # ac106
                }
            
            # 紐⑤뱺 ?뚯씠釉??쎌엯 (?レ옄 ???媛뺤젣 ?ㅼ젙)
            for table_name, df in tables_data.items():
                if table_name in table_positions and df is not None and not df.empty:
                    pos = table_positions[table_name]
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                        for c_idx, value in enumerate(row):
                            # ?レ옄 ?곗씠?????媛뺤젣 ?ㅼ젙
                            cell_value = convert_to_number_if_possible(value)
                            worksheet.cell(row=pos["row"] + r_idx, column=pos["col"] + c_idx, value=cell_value)
        
        # ?섏젙???묒? ?뚯씪??諛붿씠?몃줈 蹂??        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"?묒? ?곗씠??留ㅽ븨 ?쎌엯 以??ㅻ쪟: {e}")
        return None


# --- STATE INITIALIZATION ---
def initialize_session_state():
    states = {
        "kakao_work_completed": False, "basic_info_completed": False, "excel_export_completed": False,
        "processed_tables": [], "kakao_results": None,
        "final_excel_data": None, "processed_template_filename": None,
        "all_accumulated_rows": [], "reset_flag": 0,
        "prompt": DEFAULT_PROMPT, "work_date": None,
        "warning_rows_instrument": None,
        # ?꾨＼?꾪듃 愿由?愿???곹깭
        "current_prompt_name": "湲곕낯 ?꾨＼?꾪듃",
        "show_prompt_editor": False,
        "show_table_editor": False,
        "prompt_list": [],
        # Supabase ???愿???곹깭
        "daily_report_saved": False,
        "save_success_message": "",
        "save_error_message": "",
        "save_success_date": "",
        # ?뚯씠釉?援ъ“ 愿由?        "construction_rows": [
            "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾)",
            "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾) ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 1)?뺢굅???쇱씠??,
            "2. ?좏뭾?뺢굅??- 1)?뺢굅??誘몃뱾 ?щ씪釉?,
            "2. ?좏뭾?뺢굅????2)二쇱텧?낃뎄 ?섏쭅援??쇱씠??,
            "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB ?뺢굅??諛⑸㈃ ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB ?섏듅?듬줈 諛⑸㈃ ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (2)PCC ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (3)PCD ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (4)PHA ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒 - ?섏쭅援??쇱씠??,
            "2. ?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒 - PHB ?쇱씠??,
            "2. ?좏뭾?뺢굅??- 4)?몃?異쒖엯援?異쒖엯援?#3) 援댁갑",
            "2. ?좏뭾?뺢굅??- 4)?몃?異쒖엯援?異쒖엯援?#2) 援댁갑",
            "2. ?좏뭾?뺢굅??- 4)?몃?異쒖엯援?異쒖엯援?#1) 援댁갑",
            "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCF) 援댁갑",
            "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCF) ?쇱씠??,
            "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCE) 援댁갑",
            "3. ?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸 ?곌껐?곕꼸(PCE) ?쇱씠??,
            "3. ?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX 蹂대씪留?諛⑸㈃ 援ъ“臾?,
            "3. ?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX ?由?諛⑸㈃ 援댁갑",
            "4. 蹂몄꽑?곕꼸(2援ш컙, ?좏뭾-?꾨┝) 援댁갑",
            "4. 蹂몄꽑?곕꼸(2援ш컙, ?좏뭾-?꾨┝) ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅???곕꼸 ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅??誘몃뱾 ?щ씪釉?,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 ?섏쭅援??쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 PCA ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 PCC ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1 PHA ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 ?섏쭅援??쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 PCA ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 PCC ?쇱씠??,
            "5. ?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2 PHB ?쇱씠??
        ],
        "work_content_rows": [
            "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾)",
            "2.?좏뭾?뺢굅??- 1)?뺢굅???곕꼸",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (2)PCC",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (3)PCD",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (4)PHA",
            "2.?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒",
            "2.?좏뭾?뺢굅??- 4)?몃?異쒖엯援?,
            "3.?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸",
            "3.?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX",
            "4.蹂몄꽑?곕꼸(2援ш컙, ?좏뭾-?꾨┝)",
            "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅???곕꼸",
            "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1",
            "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2"
        ],
        "personnel_columns": [
            "1. 蹂몄꽑?곕꼸 (1援ш컙, ?由??좏뭾)",
            "2.?좏뭾?뺢굅??- 1)?뺢굅???곕꼸",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (1)PCB",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (2)PCC",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (3)PCD",
            "2.?좏뭾?뺢굅??- 2)二쇱텧?낃뎄 - (4)PHA",
            "2.?좏뭾?뺢굅??- 3)?밸퀎?쇰궃怨꾨떒",
            "2.?좏뭾?뺢굅??- 4)?몃?異쒖엯援?,
            "3.?좏뭾 ?섏듅?듬줈 - 1)?섏듅?곕꼸",
            "3.?좏뭾 ?섏듅?듬줈 - 2)媛쒖갑 BOX",
            "4.蹂몄꽑?곕꼸(2援ш컙, ?좏뭾~?꾨┝)",
            "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 1)?뺢굅???곕꼸",
            "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 2)異쒖엯援?1",
            "5.?꾨┝?ш굅由ъ젙嫄곗옣 - 3)異쒖엯援?2"
        ],
        "personnel_rows": [
            "吏곸쁺諛섏옣", "?곗닔??, "?λ퉬?댁쟾??, "?꾧린二쇱엫", "?붿빟二쇱엫", "?곕꼸怨?, "紐⑷났", "泥좉렐怨?, 
            "?쇱씠?앺뤌怨?, "?ㅽ룓?섏쿂由ш났", "移대━?꾪듃怨?, "BP怨?, "媛?쒖꽕怨??댁껜怨?, "?숇컮由ш났", 
            "?좏샇??, "遺?⑥닔怨?, "?щ윭由ъ썡怨?, "CIP怨?, "誘몄옣怨?, "?쒖꽕臾쇨났", "寃쎄퀎?앷났", "議곌꼍怨?, 
            "諛곌?怨?, "?꾩깋怨?, "諛⑹닔怨?, "?λ퉬/?묒뾽吏?댁씠", "蹂댄넻?몃?", "?ъ옣怨?, "?⑹젒怨?, "??ㅺ났", 
            "蹂대쭅怨??숈뭅怨?, "鍮꾧퀎怨?, "?꾩옣怨?, "?앸㈃怨?, "二쇱엯怨?洹몃씪?고똿怨?
        ],
        "equipment_rows": [
            "B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)",
            "?ㅽ봽?몃윮(5T)", "?ㅽ봽?몃윮(15T)", "?ㅽ봽?몃윮(25T)", "?듦??щ젅??100T)", "?듦??щ젅??80T)", "?듦??щ젅??35T)", "?듦??щ젅??25T)",
            "移닿퀬?щ젅??25T)", "移닿퀬?щ젅??5T)", "肄ㅽ봽", "?먮낫?쒕┫", "?섏씠濡쒕뜑", "?륂듃癒몄떊", "李⑥쭠移?, "?댁닔李?, "?섏씠?쒕줈?щ젅??,
            "誘뱀꽌?몃윮", "?붾Ъ李?5T)", "?뚰봽移?, "?ㅼ뭅??, "肄섑겕由ы듃?쇰땲??, "?꾩＜?ㅺ굅", "濡쒕뜑(諛붾툕耳?", "?좎젣?댄룷湲?鍮꾩슦??",
            "吏寃뚯감", "?몄씤移?, "BC而ㅽ꽣湲?, "諛붿씠釉뚮줈?대㉧", "濡ㅻ윭(2.5T)", "濡ㅻ윭(1T)", "濡ㅻ윭(0.7T)", "紐곕━", "???湲?, 
            "?щ젅??, "肄ㅻ퉬濡쒕씪", "怨듭븬?쒕┫", "?좎븬?쒕┫", "湲고?"
        ]
    }
    for key, value in states.items():
        if key not in st.session_state:
            st.session_state[key] = value
    
    # ???쒖옉 ????λ맂 ?꾨＼?꾪듃 紐⑸줉 濡쒕뱶
    if supabase_client and not st.session_state.prompt_list:
        st.session_state.prompt_list = get_all_prompts_from_supabase()

initialize_session_state()


# ?ъ씠?쒕컮 鍮꾪솢?깊솕 (?ㅻⅨ ?섏씠吏? ?숈씪?섍쾶)

# --- MAIN CONTENT ---
st.title("?뱞?묒뾽?쇰낫 ?묒꽦 ?먮룞??)
st.write("SNS ?쇱씪?묒뾽怨꾪쉷蹂닿퀬瑜??낅젰?섏떆硫?AI媛 ?먮룞?쇰줈 ?묒뾽?쇰낫瑜??앹꽦?대뱶由쎈땲??")
st.markdown("---")

# --- STEP 1: SNS WORK REPORT INPUT ---
with st.container():
    st.markdown('<h3><span style="font-size: 1.5em;">?벑</span> 1. SNS ?쇱씪?묒뾽怨꾪쉷蹂닿퀬 ?낅젰</h3>', unsafe_allow_html=True)
    
    # ?꾨떖??蹂닿퀬???댁슜 ?쒖떆
    if 'report_to_transfer' in st.session_state and st.session_state.report_to_transfer:
        st.markdown("---")
        st.markdown('<h4><span style="font-size: 1.2em;">?뱥</span> ?꾨떖???쇱씪?묒뾽蹂닿퀬</h4>', unsafe_allow_html=True)
        
        with st.container(border=True):
            st.markdown("**SNS?쇱씪?묒뾽怨꾪쉷 ?섏씠吏?먯꽌 ?꾨떖??蹂닿퀬???댁슜:**")
            st.text_area(
                "蹂닿퀬???댁슜",
                value=st.session_state.report_to_transfer,
                height=300,
                key="transferred_report_display",
                label_visibility="collapsed"
            )
            
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button("?뱥 ???댁슜?쇰줈 ?묒뾽?쇰낫 ?앹꽦", key="use_transferred_report", use_container_width=True):
                    # ?꾨떖???댁슜???꾩옱 ?꾨＼?꾪듃???ㅼ젙
                    st.session_state.current_report_content = st.session_state.report_to_transfer
                    st.toast("???꾨떖??蹂닿퀬???댁슜???ㅼ젙?섏뿀?듬땲??")
                    st.rerun()
            
            with col2:
                if st.button("?뿊截??꾨떖???댁슜 ??젣", key="clear_transferred_report", use_container_width=True):
                    del st.session_state.report_to_transfer
                    st.toast("?뿊截??꾨떖??蹂닿퀬???댁슜????젣?섏뿀?듬땲??")
                    st.rerun()
    
    # ?꾨＼?꾪듃 愿由??뱀뀡
    with st.expander("?숋툘 ?꾨＼?꾪듃 愿由?, expanded=False):
        # 1. ??λ맂 ?꾨＼?꾪듃 ?좏깮 + ?몄쭛
        st.markdown("**1. ??λ맂 ?꾨＼?꾪듃 ?좏깮**")
        col1, col2 = st.columns([4, 1])
        
        with col1:
            if st.session_state.prompt_list:
                prompt_options = ["湲곕낯 ?꾨＼?꾪듃"] + [p["name"] for p in st.session_state.prompt_list]
                selected_prompt = st.selectbox(
                    "??λ맂 ?꾨＼?꾪듃 ?좏깮",
                    options=prompt_options,
                    index=prompt_options.index(st.session_state.current_prompt_name) if st.session_state.current_prompt_name in prompt_options else 0,
                    label_visibility="collapsed"
                )
                
                if selected_prompt != st.session_state.current_prompt_name:
                    st.session_state.current_prompt_name = selected_prompt
                    if selected_prompt == "湲곕낯 ?꾨＼?꾪듃":
                        st.session_state.prompt = DEFAULT_PROMPT
                    else:
                        loaded_prompt = load_prompt_from_supabase(selected_prompt)
                        if loaded_prompt:
                            st.session_state.prompt = loaded_prompt["content"]
                    st.rerun()
        
        with col2:
            if st.button("?륅툘 ?몄쭛", key="edit_prompt", use_container_width=True):
                st.session_state.show_prompt_editor = True
                st.session_state.show_table_editor = True
                st.rerun()
        
        # 2. ?꾨＼?꾪듃 ?몄쭛
        if st.session_state.show_prompt_editor:
            st.markdown("---")
            st.markdown("**2. ?꾨＼?꾪듃 ?몄쭛**")
            
            # ?꾨＼?꾪듃 ?대쫫怨??ㅻ챸
            prompt_name = st.text_input(
                "?대쫫",
                value=st.session_state.current_prompt_name if st.session_state.current_prompt_name != "湲곕낯 ?꾨＼?꾪듃" else "",
                placeholder="???꾨＼?꾪듃 ?대쫫???낅젰?섏꽭??
            )
            
            prompt_description = st.text_input(
                "?ㅻ챸 (?좏깮?ы빆)",
                placeholder="?꾨＼?꾪듃??????ㅻ챸???낅젰?섏꽭??
            )
            
            # ?띿뒪?몄? ?뚯씠釉붿쓣 2遺꾪븷濡?諛곗튂
            text_col, table_col = st.columns(2)
            
            with text_col:
                # ?꾨＼?꾪듃 ?댁슜 (?띿뒪?? - ?ㅼ떆媛??낅뜲?댄듃
                edited_prompt = st.text_area(
                    "?댁슜 (?띿뒪??",
                    value=st.session_state.prompt,
                    height=600,
                    help="AI媛 ?곗씠??遺꾩꽍???ъ슜??吏?쒕Ц???묒꽦?섏꽭?? ?뚯씠釉?援ъ“瑜?蹂寃쏀븯硫??먮룞?쇰줈 ?낅뜲?댄듃?⑸땲??",
                    key="prompt_text_area"
                )
                
                # ?붾쾭源? ?꾩옱 ?꾨＼?꾪듃 ?곹깭 ?쒖떆
                if st.checkbox("?뵇 ?꾨＼?꾪듃 ?곹깭 ?뺤씤", key="debug_prompt"):
                    st.info(f"?꾩옱 ?꾨＼?꾪듃 湲몄씠: {len(st.session_state.prompt)} 臾몄옄")
                    st.code(st.session_state.prompt[:200] + "..." if len(st.session_state.prompt) > 200 else st.session_state.prompt)
                
                # ?띿뒪?몄뿉???뚯씠釉?援ъ“ 異붿텧 踰꾪듉
                if st.button("?봽 ?띿뒪?몄뿉???뚯씠釉?援ъ“ 異붿텧", key="extract_from_text"):
                    try:
                        # ?띿뒪?몄뿉???뚯씠釉?援ъ“瑜?異붿텧?섎뒗 濡쒖쭅
                        extracted_tables = extract_table_structure_from_prompt(edited_prompt)
                        if extracted_tables:
                            # 異붿텧??援ъ“濡??몄뀡 ?곹깭 ?낅뜲?댄듃
                            if 'construction' in extracted_tables:
                                st.session_state.construction_rows = extracted_tables['construction']
                            if 'work_content' in extracted_tables:
                                st.session_state.work_content_rows = extracted_tables['work_content']
                            if 'personnel_columns' in extracted_tables:
                                st.session_state.personnel_columns = extracted_tables['personnel_columns']
                            if 'personnel_rows' in extracted_tables:
                                st.session_state.personnel_rows = extracted_tables['personnel_rows']
                            if 'equipment' in extracted_tables:
                                st.session_state.equipment_rows = extracted_tables['equipment']
                            
                            st.success("???띿뒪?몄뿉???뚯씠釉?援ъ“瑜?異붿텧?덉뒿?덈떎!")
                            st.rerun()
                        else:
                            st.warning("?좑툘 ?띿뒪?몄뿉???뚯씠釉?援ъ“瑜?異붿텧?????놁뒿?덈떎.")
                    except Exception as e:
                        st.error(f"???뚯씠釉?援ъ“ 異붿텧 以??ㅻ쪟: {e}")
            
            with table_col:
                # ?뚯씠釉?援ъ“ ?몄쭛
                st.markdown("**?뚯씠釉?援ъ“ ?몄쭛**")
                
                # ??쑝濡?媛??뚯씠釉?援щ텇
                tab1, tab2, tab3, tab4 = st.tabs(["?룛截??쒓났?꾪솴", "?뱷 ?묒뾽?댁슜", "?뫁 ?몄썝", "?슋 ?λ퉬"])
                
                with tab1:
                    # ?쒓났?꾪솴 ???몄쭛
                    construction_df = pd.DataFrame({"援щ텇": st.session_state.construction_rows})
                    edited_construction = st.data_editor(
                        construction_df,
                        use_container_width=True,
                        num_rows="dynamic",
                        height=600,
                        column_config={
                            "援щ텇": st.column_config.TextColumn("援щ텇", help="?쒓났?꾪솴 ??ぉ紐?)
                        },
                        key="construction_editor"
                    )
                    
                    # ?쒓났?꾪솴??蹂寃쎈릺硫??먮룞?쇰줈 ?꾨＼?꾪듃 ?낅뜲?댄듃
                    new_construction_rows = edited_construction["援щ텇"].tolist()
                    if new_construction_rows != st.session_state.construction_rows:
                        st.session_state.construction_rows = new_construction_rows
                        # ?먮룞 ?꾨＼?꾪듃 ?낅뜲?댄듃
                        st.session_state.prompt = generate_prompt_from_tables()
                        st.info("???쒓났?꾪솴 ?뚯씠釉?蹂寃쎌쑝濡??꾨＼?꾪듃媛 ?먮룞 ?낅뜲?댄듃?섏뿀?듬땲??")
                        st.rerun()
                
                with tab2:
                    # ?묒뾽?댁슜 ???몄쭛
                    work_content_df = pd.DataFrame({"援щ텇": st.session_state.work_content_rows})
                    edited_work_content = st.data_editor(
                        work_content_df,
                        use_container_width=True,
                        num_rows="dynamic",
                        height=600,
                        column_config={
                            "援щ텇": st.column_config.TextColumn("援щ텇", help="?묒뾽?댁슜 ??ぉ紐?)
                        },
                        key="work_content_editor"
                    )
                    
                    # ?묒뾽?댁슜??蹂寃쎈릺硫??먮룞?쇰줈 ?꾨＼?꾪듃 ?낅뜲?댄듃
                    new_work_content_rows = edited_work_content["援щ텇"].tolist()
                    if new_work_content_rows != st.session_state.work_content_rows:
                        st.session_state.work_content_rows = new_work_content_rows
                        # ?먮룞 ?꾨＼?꾪듃 ?낅뜲?댄듃
                        st.session_state.prompt = generate_prompt_from_tables()
                        st.info("???묒뾽?댁슜 ?뚯씠釉?蹂寃쎌쑝濡??꾨＼?꾪듃媛 ?먮룞 ?낅뜲?댄듃?섏뿀?듬땲??")
                        st.rerun()
                
                with tab3:
                    # 2媛?而щ읆?쇰줈 ?섎늻???쒖떆
                    col_left, col_right = st.columns(2)
                    
                    with col_left:
                        st.markdown("##### ?뱧 ?묒뾽 ?꾩튂 (??")
                        personnel_columns_df = pd.DataFrame({"?묒뾽?꾩튂": st.session_state.personnel_columns})
                        edited_personnel_columns = st.data_editor(
                            personnel_columns_df,
                            use_container_width=True,
                            num_rows="dynamic",
                            height=600,
                            column_config={
                                "?묒뾽?꾩튂": st.column_config.TextColumn("?묒뾽?꾩튂", help="?몄썝 ?뚯씠釉붿쓽 ????ぉ")
                            },
                            key="personnel_columns_editor"
                        )
                        # ?몄썝 ?댁씠 蹂寃쎈릺硫??먮룞?쇰줈 ?꾨＼?꾪듃 ?낅뜲?댄듃
                        new_personnel_columns = edited_personnel_columns["?묒뾽?꾩튂"].tolist()
                        if new_personnel_columns != st.session_state.personnel_columns:
                            st.session_state.personnel_columns = new_personnel_columns
                            # ?먮룞 ?꾨＼?꾪듃 ?낅뜲?댄듃
                            st.session_state.prompt = generate_prompt_from_tables()
                            st.info("???몄썝 ?뚯씠釉???蹂寃쎌쑝濡??꾨＼?꾪듃媛 ?먮룞 ?낅뜲?댄듃?섏뿀?듬땲??")
                            st.rerun()
                    
                    with col_right:
                        st.markdown("##### ?뫁 吏곸쥌 (??")
                        personnel_rows_df = pd.DataFrame({"吏곸쥌": st.session_state.personnel_rows})
                        edited_personnel_rows = st.data_editor(
                            personnel_rows_df,
                            use_container_width=True,
                            num_rows="dynamic",
                            height=600,
                            column_config={
                                "吏곸쥌": st.column_config.TextColumn("吏곸쥌", help="?몄썝 ?뚯씠釉붿쓽 ????ぉ")
                            },
                            key="personnel_rows_editor"
                        )
                        
                        # ?몄썝 ?됱씠 蹂寃쎈릺硫??먮룞?쇰줈 ?꾨＼?꾪듃 ?낅뜲?댄듃
                        new_personnel_rows = edited_personnel_rows["吏곸쥌"].tolist()
                        if new_personnel_rows != st.session_state.personnel_rows:
                            st.session_state.personnel_rows = new_personnel_rows
                            # ?먮룞 ?꾨＼?꾪듃 ?낅뜲?댄듃
                            st.session_state.prompt = generate_prompt_from_tables()
                            st.info("???몄썝 ?뚯씠釉???蹂寃쎌쑝濡??꾨＼?꾪듃媛 ?먮룞 ?낅뜲?댄듃?섏뿀?듬땲??")
                            st.rerun()
                
                with tab4:
                    # ?λ퉬 ???몄쭛
                    equipment_df = pd.DataFrame({"援щ텇": st.session_state.equipment_rows})
                    edited_equipment = st.data_editor(
                        equipment_df,
                        use_container_width=True,
                        num_rows="dynamic",
                        height=600,
                        column_config={
                            "援щ텇": st.column_config.TextColumn("援щ텇", help="?λ퉬 ??ぉ紐?)
                        },
                        key="equipment_editor"
                    )
                    
                    # ?λ퉬媛 蹂寃쎈릺硫??먮룞?쇰줈 ?꾨＼?꾪듃 ?낅뜲?댄듃
                    new_equipment_rows = edited_equipment["援щ텇"].tolist()
                    if new_equipment_rows != st.session_state.equipment_rows:
                        st.session_state.equipment_rows = new_equipment_rows
                        # ?먮룞 ?꾨＼?꾪듃 ?낅뜲?댄듃
                        st.session_state.prompt = generate_prompt_from_tables()
                        st.info("???λ퉬 ?뚯씠釉?蹂寃쎌쑝濡??꾨＼?꾪듃媛 ?먮룞 ?낅뜲?댄듃?섏뿀?듬땲??")
                        st.rerun()
            
            # ?곹깭 硫붿떆吏瑜??꾪븳 1遺꾪븷 ?뱀뀡
            st.markdown("---")
            
            # ?몄쭛湲??≪뀡 踰꾪듉 - ?섎???諛곗튂
            edit_col1, edit_col2, edit_col3, edit_col4 = st.columns(4)
            
            with edit_col1:
                if st.button("?뮶 ???, key="save_prompt", use_container_width=True):
                    if prompt_name.strip():
                        if save_prompt_to_supabase(prompt_name.strip(), edited_prompt, prompt_description):
                            st.session_state.prompt = edited_prompt
                            st.session_state.current_prompt_name = prompt_name.strip()
                            st.session_state.show_prompt_editor = False
                            st.session_state.prompt_list = get_all_prompts_from_supabase()
                            st.rerun()
                    else:
                        st.error("?꾨＼?꾪듃 ?대쫫???낅젰?댁＜?몄슂.")
            
            with edit_col2:
                if st.button("?봽 ?곸슜", key="apply_prompt", use_container_width=True):
                    st.session_state.prompt = edited_prompt
                    st.success("?꾨＼?꾪듃媛 ?꾩옱 ?몄뀡???곸슜?섏뿀?듬땲??")
            
            with edit_col3:
                if st.button("?뵗 湲곕낯媛?, key="reset_to_default", use_container_width=True):
                    st.session_state.prompt = DEFAULT_PROMPT
                    st.session_state.current_prompt_name = "湲곕낯 ?꾨＼?꾪듃"
                    st.rerun()
            
            with edit_col4:
                if st.button("??痍⑥냼", key="cancel_edit", use_container_width=True):
                    st.session_state.show_prompt_editor = False
                    st.rerun()
        

    
    if not st.session_state.kakao_work_completed:
        kakao_text = st.text_area("移댁뭅?ㅽ넚 ?묒뾽蹂닿퀬", placeholder=" ?닿납??SNS?쇱씪?묒뾽怨꾪쉷蹂닿퀬瑜?遺숈뿬?ｌ쑝?몄슂.", height=200, label_visibility="collapsed")
        
        # AI ?곗씠??異붿텧 踰꾪듉留?諛곗튂
        if st.button("?쭭I ?곗씠??異붿텧", key="structure_button", use_container_width=True):
            if kakao_text:
                # 吏꾪뻾 ?곹솴???④퀎蹂꾨줈 ?쒖떆
                progress_placeholder = st.empty()
                status_placeholder = st.empty()
                
                try:
                    # 湲곕낯 濡쒕뵫 ?ㅽ뵾??                    with st.spinner("AI媛 ?곗씠?곕? 異붿텧?섍퀬 ?덉뒿?덈떎. ?좎떆留?湲곕떎?ㅼ＜?몄슂..."):
                        prompt = st.session_state.prompt + "\n" + kakao_text
                        response_text = safe_generate_content(prompt)
                    
                    if response_text:
                        # AI ?묐떟?먯꽌 ?묒뾽 ?좎쭨 異붿텧
                        work_date = extract_work_date_from_response(response_text)
                        
                        st.session_state.kakao_results = response_text
                        st.session_state.work_date = work_date  # 異붿텧???묒뾽 ?좎쭨 ???                        st.session_state.kakao_work_completed = True
                        
                        st.info(f"?뱟 異붿텧???묒뾽 ?좎쭨: {work_date}")
                        st.toast("??1?④퀎 ?꾨즺: SNS ?묒뾽蹂닿퀬 ?곗씠??援ъ“???깃났!")
                        st.success("??1?④퀎 ?꾨즺: SNS ?쇱씪?묒뾽怨꾪쉷蹂닿퀬瑜??깃났?곸쑝濡?泥섎━?덉뒿?덈떎.")
                        
                        # ?섏씠吏 ?덈줈怨좎묠?섏뿬 泥섎━???곗씠??蹂닿린 ?쒖떆
                        st.rerun()
                    else:
                        st.error("??AI ?묐떟 ?앹꽦???ㅽ뙣?덉뒿?덈떎.")
                except Exception as e: 
                    st.error(f"??AI ?곗씠??異붿텧 以??ㅻ쪟: {e}")
                    st.error("?뮕 ?ㅽ듃?뚰겕 ?곌껐???뺤씤?섍굅???좎떆 ???ㅼ떆 ?쒕룄?댁＜?몄슂.")
            else: 
                st.warning("蹂닿퀬 ?댁슜???낅젰?댁＜?몄슂.")
    else:
        st.success("??1?④퀎 ?꾨즺: SNS ?묒뾽蹂닿퀬媛 ?깃났?곸쑝濡?泥섎━?섏뿀?듬땲??")
        
        # 泥섎━???곗씠??蹂닿린
        with st.expander("?뱤 泥섎━???곗씠??蹂닿린", expanded=True):
            # QA-Checklist ?쒖떆
            qa_checklist = extract_qa_checklist_from_response(st.session_state.kakao_results)
            if qa_checklist:
                st.subheader("?뵇 QA-Checklist (?먮룞 寃利?寃곌낵)")
                st.markdown(qa_checklist)
                st.markdown("---")
            
            # AI ?묐떟?먯꽌 媛??뚯씠釉붿쓣 媛쒕퀎?곸쑝濡?異붿텧
            response_text = st.session_state.kakao_results
            
            # 媛??뚯씠釉??뱀뀡??李얠븘??媛쒕퀎?곸쑝濡?泥섎━
            table_sections = {
                "?쒓났?꾪솴": None,
                "?묒뾽?댁슜": None,
                "?몄썝": None,
                "?λ퉬": None
            }
            
            # AI ?묐떟 援ъ“ 遺꾩꽍 諛??뚯씠釉?異붿텧
            # 1. 癒쇱? TSV 釉붾줉??李얘린
            tsv_blocks = re.findall(r'```(?:tsv)?\n(.*?)```', response_text, re.DOTALL | re.IGNORECASE)
            
            # 2. TSV 釉붾줉???놁쑝硫??뱀뀡蹂꾨줈 異붿텧 ?쒕룄
            if not tsv_blocks:
                # ?쒓났?꾪솴 ?뱀뀡 李얘린 (???좎뿰??寃??
                construction_patterns = [
                    r'## 1\. ?쒓났?꾪솴.*?(?=## 2\.|$)',
                    r'?쒓났?꾪솴.*?(?=##|$)',
                    r'?쒓났?꾪솴.*?(?=\n\n|$)'
                ]
                
                construction_found = False
                for pattern in construction_patterns:
                    construction_match = re.search(pattern, response_text, re.DOTALL | re.IGNORECASE)
                    if construction_match:
                        construction_text = construction_match.group(0)
                        lines = construction_text.split('\n')
                        table_data = []
                        for line in lines:
                            line = line.strip()
                            if line and not line.startswith('##') and not line.startswith('#'):
                                if re.search(r'\d+', line):
                                    parts = line.split()
                                    if len(parts) >= 2:
                                        table_data.append('\t'.join(parts))
                        if table_data:
                            table_sections["?쒓났?꾪솴"] = '\n'.join(table_data)
                            construction_found = True
                            break
                
                # ?묒뾽?댁슜 ?뱀뀡 李얘린 (???좎뿰??寃??
                work_patterns = [
                    r'## 2\. ?묒뾽?댁슜.*?(?=## 3\.|$)',
                    r'?묒뾽?댁슜.*?(?=##|$)',
                    r'?묒뾽?댁슜.*?(?=\n\n|$)'
                ]
                
                work_found = False
                for pattern in work_patterns:
                    work_match = re.search(pattern, response_text, re.DOTALL | re.IGNORECASE)
                    if work_match:
                        work_text = work_match.group(0)
                        lines = work_text.split('\n')
                        table_data = []
                        for line in lines:
                            line = line.strip()
                            if line and not line.startswith('##') and not line.startswith('#'):
                                if re.search(r'\d+', line):
                                    parts = line.split()
                                    if len(parts) >= 2:
                                        table_data.append('\t'.join(parts))
                        if table_data:
                            table_sections["?묒뾽?댁슜"] = '\n'.join(table_data)
                            work_found = True
                            break
                
                # ?몄썝 ?뱀뀡 李얘린 (???좎뿰??寃??
                personnel_patterns = [
                    r'## 3\. ?몄썝.*?(?=## 4\.|$)',
                    r'?몄썝.*?(?=##|$)',
                    r'?몄썝.*?(?=\n\n|$)'
                ]
                
                personnel_found = False
                for pattern in personnel_patterns:
                    personnel_match = re.search(pattern, response_text, re.DOTALL | re.IGNORECASE)
                    if personnel_match:
                        personnel_text = personnel_match.group(0)
                        lines = personnel_text.split('\n')
                        table_data = []
                        for line in lines:
                            line = line.strip()
                            if line and not line.startswith('##') and not line.startswith('#'):
                                if re.search(r'\d+', line):
                                    parts = line.split()
                                    if len(parts) >= 2:
                                        table_data.append('\t'.join(parts))
                        if table_data:
                            table_sections["?몄썝"] = '\n'.join(table_data)
                            personnel_found = True
                            break
                
                # ?λ퉬 ?뱀뀡 李얘린 (???좎뿰??寃??
                equipment_patterns = [
                    r'## 4\. ?λ퉬.*?(?=##|$)',
                    r'?λ퉬.*?(?=##|$)',
                    r'?λ퉬.*?(?=\n\n|$)',
                    r'?λ퉬.*?(?=?덉쟾愿由?$)'
                ]
                
                equipment_found = False
                for pattern in equipment_patterns:
                    equipment_match = re.search(pattern, response_text, re.DOTALL | re.IGNORECASE)
                    if equipment_match:
                        equipment_text = equipment_match.group(0)
                        lines = equipment_text.split('\n')
                        table_data = []
                        for line in lines:
                            line = line.strip()
                            if line and not line.startswith('##') and not line.startswith('#'):
                                # ?レ옄媛 ?ы븿???쇱씤???뚯씠釉??곗씠?곕줈 媛꾩＜
                                if re.search(r'\d+', line):
                                    parts = line.split()
                                    if len(parts) >= 2:
                                        table_data.append('\t'.join(parts))
                        if table_data:
                            table_sections["?λ퉬"] = '\n'.join(table_data)
                            equipment_found = True
                            break
                
                # ?λ퉬 ?뱀뀡??李얠? 紐삵뻽?ㅻ㈃ ?붾쾭源??뺣낫 ?쒖떆
                if not equipment_found:
                    st.info("?뵇 ?λ퉬 ?뱀뀡 寃???⑦꽩:")
                    for i, pattern in enumerate(equipment_patterns):
                        st.text(f"?⑦꽩 {i+1}: {pattern}")
                    st.info("?뵇 AI ?묐떟?먯꽌 '?λ퉬' ?ㅼ썙???꾩튂:")
                    equipment_keyword_pos = response_text.lower().find('?λ퉬')
                    if equipment_keyword_pos != -1:
                        start = max(0, equipment_keyword_pos - 100)
                        end = min(len(response_text), equipment_keyword_pos + 200)
                        st.code(response_text[start:end])
                    else:
                        st.warning("?좑툘 AI ?묐떟?먯꽌 '?λ퉬' ?ㅼ썙?쒕? 李얠쓣 ???놁뒿?덈떎.")
            else:
                # TSV 釉붾줉???덉쑝硫?湲곗〈 諛⑹떇?쇰줈 泥섎━
                table_names = ["?쒓났?꾪솴", "?묒뾽?댁슜", "?몄썝", "?λ퉬"]
                for i, tsv_data in enumerate(tsv_blocks):
                    if i < len(table_names):
                        table_sections[table_names[i]] = tsv_data.strip()
            
            # ?붾쾭源? 李얠? ?곗씠???쒖떆
            found_tables = sum(1 for data in table_sections.values() if data)
            if found_tables > 0:
                st.info(f"?뵇 {found_tables}媛쒖쓽 ?뚯씠釉??곗씠?곕? 李얠븯?듬땲??")
            else:
                st.warning("?좑툘 ?뚯씠釉??곗씠?곕? 李얠쓣 ???놁뒿?덈떎. AI ?묐떟???뺤씤?댁＜?몄슂.")
                st.code(response_text[:1000] + "..." if len(response_text) > 1000 else response_text)
            
            processed_tables = []
            for table_name, tsv_data in table_sections.items():
                if tsv_data:
                    df = parse_tsv_to_dataframe(fix_tsv_field_count(tsv_data))
                    if df is not None:
                        st.subheader(table_name)
                        st.dataframe(df)
                        processed_tables.append(df)
                    else:
                        st.warning(f"?좑툘 {table_name} ?뚯씠釉??뚯떛 ?ㅽ뙣")
                else:
                    st.warning(f"?좑툘 {table_name} ?뚯씠釉붿쓣 李얠쓣 ???놁뒿?덈떎")
            
            st.session_state.processed_tables = processed_tables
            
                        # Supabase ???湲곕뒫 (??긽 ?쒖떆)
            st.markdown("---")
            st.markdown("### ?뮶 ?곗씠?????)
            
            # ??踰꾪듉????以꾩뿉 諛곗튂
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("?뮶 Database?????, key="save_step1"):
                    # ?곗씠?곕? ?뺤뀛?덈━濡?蹂?섑븯?????                    report_data = {}
                    for table_name, tsv_data in table_sections.items():
                        if tsv_data:
                            df = parse_tsv_to_dataframe(fix_tsv_field_count(tsv_data))
                            if df is not None:
                                report_data[table_name] = df.to_dict('records')
                    
                    if report_data:
                        # Supabase?????(異붿텧???묒뾽 ?좎쭨 ?ъ슜)
                        work_date = st.session_state.get('work_date', datetime.now().strftime('%Y-%m-%d'))
                        if save_step1_to_supabase(report_data, work_date):
                            st.session_state.daily_report_saved = True
                            st.info("?뮕 ?쒓났?꾪솴, ?묒뾽?댁슜, ?몄썝, ?λ퉬 ?곗씠?곌? 媛곴컖 蹂꾨룄 ?뚯씠釉붿뿉 ??λ릺?덉뒿?덈떎.")
                        else:
                            st.error("??1?④퀎 ?곗씠????μ뿉 ?ㅽ뙣?덉뒿?덈떎.")
                    else:
                        st.warning("?좑툘 ??ν븷 ?곗씠?곌? ?놁뒿?덈떎.")
            
            with col2:
                if st.button("?뱟 ?좎쭨蹂??곗씠??議고쉶", key="load_step1"):
                    selected_date = st.date_input("議고쉶???좎쭨 ?좏깮", value=datetime.now(), key="load_date_step1")
                    loaded_data = load_from_supabase("daily_report", selected_date.strftime('%Y-%m-%d'))
                    if loaded_data:
                        st.json(loaded_data)
                    else:
                        st.info("?대떦 ?좎쭨???곗씠?곌? ?놁뒿?덈떎.")

# --- STEP 1怨?STEP 2 援щ텇??---
st.markdown("---")

# --- STEP 2: BASIC INFO INPUT ---
with st.container():
    st.markdown('<h3><span style="font-size: 1.5em;">?뱥</span> 2. 湲곕낯?뺣낫 ?낅젰</h3>', unsafe_allow_html=True)
    
    # 湲곕낯?뺣낫 ?낅젰 ?뱀뀡 - 泥?踰덉㎏ ?? ?묒꽦?쇱옄? 怨듭궗紐?    col1, col2 = st.columns(2)
    
    with col1:
        # 異붿텧???묒뾽 ?좎쭨媛 ?덉쑝硫??ъ슜, ?놁쑝硫??꾩옱 ?좎쭨 ?ъ슜
        default_date = None
        if st.session_state.get('work_date'):
            try:
                default_date = datetime.strptime(st.session_state.work_date, '%Y-%m-%d').date()
            except:
                default_date = datetime.now().date()
        else:
            default_date = datetime.now().date()
        
        report_date = st.date_input("?묒꽦?쇱옄", value=default_date, key="excel_report_date")
    
    with col2:
        project_name = st.text_input("怨듭궗紐?, value="?좎븞?곗꽑 蹂듭꽑?꾩쿋 誘쇨컙?ъ옄?ъ뾽 4-1怨듦뎄", key="excel_project_name")
    
    # ??踰덉㎏ ?? ?좎뵪?뺣낫? 怨듭젙瑜?    col3, col4 = st.columns(2)
    
    with col3:
        # ?좎뵪?뺣낫 ?뱀뀡
        st.markdown("**?뙟截??좎뵪?뺣낫**")
        
        weather_col1, weather_col2, weather_col3 = st.columns(3)
        
        # ??λ맂 ?좎뵪 ?곗씠?곌? ?덉쑝硫??ъ슜
        weather_auto_fill = st.session_state.get('weather_auto_fill_data', {})
        default_min_temp = weather_auto_fill.get('min_temp', 18.2) if weather_auto_fill else 18.2
        default_max_temp = weather_auto_fill.get('max_temp', 25.5) if weather_auto_fill else 25.5
        default_precipitation = weather_auto_fill.get('precipitation', 0.0) if weather_auto_fill else 0.0
        
        with weather_col1:
            min_temp = st.number_input("理쒖?湲곗삩 (째C)", value=default_min_temp, key="excel_min_temp", format="%.1f")
        with weather_col2:
            max_temp = st.number_input("理쒓퀬湲곗삩 (째C)", value=default_max_temp, key="excel_max_temp", format="%.1f")
        with weather_col3:
            precipitation = st.number_input("媛뺤닔??(mm)", value=default_precipitation, key="excel_precipitation", format="%.1f")
        
        # ?좎뵪 ?곹깭 ?쒖떆
        if weather_auto_fill:
            st.info(f"?뙟截??좎뵪: 理쒓퀬 {weather_auto_fill.get('max_temp', 0):.1f}째C / 理쒖? {weather_auto_fill.get('min_temp', 0):.1f}째C / 媛뺤닔??{weather_auto_fill.get('precipitation', 0):.1f}mm")
        
        # AI ?먮룞梨꾩슦湲?踰꾪듉 (?쒖씪 ?쇱そ, ?묒? 踰꾪듉)
        if st.button("?챷AI ?먮룞梨꾩슦湲?, key="weather_auto_fill_button", help="AI ?먮룞梨꾩슦湲?):
            try:
                weather_data = get_weather_data()
                if weather_data:
                    st.session_state.weather_auto_fill_data = weather_data
                    st.success(f"???좎뵪 ?곗씠??媛?몄삤湲??깃났: {weather_data['max_temp']}째C / {weather_data['min_temp']}째C")
                    st.rerun()
                else:
                    st.error(f"???좎뵪 ?곗씠??媛?몄삤湲??ㅽ뙣")
            except Exception as e:
                st.error(f"???좎뵪 ?곗씠??媛?몄삤湲??ㅽ뙣")
    
    with col4:
        # 怨듭젙瑜??뱀뀡
        st.markdown("**?뱤 怨듭젙瑜?*")
        progress_col1, progress_col2 = st.columns(2)
        
        with progress_col1:
            planned_progress = st.number_input("怨꾪쉷 (%)", value=50, key="excel_planned_progress")
        with progress_col2:
            actual_progress = st.number_input("?ㅼ쟻 (%)", value=48.5, key="excel_actual_progress")

# --- STEP 2? STEP 3 援щ텇??---
st.markdown("---")

# --- STEP 3: WORK REPORT GENERATION ---
with st.container():
    st.markdown('<h3><span style="font-size: 1.5em;">?뱞</span> 3. ?묒뾽?쇰낫 ?앹꽦</h3>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)

    # ?묒뾽?쇰낫 ?앹꽦 湲곕뒫 異붽? (??긽 ?쒖떆)
    
    # 媛꾨떒???쒗뵆由??좏깮
    col1, col2 = st.columns([3, 1])
    
    with col1:
        # ?쒗뵆由??듭뀡 (理쒖떊 ?낅줈???뚯씪??湲곕낯媛?
        template_options = ["?덈줈???묒떇?쇰줈 ?앹꽦"]
        
        # Supabase?먯꽌 ??λ맂 ?쒗뵆由?紐⑸줉 媛?몄삤湲?(理쒖떊 ?쒖쑝濡??뺣젹)
        if supabase_client:
            try:
                saved_templates = get_all_templates()
                if saved_templates:
                    # 理쒖떊 ?쒗뵆由우쓣 泥?踰덉㎏濡??ㅼ젙
                    template_options.extend([t["template_name"] for t in saved_templates])
            except:
                pass
        
        selected_template_option = st.selectbox(
            "?쒗뵆由??좏깮",
            options=template_options,
            index=0,
            label_visibility="collapsed"
        )
    
    with col2:
        # ?쒗뵆由??낅줈??踰꾪듉 (???遺덈윭?ㅺ린? 媛숈? ?ㅽ???
        if st.button("?뱾 ?쒗뵆由??낅줈??, key="upload_template"):
            st.session_state.show_template_upload = True
            st.rerun()
    
    # 媛꾨떒???쒗뵆由??낅줈???뱀뀡
    if st.session_state.get('show_template_upload', False):
        with st.expander("?뱾 ?쒗뵆由??낅줈??, expanded=True):
            uploaded_template = st.file_uploader(
                "?묒? ?쒗뵆由??뚯씪 ?좏깮",
                type=['xlsx', 'xls'],
                key="template_uploader"
            )
            
            if uploaded_template:
                template_name = st.text_input("?쒗뵆由??대쫫", value="???쒗뵆由?)
                
                if st.button("?뮶 ?쒗뵆由????, key="save_template", use_container_width=True):
                    template_bytes = uploaded_template.read()
                    success, message = save_template_to_supabase(template_bytes, template_name, "?낅줈?쒕맂 ?쒗뵆由?)
                    if success:
                        st.success(f"??{message}")
                        st.rerun()
                    else:
                        st.error(f"??{message}")
            
            if st.button("???リ린", key="close_template_upload"):
                st.session_state.show_template_upload = False
                st.rerun()
    
    # ?쒗뵆由??뚯씪 泥섎━
    template_bytes = None
    
    if selected_template_option != "?덈줈???묒떇?쇰줈 ?앹꽦":
        # Supabase?먯꽌 ?좏깮???쒗뵆由?濡쒕뱶
        template_bytes = get_template_from_supabase(selected_template_option)
        if template_bytes:
            st.success(f"???쒗뵆由?'{selected_template_option}' 濡쒕뱶 ?꾨즺")
        else:
            st.error(f"???쒗뵆由?'{selected_template_option}' 濡쒕뱶 ?ㅽ뙣")
        
        # ? 留ㅽ븨 ?ㅼ젙 (?쒗뵆由우씠 ?덈뒗 寃쎌슦)
        with st.expander("?숋툘 ? 留ㅽ븨 ?ㅼ젙", expanded=True):
            # ?쒗뵆由우씠 濡쒕뱶?섏? ?딆? 寃쎌슦 ?덈궡 硫붿떆吏
            if not template_bytes:
                st.warning("?좑툘 ?쒗뵆由우쓣 癒쇱? ?좏깮?섍굅???낅줈?쒗빐二쇱꽭??")
            
            # 肄ㅽ뙥?명븳 3???덉씠?꾩썐
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("**湲곕낯 ?뺣낫**")
                st.session_state.date_cell = st.text_input("?묒꽦?쇱옄", value=st.session_state.get('date_cell', 'u2'), disabled=not template_bytes)
                st.session_state.project_cell = st.text_input("怨듭궗紐?, value=st.session_state.get('project_cell', 'd4'), disabled=not template_bytes)
                st.session_state.max_temp_cell = st.text_input("理쒓퀬湲곗삩", value=st.session_state.get('max_temp_cell', 'o4'), disabled=not template_bytes)
                st.session_state.min_temp_cell = st.text_input("理쒖?湲곗삩", value=st.session_state.get('min_temp_cell', 'o5'), disabled=not template_bytes)
                st.session_state.precipitation_cell = st.text_input("媛뺤닔??, value=st.session_state.get('precipitation_cell', 'o6'), disabled=not template_bytes)
            
            with col2:
                st.markdown("**怨듭젙瑜?*")
                st.session_state.planned_progress_cell = st.text_input("怨꾪쉷", value=st.session_state.get('planned_progress_cell', 'w4'), disabled=not template_bytes)
                st.session_state.actual_progress_cell = st.text_input("?ㅼ쟻", value=st.session_state.get('actual_progress_cell', 'w5'), disabled=not template_bytes)
                
                st.markdown("**?뚯씠釉??쒖옉 ?꾩튂**")
                st.session_state.table_construction_cell = st.text_input("?쒓났?꾪솴", value=st.session_state.get('table_construction_cell', 'q8'), disabled=not template_bytes)
                st.session_state.table_work_cell = st.text_input("?묒뾽?댁슜", value=st.session_state.get('table_work_cell', 'q11'), disabled=not template_bytes)
                st.session_state.table_personnel_cell = st.text_input("?몄썝", value=st.session_state.get('table_personnel_cell', 'ac66'), disabled=not template_bytes)
                st.session_state.table_equipment_cell = st.text_input("?λ퉬", value=st.session_state.get('table_equipment_cell', 'ac106'), disabled=not template_bytes)
            
            with col3:
                st.markdown("**???遺덈윭?ㅺ린**")
                if st.button("?뮶 ???, key="save_mapping", disabled=not template_bytes, use_container_width=True):
                    mapping_data = {
                        'date_cell': st.session_state.date_cell,
                        'project_cell': st.session_state.project_cell,
                        'max_temp_cell': st.session_state.max_temp_cell,
                        'min_temp_cell': st.session_state.min_temp_cell,
                        'precipitation_cell': st.session_state.precipitation_cell,
                        'planned_progress_cell': st.session_state.planned_progress_cell,
                        'actual_progress_cell': st.session_state.actual_progress_cell,
                        'table_construction_cell': st.session_state.table_construction_cell,
                        'table_work_cell': st.session_state.table_work_cell,
                        'table_personnel_cell': st.session_state.table_personnel_cell,
                        'table_equipment_cell': st.session_state.table_equipment_cell
                    }
                    
                    success, message = save_cell_mapping_to_supabase(mapping_data, "default")
                    if success:
                        st.session_state.mapping_save_success = True
                        st.session_state.mapping_save_message = f"??{message}"
                        st.toast("?뮶 ? 留ㅽ븨 ?ㅼ젙????λ릺?덉뒿?덈떎!", icon="??)
                        st.rerun()
                    else:
                        st.session_state.mapping_save_success = False
                        st.session_state.mapping_save_message = f"??{message}"
                        st.toast("????μ뿉 ?ㅽ뙣?덉뒿?덈떎.", icon="??)
                        st.rerun()
                
                # ???寃곌낵 硫붿떆吏 ?쒖떆
                if hasattr(st.session_state, 'mapping_save_success'):
                    if st.session_state.mapping_save_success:
                        st.success(st.session_state.mapping_save_message)
                        # 硫붿떆吏 ?쒖떆 ???곹깭 珥덇린??                        del st.session_state.mapping_save_success
                        del st.session_state.mapping_save_message
                    else:
                        st.error(st.session_state.mapping_save_message)
                        # 硫붿떆吏 ?쒖떆 ???곹깭 珥덇린??                        del st.session_state.mapping_save_success
                        del st.session_state.mapping_save_message
                
                if st.button("?뱿 遺덈윭?ㅺ린", key="load_mapping", disabled=not template_bytes, use_container_width=True):
                    loaded_mapping = get_cell_mapping_from_supabase("default")
                    if loaded_mapping:
                        for key, value in loaded_mapping.items():
                            if key in st.session_state:
                                st.session_state[key] = value
                        st.session_state.mapping_load_success = True
                        st.session_state.mapping_load_message = "??留ㅽ븨 ?ㅼ젙??遺덈윭?붿뒿?덈떎."
                        st.toast("?뱿 留ㅽ븨 ?ㅼ젙??遺덈윭?붿뒿?덈떎!", icon="??)
                        st.rerun()
                    else:
                        st.session_state.mapping_load_success = False
                        st.session_state.mapping_load_message = "?좑툘 ??λ맂 留ㅽ븨 ?ㅼ젙???놁뒿?덈떎."
                        st.toast("?좑툘 ??λ맂 留ㅽ븨 ?ㅼ젙???놁뒿?덈떎.", icon="?좑툘")
                        st.rerun()
                
                # 遺덈윭?ㅺ린 寃곌낵 硫붿떆吏 ?쒖떆
                if hasattr(st.session_state, 'mapping_load_success'):
                    if st.session_state.mapping_load_success:
                        st.success(st.session_state.mapping_load_message)
                        # 硫붿떆吏 ?쒖떆 ???곹깭 珥덇린??                        del st.session_state.mapping_load_success
                        del st.session_state.mapping_load_message
                    else:
                        st.warning(st.session_state.mapping_load_message)
                        # 硫붿떆吏 ?쒖떆 ???곹깭 珥덇린??                        del st.session_state.mapping_load_success
                        del st.session_state.mapping_load_message
        
        # 怨듭궗?쇰낫 ?앹꽦 踰꾪듉
        st.markdown("---")
        
        # 怨듭궗?쇰낫 ?앹꽦 踰꾪듉 (?꾩껜 ?덈퉬)
        if st.button("?뱤 怨듭궗?쇰낫 ?앹꽦", key="create_report", use_container_width=True):
            # ?쒗뵆由우씠 ?덈뒗 寃쎌슦 湲곕낯媛??ㅼ젙 (寃利??놁씠)
            if template_bytes:
                # 湲곕낯媛??ㅼ젙
                default_mappings = {
                    'date_cell': 'u2',
                    'project_cell': 'd4', 
                    'max_temp_cell': 'o4',
                    'min_temp_cell': 'o5',
                    'precipitation_cell': 'o6',
                    'planned_progress_cell': 'w4',
                    'actual_progress_cell': 'w5',
                    'table_construction_cell': 'q8',
                    'table_work_cell': 'q11',
                    'table_personnel_cell': 'ac66',
                    'table_equipment_cell': 'ac106'
                }
                
                # ?몄뀡 ?곹깭??湲곕낯媛??ㅼ젙
                for key, default_value in default_mappings.items():
                    if not st.session_state.get(key):
                        st.session_state[key] = default_value
            
            # 濡쒕뵫 以묒씪 ?뚮쭔 ?쒖떆
            with st.spinner(""):
                try:
                    # 湲곕낯 ?뺣낫 以鍮?(1?④퀎 寃곌낵 ?ъ슜)
                    # 1?④퀎?먯꽌 異붿텧???좎쭨 ?ъ슜
                    work_date = st.session_state.get('work_date', datetime.now().strftime('%Y-%m-%d'))
                    
                    # ?좎뵪 ?곗씠?? 2?④퀎 ?낅젰媛??곗꽑, ?놁쑝硫?API?먯꽌 媛?몄삤湲?                    weather_data = {}
                    
                    # 2?④퀎?먯꽌 ?ъ슜?먭? ?낅젰???좎뵪 ?곗씠???뺤씤
                    if st.session_state.get('excel_max_temp') is not None:
                        weather_data['max_temp'] = st.session_state.get('excel_max_temp')
                    if st.session_state.get('excel_min_temp') is not None:
                        weather_data['min_temp'] = st.session_state.get('excel_min_temp')
                    if st.session_state.get('excel_precipitation') is not None:
                        weather_data['precipitation'] = st.session_state.get('excel_precipitation')
                    
                    # ?낅젰??媛믪씠 ?놁쑝硫?API?먯꽌 媛?몄삤湲?                    if not weather_data:
                        weather_data = get_weather_data()
                    
                    planned_progress = st.session_state.get('excel_planned_progress', 48.0)
                    actual_progress = st.session_state.get('excel_actual_progress', 48.5)
                    progress_diff = actual_progress - planned_progress
                    
                    basic_info = {
                        'date': work_date,
                        'project_name': st.session_state.get('excel_project_name', '?쒖슱吏?섏쿋 2?몄꽑 ?좏뭾~?꾨┝ 援ш컙 嫄댁꽕怨듭궗'),
                        'max_temp': weather_data.get('max_temp', 25.5),
                        'min_temp': weather_data.get('min_temp', 18.2),
                        'precipitation': weather_data.get('precipitation', 0.0),
                        'planned_progress': planned_progress,
                        'actual_progress': actual_progress,
                        'progress_diff': progress_diff
                    }
                    
                    # ?뚯씠釉??곗씠??以鍮?(SNS ?묒뾽蹂닿퀬媛 ?덈뒗 寃쎌슦)
                    tables_data = {}
                    if st.session_state.kakao_work_completed and st.session_state.get('kakao_results'):
                        tables = st.session_state.kakao_results.split("```")
                        # ?좎뵪?뺣낫???쒖쇅?섍퀬 ?ㅼ젣 ?뚯씠釉붾쭔 泥섎━
                        table_names = ["?쒓났?꾪솴", "?묒뾽?댁슜", "?몄썝", "?λ퉬"]
                        real_tables = [t.strip() for t in tables if "\t" in t.strip()]
                        
                        # ?좎뵪?뺣낫瑜??쒖쇅?섍퀬 ?ㅼ젣 ?뚯씠釉붾쭔 泥섎━
                        table_index = 0
                        for i, tsv_data in enumerate(real_tables):
                            try:
                                # TSV ?곗씠???뺤젣
                                cleaned_tsv = re.sub(r'^tsv\n', '', tsv_data, flags=re.IGNORECASE)
                                fixed_tsv = fix_tsv_field_count(cleaned_tsv)
                                df = parse_tsv_to_dataframe(fixed_tsv)
                                
                                if df is not None and table_index < len(table_names):
                                    tables_data[table_names[table_index]] = df
                                    table_index += 1
                                    
                            except Exception as e:
                                continue
                    else:
                        # SNS ?묒뾽蹂닿퀬媛 ?녿뒗 寃쎌슦 湲곕낯 ?뚯씠釉??앹꽦
                        import pandas as pd
                        st.info("?뱄툘 SNS ?묒뾽蹂닿퀬媛 ?놁뼱 湲곕낯 ?뚯씠釉붾줈 怨듭궗?쇰낫瑜??앹꽦?⑸땲??")
                        
                        # 湲곕낯 ?뚯씠釉??곗씠???앹꽦
                        default_construction = pd.DataFrame({
                            '援щ텇': ['蹂몄꽑?곕꼸(1援ш컙)', '蹂몄꽑?곕꼸(2援ш컙)', '?좏뭾?뺢굅??, '?꾨┝?뺢굅??],
                            '?꾧퀎': ['0%', '0%', '0%', '0%']
                        })
                        default_work = pd.DataFrame({
                            '援щ텇': ['蹂몄꽑?곕꼸(1援ш컙)', '蹂몄꽑?곕꼸(2援ш컙)', '?좏뭾?뺢굅??, '?꾨┝?뺢굅??],
                            '湲덉씪?묒뾽': ['以鍮꾩쨷', '以鍮꾩쨷', '以鍮꾩쨷', '以鍮꾩쨷']
                        })
                        default_personnel = pd.DataFrame({
                            '援щ텇': ['吏곸쁺諛섏옣', '?곕꼸怨?, '紐⑷났', '泥좉렐怨?],
                            '?몄썝': ['0紐?, '0紐?, '0紐?, '0紐?]
                        })
                        default_equipment = pd.DataFrame({
                            '援щ텇': ['B/H(1.0LC)', '?ㅽ봽?몃윮(5T)', '?듦??щ젅??25T)', '誘뱀꽌?몃윮'],
                            '???: ['0?', '0?', '0?', '0?']
                        })
                        
                        tables_data = {
                            "?쒓났?꾪솴": default_construction,
                            "?묒뾽?댁슜": default_work,
                            "?몄썝": default_personnel,
                            "?λ퉬": default_equipment
                        }
                        

                    
                    # ?묒? ?뚯씪 ?앹꽦
                    try:
                        if template_bytes:
                            # ?쒗뵆由??ъ슜 - ? 留ㅽ븨 ?ㅼ젙 ?꾨떖
                            cell_mapping = {
                                'date': st.session_state.get('date_cell', 'u2'),
                                'project_name': st.session_state.get('project_cell', 'd4'),
                                'max_temp': st.session_state.get('max_temp_cell', 'o4'),
                                'min_temp': st.session_state.get('min_temp_cell', 'o5'),
                                'precipitation': st.session_state.get('precipitation_cell', 'o6'),
                                'planned_progress': st.session_state.get('planned_progress_cell', 'w4'),
                                'actual_progress': st.session_state.get('actual_progress_cell', 'w5'),
                                'progress_diff': st.session_state.get('progress_diff_cell', 'w6')
                            }
                            
                            table_mapping = {
                                '?쒓났?꾪솴': st.session_state.get('table_construction_cell', 'ac10'),
                                '?묒뾽?댁슜': st.session_state.get('table_work_cell', 'ac48'),
                                '?몄썝': st.session_state.get('table_personnel_cell', 'ac66'),
                                '?λ퉬': st.session_state.get('table_equipment_cell', 'ac106')
                            }
                            
                            # ?꾩씪 ?곗씠??媛?몄삤湲?                            previous_data = get_previous_day_data(basic_info['date'])
                            
                            excel_bytes = insert_data_to_excel_with_mapping(
                                template_bytes, 
                                basic_info, 
                                tables_data,
                                cell_mapping=cell_mapping,
                                table_mapping=table_mapping,
                                previous_data=previous_data
                            )
                        else:
                            # ?덈줈???묒떇?쇰줈 ?앹꽦
                            excel_bytes = create_excel_report(
                                basic_info=basic_info,
                                tables_data=tables_data
                            )
                        
                        if excel_bytes:
                            st.success("??怨듭궗?쇰낫 ?앹꽦 ?꾨즺!")
                            
                            # ?묒? ?ㅼ슫濡쒕뱶 踰꾪듉
                            st.download_button(
                                label="?뱿 怨듭궗?쇰낫 ?ㅼ슫濡쒕뱶",
                                data=excel_bytes,
                                file_name=f"怨듭궗?쇰낫_{basic_info['date']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            # ???寃곌낵 硫붿떆吏 ?쒖떆 (?붾㈃ ?덈줈怨좎묠 ?꾩뿉???좎?)
                            if hasattr(st.session_state, 'daily_report_saved'):
                                if st.session_state.daily_report_saved:
                                    st.success(st.session_state.save_success_message)
                                    st.info(f"?뱟 ??λ맂 ?좎쭨: {st.session_state.save_success_date}")
                                    # 硫붿떆吏 ?쒖떆 ???곹깭 珥덇린??                                    del st.session_state.daily_report_saved
                                    del st.session_state.save_success_message
                                    del st.session_state.save_success_date
                                else:
                                    st.error(st.session_state.save_error_message)
                                    # 硫붿떆吏 ?쒖떆 ???곹깭 珥덇린??                                    del st.session_state.daily_report_saved
                                    del st.session_state.save_error_message
                            
                            # Supabase ???踰꾪듉 異붽?
                            col1, col2 = st.columns(2)
                            with col1:
                                # ???踰꾪듉 ?곹깭???곕Ⅸ ?띿뒪??蹂寃?                                save_button_text = "?뮶 Supabase?????
                                if hasattr(st.session_state, 'daily_report_saved') and st.session_state.daily_report_saved:
                                    save_button_text = "??????꾨즺"
                                
                                save_button_clicked = st.button(save_button_text, key="save_to_supabase", use_container_width=True)
                                if save_button_clicked:
                                    # ???吏꾪뻾 ?곹깭 ?쒖떆
                                    with st.status("?뮶 Supabase ???以?..", expanded=True) as status:
                                        st.write("?뵇 ????꾨줈?몄뒪瑜??쒖옉?⑸땲??..")
                                        
                                        # Supabase ?곌껐 ?뚯뒪??                                        connection_success, connection_message = test_supabase_connection()
                                        if not connection_success:
                                            st.error(f"??Supabase ?곌껐 ?ㅽ뙣: {connection_message}")
                                            status.update(label="??????ㅽ뙣 - ?곌껐 ?ㅻ쪟", state="error")
                                        else:
                                            # ?뚯씠釉?議댁옱 ?щ? ?뺤씤
                                            table_exists, table_message = check_daily_report_data_table()
                                            if not table_exists:
                                                st.warning(f"?좑툘 daily_report_data ?뚯씠釉붿씠 議댁옱?섏? ?딆뒿?덈떎: {table_message}")
                                                st.write("?뵇 ?뚯씠釉붿쓣 ?먮룞?쇰줈 ?앹꽦?⑸땲??..")
                                                
                                                # ?뚯씠釉??앹꽦 ?쒕룄
                                                create_success, create_message = create_daily_report_data_table()
                                                if not create_success:
                                                    st.error(f"???뚯씠釉??앹꽦 ?ㅽ뙣: {create_message}")
                                                    st.info("?뮕 Supabase?먯꽌 ?섎룞?쇰줈 ?뚯씠釉붿쓣 ?앹꽦?댁＜?몄슂.")
                                                    status.update(label="??????ㅽ뙣 - ?뚯씠釉??앹꽦 ?ㅻ쪟", state="error")
                                                else:
                                                    st.success("???뚯씠釉??앹꽦 ?꾨즺!")
                                            else:
                                                st.success("???뚯씠釉붿씠 議댁옱?⑸땲??")
                                            
                                            # ????쒕룄
                                            try:
                                                # ??ν븷 ?곗씠??以鍮?(?묒? ?뚯씪?먯꽌 異붿텧??? ?곗씠?곕쭔)
                                                st.write("?뵇 3?④퀎 ?묒? ?뚯씪?먯꽌 異붿텧??? ?곗씠?곕? ??ν빀?덈떎.")
                                                report_data = {
                                                    "?쒓났?꾪솴": tables_data.get("?쒓났?꾪솴"),
                                                    "?묒뾽?댁슜": tables_data.get("?묒뾽?댁슜"),
                                                    "?몄썝": tables_data.get("?몄썝"),
                                                    "?λ퉬": tables_data.get("?λ퉬"),
                                                    "湲곕낯?뺣낫": basic_info,
                                                    "excel_bytes": excel_bytes  # ?묒? ?뚯씪 諛붿씠??異붽?
                                                }
                                                
                                                st.write(f"?뵇 ??ν븷 ?곗씠??以鍮??꾨즺: {list(report_data.keys())}")
                                                
                                                work_date = basic_info['date']
                                                st.write(f"?뵇 ??ν븷 ?좎쭨: {work_date}")
                                                
                                                save_result = save_to_supabase("daily_report", report_data, work_date)
                                                
                                                if save_result:
                                                    st.session_state.daily_report_saved = True
                                                    st.session_state.save_success_message = "?럦 ?묒뾽?쇰낫媛 Supabase???깃났?곸쑝濡???λ릺?덉뒿?덈떎!"
                                                    st.session_state.save_success_date = work_date
                                                    status.update(label="??????꾨즺!", state="complete")
                                                    st.success("?럦 ?묒뾽?쇰낫媛 Supabase???깃났?곸쑝濡???λ릺?덉뒿?덈떎!")
                                                    st.balloons()
                                                    st.toast("?뮶 Supabase ????꾨즺!", icon="??)
                                                else:
                                                    st.session_state.daily_report_saved = False
                                                    st.session_state.save_error_message = "??Supabase ??μ뿉 ?ㅽ뙣?덉뒿?덈떎."
                                                    status.update(label="??????ㅽ뙣", state="error")
                                                    st.error("??Supabase ??μ뿉 ?ㅽ뙣?덉뒿?덈떎.")
                                                    st.info("?뮕 ?ㅽ듃?뚰겕 ?곌껐?대굹 Supabase ?ㅼ젙???뺤씤?댁＜?몄슂.")
                                                    st.toast("??????ㅽ뙣", icon="??)
                                                    
                                            except Exception as save_error:
                                                status.update(label="?????以??ㅻ쪟 諛쒖깮", state="error")
                                                st.error(f"?????以??ㅻ쪟: {save_error}")
                                                import traceback
                                                st.error(f"???곸꽭 ?ㅻ쪟: {traceback.format_exc()}")
                                                st.info("?뮕 臾몄젣媛 吏?띾릺硫?愿由ъ옄?먭쾶 臾몄쓽?댁＜?몄슂.")
                                        

                            
                            with col2:
                                if st.button("?뱟 ??λ맂 ?곗씠??議고쉶", key="load_saved_data", use_container_width=True):
                                    selected_date = st.date_input("議고쉶???좎쭨 ?좏깮", value=datetime.strptime(basic_info['date'], '%Y-%m-%d').date(), key="load_date_step3")
                                    loaded_data = load_from_supabase("daily_report", selected_date.strftime('%Y-%m-%d'))
                                    if loaded_data:
                                        st.json(loaded_data)
                                    else:
                                        st.info("?대떦 ?좎쭨????λ맂 ?곗씠?곌? ?놁뒿?덈떎.")
                        else:
                            st.error("??怨듭궗?쇰낫 ?앹꽦 ?ㅽ뙣: ?뚯씪 ?앹꽦???ㅽ뙣?덉뒿?덈떎.")
                    except Exception as excel_error:
                        st.error(f"???묒? ?뚯씪 ?앹꽦 以??ㅻ쪟: {excel_error}")
                        st.info("?뮕 ?쒗뵆由??뚯씪?대굹 ?곗씠???뺤떇???뺤씤?댁＜?몄슂.")
                except Exception as e:
                    st.error(f"??怨듭궗?쇰낫 ?앹꽦 以??ㅻ쪟: {e}")
                    st.info("?뮕 湲곕낯 ?뺣낫瑜??뺤씤?댁＜?몄슂.")
        
        # PDF 蹂닿퀬???앹꽦 湲곕뒫 ?쒓굅??




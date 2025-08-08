import streamlit as st
import pandas as pd
from datetime import datetime
import os
import google.generativeai as genai
import io
import re
import pdfplumber
from pdf2image import convert_from_bytes
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import requests
import json
try:
    from streamlit_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode
    from streamlit_aggrid.shared import JsCode
    AGGRID_AVAILABLE = True
except ImportError:
    AGGRID_AVAILABLE = False
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

# --- CONFIG & SETUP ---
st.set_page_config(
    page_title="공사일보 자동화",
    page_icon="🏗️",
    layout="wide"
)

# Supabase 클라이언트는 함수 정의 후에 초기화됩니다

# --- STYLING ---
st.markdown("""
<style>
    /* Main App Font */
    html, body, [class*="st-"], .stTextArea, .stButton>button, .stFileUploader, .stSelectbox {
        font-family: 'Pretendard', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    }
    /* Main container */
    .main .block-container {
        padding: 2rem 2rem 5rem 2rem;
        max-width: 1000px;
    }
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #F8F9FA;
        border-right: 1px solid #E5E7EB;
    }
    [data-testid="stSidebar"] h1 {
        font-size: 1.5rem;
        color: #1E3A8A;
        font-weight: 700;
        padding: 1rem 0;
    }
    /* Step container in sidebar */
    .step-container {
        padding-top: 1rem;
    }
    .step {
        display: flex;
        align-items: center;
        margin-bottom: 1.25rem;
        padding: 0.75rem;
        border-radius: 0.5rem;
        transition: background-color 0.3s, border-color 0.3s;
        border-left: 5px solid #E5E7EB;
    }
    .step.active {
        border-left-color: #2563EB;
        background-color: #EFF6FF;
    }
    .step.completed {
        border-left-color: #16A34A;
    }
    .step-icon {
        font-size: 1.5rem;
        margin-right: 1rem;
    }
    .step-text {
        font-size: 1rem;
        font-weight: 500;
        color: #374151;
    }
    .step.completed .step-text {
        color: #115E59;
    }
    /* Main content cards */
    .card {
        background-color: white;
        border-radius: 0.75rem;
        padding: 2rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -2px rgba(0, 0, 0, 0.05);
        border: 1px solid #E5E7EB;
        margin-bottom: 2rem;
    }
    .card-title {
        font-size: 1.75rem;
        font-weight: 700;
        color: #1E3A8A;
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
    }
    .card-title .icon {
        font-size: 2rem;
        margin-right: 0.75rem;
    }
    .card-description {
        color: #4B5563;
        margin-bottom: 1.5rem;
    }
    /* Custom button style */
    .stButton>button {
        background-color: #2563EB;
        color: white;
        border: none;
        padding: 0.75rem 1.5rem;
        border-radius: 0.5rem;
        font-weight: 600;
        width: 100%;
        transition: background-color 0.3s;
    }
    .stButton>button:hover {
        background-color: #1D4ED8;
    }
    .stButton>button:disabled {
        background-color: #9CA3AF;
        color: #E5E7EB;
    }
    .stButton>button.reset-button {
        background-color: #D1D5DB;
        color: #4B5563;
    }
    .stButton>button.reset-button:hover {
        background-color: #9CA3AF;
        color: #1F2937;
    }
    
    /* 실시간 양식 스타일 */
    .live-form-container {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        border-radius: 15px;
        padding: 20px;
        margin: 10px 0;
        border: 2px solid #e0e6ed;
        box-shadow: 0 8px 25px rgba(0,0,0,0.1);
    }
    
    .live-form-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 15px;
        border-radius: 10px;
        margin-bottom: 20px;
        text-align: center;
        font-size: 1.2rem;
        font-weight: bold;
    }
    
    .live-form-section {
        background: white;
        border-radius: 10px;
        padding: 15px;
        margin: 10px 0;
        border: 1px solid #e1e5e9;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05);
    }
    
    .metric-container {
        background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%);
        border-radius: 8px;
        padding: 10px;
        margin: 5px 0;
        text-align: center;
    }
    
    /* 데이터 에디터 스타일 개선 */
    .stDataFrame {
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid #e1e5e9;
    }
    
    /* 탭 스타일 개선 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        background-color: #f8f9fa;
        border: 1px solid #dee2e6;
        padding: 10px 16px;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #007bff;
        color: white;
        border-color: #007bff;
    }
</style>
""", unsafe_allow_html=True)


# --- GLOBAL CONSTANTS & API SETUP ---
# Streamlit secrets에서 설정 가져오기
TEAMS_WEBHOOK_URL = st.secrets.get("TEAMS_WEBHOOK_URL", "https://poscoenc365.webhook.office.com/webhookb2/f6efcf11-c6a7-4385-903f-f3fd8937de55@ec1d3aa9-13ec-4dc5-8672-06fc64ca7701/IncomingWebhook/1fb9d9ce7f4c4093ba4fe9a8db67dc2f/1a2e3f7d-551b-40ec-90a1-e815373c81a7/V2qbqRtbAap4il8cvVljyk_ApZuHTDE0AfOYLQ8V9SqQs1")
GENAI_API_KEY = st.secrets.get("GENAI_API_KEY", "AIzaSyD69-wKYfZSID327fczrkx-JveJdGYIUIk")

# Supabase 클라이언트 초기화
@st.cache_resource
def init_supabase():
    """Supabase 클라이언트를 초기화하고 반환합니다."""
    if not SUPABASE_AVAILABLE:
        return None
        
    try:
        supabase_url = st.secrets.get("SUPABASE_URL")
        supabase_key = st.secrets.get("SUPABASE_KEY")
        
        if not supabase_url or not supabase_key:
            st.warning("⚠️ Supabase 설정이 완료되지 않았습니다. .streamlit/secrets.toml 파일을 확인해주세요.")
            return None
            
        if supabase_url == "https://your-project-id.supabase.co" or supabase_key == "your-anon-key-here":
            st.warning("⚠️ Supabase 설정이 기본값으로 되어 있습니다. 실제 프로젝트 정보로 업데이트해주세요.")
            return None
        
        # SSL 인증서 검증 우회 설정
        import ssl
        import urllib3
        import os
        
        # SSL 경고 비활성화
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # 환경 변수를 통한 SSL 검증 비활성화
        os.environ['PYTHONHTTPSVERIFY'] = '0'
        
        # SSL 컨텍스트 생성 (인증서 검증 우회)
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        # Supabase 클라이언트 생성 시 SSL 설정 적용
        client = create_client(supabase_url, supabase_key)
        
        # HTTP 클라이언트에 SSL 설정 적용
        if hasattr(client, 'rest') and hasattr(client.rest, 'transport'):
            import httpx
            transport = httpx.HTTPTransport(verify=False)
            client.rest.transport = transport
            
        return client
    except Exception as e:
        st.error(f"❌ Supabase 연결 실패: {e}")
        return None

# 전역 Supabase 클라이언트 변수
supabase_client = None

# Supabase 클라이언트 초기화
if SUPABASE_AVAILABLE:
    try:
        supabase_client = init_supabase()
        if supabase_client:
            st.success("✅ Supabase 연결 성공!")
        else:
            st.warning("⚠️ Supabase 연결 실패 - 기본 기능만 사용 가능합니다.")
    except Exception as e:
        st.warning(f"⚠️ Supabase 초기화 실패: {e}")
        supabase_client = None
else:
    st.warning("⚠️ Supabase 모듈이 설치되지 않아 데이터베이스 기능을 사용할 수 없습니다.")

# Gemini AI 설정
genai.configure(api_key=GENAI_API_KEY)
GEMINI_MODEL = genai.GenerativeModel("models/gemini-2.5-flash-preview-05-20")

BLAST_EXTRACTION_PROMPT = '''
# INSTRUCTION
- 반드시 아래 예시처럼 오직 TSV(탭 구분) 데이터만 출력하세요.
- 설명, 마크다운, 코드블록, 주석, 기타 텍스트는 절대 포함하지 마세요.
- 아래 예시와 동일한 형식으로만 출력하세요.
발파일자    발파시간    지발당장약량(최소, kg)    지발당장약량(최대, kg)    폭약사용량(kg)    발파진동(cm/sec)    발파소음(dB(A))    계측위치    비고
2023-07-27    08:05    0.5    0.9    73    -    -    -    PLA-2
2023-07-27    13:47    0.4    0.8    77    0.87    53.29    티스테이션    PD-2
2023-07-27    13:47    -    -    -    0.71    61.23    양말집    PD-2
(위 예시는 형식만 참고, 실제 데이터는 입력값에 따라 동적으로 생성)
# 입력
- 입력1: 발파작업일지_TSV (아래와 같은 형식)
- 입력2: 계측일지_TSV (아래와 같은 형식, **계측일지 표는 PDF 2페이지 이후부터 추출**)
# 입력1 예시
발파일자    발파시간    지발당장약량(최소, kg)    지발당장약량(최대, kg)    폭약사용량(kg)    비고
2023-07-27    08:05    0.5    0.9    73    PLA-2
2023-07-27    13:47    0.4    0.8    77    PD-2
# 입력2 예시 (**2페이지 이후 표만**)
Date/Time    Peak Particle Vel (X_Axis) (mm/sec)    Peak Particle Vel (Y_Axis) (mm/sec)    Peak Particle Vel (Z_Axis) (mm/sec)    LMax (Sound) (dBA)    측정위치
2023/07/27 1:47:00 PM    0.71    0.36    0.71    61.23    양말집
2023/07/27 1:47:00 PM    0.87    0.56    0.87    53.29    티스테이션
# Mapping Rules
- 두 입력을 병합하여 위 예시와 동일한 TSV만 출력
- 설명, 마크다운, 코드블록, 주석, 기타 텍스트는 절대 포함하지 마세요.
- 계측일지 표는 반드시 PDF 2페이지 이후의 표만 사용 
- 최종 헤더(고정열): 발파일자, 발파시간, 지발당장약량(최소, kg), 지발당장약량(최대, kg), 폭약사용량(kg), 발파진동(cm/sec), 발파소음(dB(A)), 계측위치, 비고
- 정렬: 발파시간 오름차순, 계측위치 오름차순(필요시)
- 병합/매칭/포맷 규칙은 기존과 동일
'''
DEFAULT_PROMPT = """
# INSTRUCTIONS
1. 일일작업보고 원문에서 데이터를 파싱하여 4개 테이블("시공현황", "작업내용", "인원", "장비")과 각각을 TSV 형식의 코드블록으로 생성합니다.
2. 자동 검증 결과(QA-CHECKLIST)를 마크다운 표(Table)로 생성합니다.
3. 일일작업보고 텍스트에서 **작업 날짜**를 추출하여 첫 번째로 출력 (YYYY-MM-DD 형식)

# OUTPUT  
## 1. 시공현황 테이블  
1. 고정 열 : "구분", "누계"  
2. 고정 행(총 33행) - 아래 순서와 명칭을 그대로  
- "1. 본선터널 (1구간, 대림-신풍) 굴착"  
- "1. 본선터널 (1구간, 대림-신풍) 라이닝" 
- "2. 신풍정거장 - 1)정거장 라이닝"
- "2. 신풍정거장 - 1)정거장 미들 슬라브"
- "2. 신풍정거장 – 2)주출입구 수직구 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 정거장 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 환승통로 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (2)PCC 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (3)PCD 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (4)PHA 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - 수직구 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - PHB 라이닝"
- "2. 신풍정거장 - 4)외부출입구 출입구(#3) 굴착" 
- "2. 신풍정거장 - 4)외부출입구 출입구(#2) 굴착"
- "2. 신풍정거장 - 4)외부출입구 출입구(#1) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 라이닝"  
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 라이닝"  
- "3. 신풍 환승통로 - 2)개착 BOX 보라매 방면 구조물"  
- "3. 신풍 환승통로 - 2)개착 BOX 대림 방면 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 터널 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 미들 슬라브" 
- "5. 도림사거리정거장 - 2)출입구#1 수직구 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCA 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCC 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PHA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 수직구 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCC 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PHB 라이닝"  

3. 추출데이터  
- "누계"값만 숫자로 추출할 것 (예:945.3m / 1,116m 에서 "945.3" 추출)

## 2. 작업내용 테이블  
1. 고정 열 : "구분", "금일작업"  
2. 고정 행(총 14행) - 아래 순서와 명칭(매핑 후 결과)을 그대로  
- "1. 본선터널 (1구간, 대림-신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍-도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"  
3. 주의사항  
- '작업내용' 셀은 여러 세부 내용을 포함할 수 있습니다. 내용을 구분할 때는, 최종 TSV 출력 시 해당 셀을 큰따옴표("...")로 감싸되, 셀 내부의 각 내용은 **실제 줄바꿈 문자(예: '\\n'문자열 대신 엔터 키 입력에 해당)**를 사용하여 분리하며, "-"기호는 생략함

## 3. 인원 / 장비 테이블  
1. 고정 열 (총 15열) - 열 순서는 아래와 같음
- "구분" 
- "1. 본선터널 (1구간, 대림~신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍~도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"    

2. 고정 행(인원 테이블 – 총 36행)  
(인원 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):

"직영반장", "연수생", "장비운전원", "전기주임", "화약주임", "터널공", "목공", "철근공", "라이닝폼공", "오폐수처리공", "카리프트공", "BP공", "가시설공", "설치공/해체공", "동바리공", "신호수", "부단수공", "슬러리월공", "CIP공", "미장공", "시설물공", "경계석공", "조경공", "배관공", "도색공", "방수공", "장비/작업지킴이", "보통인부", "포장공", "용접공", "타설공", "보링공/앙카공", "비계공", "도장공", "석면공", "주입공/그라우팅공"

3. 고정 행 (장비 테이블 – 총 46행)  
(장비 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):

"B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)", "덤프트럭(5T)", "덤프트럭(15T)", "덤프트럭(25T)", "앵글크레인(100T)", "앵글크레인(80T)", "앵글크레인(35T)", "앵글크레인(25T)", "카고크레인(25T)", "카고크레인(5T)", "콤프", "점보드릴", "페이로더", "숏트머신", "차징카", "살수차", "하이드로크레인", "믹서트럭", "화물차(5T)", "펌프카", "스카이", "콘크리트피니셔", "전주오거", "로더(바브켓)", "유제살포기(비우다)", "지게차", "싸인카", "BC커터기", "바이브로해머", "롤러(2.5T)", "롤러(1T)", "롤러(0.7T)", "몰리", "항타기", "크레인", "콤비로라", "공압드릴", "유압드릴", "기타"

## 4. Parsing Rules 
1. 시공현황: "누계/설계" → **앞값(소수 허용)** 만 추출.    
2. 인원·장비: 투입현황에서 **정수만** 추출, 빈셀은 **0**    
3. 하위 섹션 매핑   
   - 정거장 터널 → 열 ②, PCB → ③, PCC → ④, PCD → ⑤,PHA → ⑥, 특별피난 → ⑦, 외부출입구 →⑧    
4. 매핑 딕셔너리 적용    
- "B/H08W" →"B/H(08W)"   
- "25톤 카고크레인" → "카고크레인(25T)"   
- "특공" → "보통인부"    
- "기계타설공" → "타설공"    
- "목공연수생" 또는 "목수연수생" → "연수생"    
- "5톤트럭" → "화물차(5T)"    
- "카리프트" → "카리프트공"    
- "하이드로크레인(20T)" → "하이드로크레인"    
- "라이닝폼조립" → "라이닝폼공"  
- "S/C타설팀" → "터널공"  
- "목수" → "목공"    
5. 사전에 없는 항목 → 유사항목, 없으면 **인원: 보통인부 / 장비: 기타** 로 합산하고 '오류요약'에 기재.

## 5. QA-CHECKLIST(자동 검증 결과)

 1. 검증 항목
아래 기준에 따라 데이터 처리 과정의 정확성을 자체 검증하고, 그 결과를 마크다운 테이블로 생성합니다.
- **구조/형식**: 4개 테이블의 행과 열 개수, 순서, 데이터 형식(숫자, 정수, 0 처리)이 지침과 일치하는가?
- **데이터 무결성**: 원본 보고서의 인원 및 장비 수량이 누락되거나 중복되지 않고 100% 정확하게 집계되었는가?
- **매핑/변환**: 지정된 매핑 규칙(용어 표준화, 유사 항목 처리 등)이 모두 올바르게 적용되었는가?
- **미분류 항목**: 사전에 정의되지 않은 항목이 '보통인부' 또는 '기타'로 적절히 분류되고 기록되었는가?

2. 출력 방식
- **요약**: 검증 결과를 아래 예시와 같이 마크다운 테이블(`QA-CHECKLIST`)으로 요약합니다.
- **변환 내역**: 데이터 처리 과정에서 변경된 내용이 있는 경우, **'변환 내역'**란에 `원문 → 결과` 형식으로 명시합니다. 변경 사항이 없으면 "이상 없음"으로 표기합니다.

3. 예시 (마크다운 테이블)
| 점검 항목 | 검증 기준 | 변환 내역 (원문 → 결과) | 상태 |
| :--- | :--- | :--- | :---: |
| **구조 및 형식** | 4개 테이블의 구조 및 데이터 형식이 지침과 일치함 | 이상 없음 | ✅ |
| **데이터 무결성** | 인원(85명), 장비(12대) 총계가 원문과 일치함 | 이상 없음 | ✅ |
| **용어 표준화** | 매핑 규칙에 따라 용어가 일괄 변환됨 | - 목수 → 목공<br>- 특공 → 보통인부<br>- B/H08W → B/H(08W) | ✅ |
| **미분류 항목 처리** | 사전에 없는 항목을 규칙에 따라 처리함 | - 노무원 → 보통인부 (합산) | ⚠️ |

---

# USER TEXT(작업계획보고 입력란)
"""

# --- PROMPT MANAGEMENT FUNCTIONS ---
def save_prompt_to_supabase(prompt_name, prompt_content, description=""):
    """프롬프트를 Supabase에 저장합니다."""
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return False
    
    try:
        data = {
            "name": prompt_name,
            "content": prompt_content,
            "description": description,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        # 기존 프롬프트가 있는지 확인
        existing = supabase_client.table("prompts").select("*").eq("name", prompt_name).execute()
        
        if existing.data:
            # 업데이트
            data["updated_at"] = datetime.now().isoformat()
            result = supabase_client.table("prompts").update(data).eq("name", prompt_name).execute()
            st.success(f"✅ 프롬프트 '{prompt_name}'이(가) 업데이트되었습니다.")
        else:
            # 새로 생성
            result = supabase_client.table("prompts").insert(data).execute()
            st.success(f"✅ 프롬프트 '{prompt_name}'이(가) 저장되었습니다.")
        
        return True
    except Exception as e:
        st.error(f"❌ 프롬프트 저장 실패: {e}")
        return False

def load_prompt_from_supabase(prompt_name):
    """Supabase에서 프롬프트를 로드합니다."""
    if not supabase_client:
        return None
    
    try:
        result = supabase_client.table("prompts").select("*").eq("name", prompt_name).execute()
        if result.data:
            return result.data[0]
        return None
    except Exception as e:
        st.error(f"❌ 프롬프트 로드 실패: {e}")
        return None

def get_all_prompts_from_supabase():
    """Supabase에서 모든 프롬프트 목록을 가져옵니다."""
    if not supabase_client:
        return []
    
    try:
        result = supabase_client.table("prompts").select("name, description, updated_at").execute()
        return result.data if result.data else []
    except Exception as e:
        st.error(f"❌ 프롬프트 목록 로드 실패: {e}")
        return []

def delete_prompt_from_supabase(prompt_name):
    """Supabase에서 프롬프트를 삭제합니다."""
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 설정되지 않았습니다.")
        return False
    
    try:
        result = supabase_client.table("prompts").delete().eq("name", prompt_name).execute()
        st.success(f"✅ 프롬프트 '{prompt_name}'이(가) 삭제되었습니다.")
        return True
    except Exception as e:
        st.error(f"❌ 프롬프트 삭제 실패: {e}")
        return False

def generate_prompt_from_tables():
    """테이블 구조 데이터를 기반으로 프롬프트를 생성합니다."""
    
    # 시공현황 행 목록
    construction_rows = st.session_state.construction_rows
    
    # 작업내용 행 목록
    work_content_rows = st.session_state.work_content_rows
    
    # 인원/장비 열 목록
    personnel_columns = st.session_state.personnel_columns
    personnel_rows = st.session_state.personnel_rows
    equipment_rows = st.session_state.equipment_rows
    
    prompt = f"""# INSTRUCTIONS
1. 일일작업보고 원문에서 데이터를 파싱하여 4개 테이블("시공현황", "작업내용", "인원", "장비")과 각각을 TSV 형식의 코드블록으로 생성합니다.
2. 자동 검증 결과(QA-CHECKLIST)를 마크다운 표(Table)로 생성합니다.
3. 일일작업보고 텍스트에서 **작업 날짜**를 추출하여 첫 번째로 출력 (YYYY-MM-DD 형식)

# OUTPUT  
## 1. 시공현황 테이블  
1. 고정 열 : "구분", "누계"  
2. 고정 행(총 {len(construction_rows)}행) - 아래 순서와 명칭을 그대로  
{chr(10).join([f'- "{row}"' for row in construction_rows])}
3. 추출데이터  
- "누계"값만 숫자로 추출할 것 (예:945.3m / 1,116m 에서 "945.3" 추출)

## 2. 작업내용 테이블  
1. 고정 열 : "구분", "금일작업"  
2. 고정 행(총 {len(work_content_rows)}행) - 아래 순서와 명칭(매핑 후 결과)을 그대로  
{chr(10).join([f'- "{row}"' for row in work_content_rows])}
3. 주의사항  
- '작업내용' 셀은 여러 세부 내용을 포함할 수 있습니다. 내용을 구분할 때는, 최종 TSV 출력 시 해당 셀을 큰따옴표("...")로 감싸되, 셀 내부의 각 내용은 **실제 줄바꿈 문자(예: '\\n'문자열 대신 엔터 키 입력에 해당)**를 사용하여 분리하며, "-"기호는 생략함

## 3. 인원 / 장비 테이블  
1. 고정 열 (총 {len(personnel_columns) + 1}열) - 열 순서는 아래와 같음
- "구분" 
{chr(10).join([f'- "{col}"' for col in personnel_columns])}
2. 고정 행(인원 테이블 – 총 {len(personnel_rows)}행)  
(인원 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):
{', '.join([f'"{row}"' for row in personnel_rows])}
3. 고정 행 (장비 테이블 – 총 {len(equipment_rows)}행)  
(장비 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):
{', '.join([f'"{row}"' for row in equipment_rows])}

## 4. Parsing Rules 
1. 시공현황: "누계/설계" → **앞값(소수 허용)** 만 추출.    
2. 인원·장비: 투입현황에서 **정수만** 추출, 빈셀은 **0**    
3. 하위 섹션 매핑   
   - 정거장 터널 → 열 ②, PCB → ③, PCC → ④, PCD → ⑤,PHA → ⑥, 특별피난 → ⑦, 외부출입구 →⑧    
4. 매핑 딕셔너리 적용    
- "B/H08W" →"B/H(08W)"   
- "25톤 카고크레인" → "카고크레인(25T)"   
- "특공" → "보통인부"    
- "기계타설공" → "타설공"    
- "목공연수생" 또는 "목수연수생" → "연수생"    
- "5톤트럭" → "화물차(5T)"    
- "카리프트" → "카리프트공"    
- "하이드로크레인(20T)" → "하이드로크레인"    
- "라이닝폼조립" → "라이닝폼공"  
- "S/C타설팀" → "터널공"  
- "목수" → "목공"    
5. 사전에 없는 항목 → 유사항목, 없으면 **인원: 보통인부 / 장비: 기타** 로 합산하고 '오류요약'에 기재.

## 5. QA-CHECKLIST(자동 검증 결과)

 1. 검증 항목
아래 기준에 따라 데이터 처리 과정의 정확성을 자체 검증하고, 그 결과를 마크다운 테이블로 생성합니다.
- **구조/형식**: 4개 테이블의 행과 열 개수, 순서, 데이터 형식(숫자, 정수, 0 처리)이 지침과 일치하는가?
- **데이터 무결성**: 원본 보고서의 인원 및 장비 수량이 누락되거나 중복되지 않고 100% 정확하게 집계되었는가?
- **매핑/변환**: 지정된 매핑 규칙(용어 표준화, 유사 항목 처리 등)이 모두 올바르게 적용되었는가?
- **미분류 항목**: 사전에 정의되지 않은 항목이 '보통인부' 또는 '기타'로 적절히 분류되고 기록되었는가?

2. 출력 방식
- **요약**: 검증 결과를 아래 예시와 같이 마크다운 테이블(`QA-CHECKLIST`)으로 요약합니다.
- **변환 내역**: 데이터 처리 과정에서 변경된 내용이 있는 경우, **'변환 내역'**란에 `원문 → 결과` 형식으로 명시합니다. 변경 사항이 없으면 "이상 없음"으로 표기합니다.

3. 예시 (마크다운 테이블)
| 점검 항목 | 검증 기준 | 변환 내역 (원문 → 결과) | 상태 |
| :--- | :--- | :--- | :---: |
| **구조 및 형식** | 4개 테이블의 구조 및 데이터 형식이 지침과 일치함 | 이상 없음 | ✅ |
| **데이터 무결성** | 인원(85명), 장비(12대) 총계가 원문과 일치함 | 이상 없음 | ✅ |
| **용어 표준화** | 매핑 규칙에 따라 용어가 일괄 변환됨 | - 목수 → 목공<br>- 특공 → 보통인부<br>- B/H08W → B/H(08W) | ✅ |
| **미분류 항목 처리** | 사전에 없는 항목을 규칙에 따라 처리함 | - 노무원 → 보통인부 (합산) | ⚠️ |

---

# USER TEXT(작업계획보고 입력란)
"""
    
    return prompt

# --- HELPER FUNCTIONS ---
def display_unified_construction_report():
    """통합된 공사일보 고정양식을 표시합니다 - 5개 테이블이 모두 하나로 합쳐진 형태"""
    
    # 세션 상태에서 데이터 가져오기 (없으면 기본값 사용)
    work_date = st.session_state.get('work_date', None)
    if work_date is None:
        work_date = datetime.now().strftime('%Y-%m-%d')
    
    # 기본 정보 가져오기
    project_name = st.session_state.get('live_project_name', '터널 건설공사')
    section_name = st.session_state.get('live_section_name', '대림-신풍-도림')
    try:
        default_date = datetime.strptime(work_date, '%Y-%m-%d')
    except (ValueError, TypeError):
        default_date = datetime.now()
    report_date = st.session_state.get('live_report_date', default_date)
    
    # 제목
    st.markdown("""
    <div style="text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                color: white; padding: 20px; border-radius: 10px; margin: 20px 0;">
        <h1 style="margin: 0; font-size: 28px; font-weight: bold;">🏗️ 터널 건설공사 일일작업보고서</h1>
    </div>
    """, unsafe_allow_html=True)
    
    # 통합 양식 데이터 초기화
    if 'unified_form_data' not in st.session_state:
        initialize_unified_form_data()
    
    # 전체 통합 양식을 하나의 큰 AgGrid로 표시
    if AGGRID_AVAILABLE:
        display_unified_aggrid()
    else:
        st.error("❌ streamlit-aggrid가 설치되지 않았습니다. 통합 양식을 표시할 수 없습니다.")
    
    # 하단 액션 버튼들
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("🔄 양식 초기화", key="reset_unified_form", use_container_width=True):
            initialize_unified_form_data()
            st.rerun()
    
    with col2:
        if st.button("📊 엑셀 다운로드", key="download_unified_excel", use_container_width=True):
            try:
                excel_data = create_excel_from_unified_data()
                st.download_button(
                    label="📥 엑셀 파일 다운로드",
                    data=excel_data,
                    file_name=f"공사일보_{report_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_unified_excel_button"
                )
                st.success("✅ 통합 양식 그대로 엑셀 파일이 생성됩니다!")
            except Exception as e:
                st.error(f"❌ 엑셀 생성 중 오류: {e}")
    
    with col3:
        if st.button("📤 템플릿 업로드", key="upload_template", use_container_width=True):
            st.session_state.show_template_upload = True
            st.rerun()
    
    with col4:
        if st.button("🤖 AI로 채우기", key="fill_with_ai", use_container_width=True):
            if st.session_state.get('processed_tables'):
                fill_unified_form_with_ai_data()
                st.rerun()
            else:
                st.warning("⚠️ 먼저 카카오톡 데이터를 분석해주세요.")
    
    # 템플릿 업로드 모달
    if st.session_state.get('show_template_upload', False):
        display_template_upload_modal()

def initialize_unified_form_data():
    """통합 양식 데이터를 초기화합니다"""
    
    # 통합 데이터프레임 생성 - 모든 테이블이 하나로 합쳐진 형태
    unified_data = []
    
    # 1. 기본 정보
    unified_data.extend([
        {"섹션": "기본정보", "구분": "프로젝트명", "내용": "터널 건설공사", "타입": "basic"},
        {"섹션": "기본정보", "구분": "구간명", "내용": "대림-신풍-도림", "타입": "basic"},
        {"섹션": "기본정보", "구분": "보고일", "내용": datetime.now().strftime('%Y-%m-%d'), "타입": "basic"},
        {"섹션": "", "구분": "", "내용": "", "타입": "separator"},  # 구분선
    ])
    
    # 2. 시공현황
    unified_data.append({"섹션": "시공현황", "구분": "섹션 헤더", "내용": "🏗️ 시공 현황", "타입": "header"})
    construction_items = st.session_state.get('construction_rows', [
        "1. 본선터널 (1구간, 대림-신풍)",
        "2. 신풍정거장 - 1)정거장 라이닝",
        "3. 신풍 환승통로 - 1)환승터널",
        "4. 본선터널(2구간, 신풍-도림) 굴착",
        "5. 도림사거리정거장 - 1)정거장 터널 라이닝"
    ])[:5]  # 상위 5개만
    for item in construction_items:
        unified_data.append({"섹션": "시공현황", "구분": item, "내용": "", "타입": "construction"})
    unified_data.append({"섹션": "", "구분": "", "내용": "", "타입": "separator"})
    
    # 3. 작업내용
    unified_data.append({"섹션": "작업내용", "구분": "섹션 헤더", "내용": "📝 금일 작업 내용", "타입": "header"})
    work_items = st.session_state.get('work_content_rows', [
        "1. 본선터널 (1구간, 대림-신풍)",
        "2.신풍정거장 - 1)정거장 터널",
        "3.신풍 환승통로 - 1)환승터널",
        "4.본선터널(2구간, 신풍-도림)",
        "5.도림사거리정거장 - 1)정거장 터널"
    ])[:5]  # 상위 5개만
    for item in work_items:
        unified_data.append({"섹션": "작업내용", "구분": item, "내용": "", "타입": "work"})
    unified_data.append({"섹션": "", "구분": "", "내용": "", "타입": "separator"})
    
    # 4. 인원현황 요약 (주요 직종만)
    unified_data.append({"섹션": "인원현황", "구분": "섹션 헤더", "내용": "👥 인원 현황", "타입": "header"})
    personnel_items = ["직영반장", "터널공", "목공", "철근공", "보통인부"]
    for item in personnel_items:
        unified_data.append({"섹션": "인원현황", "구분": item, "내용": "0", "타입": "personnel"})
    unified_data.append({"섹션": "", "구분": "", "내용": "", "타입": "separator"})
    
    # 5. 장비현황 요약 (주요 장비만)
    unified_data.append({"섹션": "장비현황", "구분": "섹션 헤더", "내용": "🚛 장비 현황", "타입": "header"})
    equipment_items = ["B/H(08W)", "덤프트럭(5T)", "앵글크레인(25T)", "믹서트럭", "기타"]
    for item in equipment_items:
        unified_data.append({"섹션": "장비현황", "구분": item, "내용": "0", "타입": "equipment"})
    
    # 데이터프레임으로 변환
    st.session_state.unified_form_data = pd.DataFrame(unified_data)

def display_unified_aggrid():
    """통합 AgGrid를 표시합니다"""
    df = st.session_state.unified_form_data.copy()
    
    # 테이블 편집 버튼들
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("➕ 행 추가", key="add_row_btn", use_container_width=True):
            add_row_to_unified_form()
            st.rerun()
    
    with col2:
        if st.button("➖ 행 삭제", key="delete_row_btn", use_container_width=True):
            if len(df) > 0:
                # 마지막 행 삭제 (separator가 아닌 행)
                for i in range(len(df)-1, -1, -1):
                    if df.iloc[i]['타입'] not in ['header', 'separator']:
                        st.session_state.unified_form_data = df.drop(df.index[i]).reset_index(drop=True)
                        st.rerun()
                        break
    
    with col3:
        section_type = st.selectbox(
            "추가할 섹션:", 
            ["시공현황", "작업내용", "인원현황", "장비현황"],
            key="section_select"
        )
    
    with col4:
        if st.button("📝 섹션 추가", key="add_section_btn", use_container_width=True):
            add_section_to_unified_form(section_type)
            st.rerun()
    
    # GridOptionsBuilder 설정
    gb = GridOptionsBuilder.from_dataframe(df)
    
    # 기본 설정
    gb.configure_default_column(
        editable=True, 
        groupable=False, 
        sortable=False, 
        filter=False,
        resizable=True
    )
    
    # 각 컬럼 설정
    gb.configure_column("섹션", headerName="섹션", width=120, editable=True)
    gb.configure_column("구분", headerName="구분", width=300, editable=True)
    gb.configure_column("내용", headerName="내용", width=250, editable=True)
    gb.configure_column("타입", hide=True)  # 타입 컬럼 숨기기
    
    # 그리드 옵션
    gb.configure_grid_options(
        enableRangeSelection=True,
        enableRowSelection=True,
        domLayout='normal',
        rowHeight=35
    )
    
    grid_options = gb.build()
    
    # 조건부 스타일링을 위한 JS 코드
    row_style_jscode = JsCode("""
    function(params) {
        if (params.data.타입 === 'header') {
            return {
                'background-color': '#4F81BD',
                'color': 'white',
                'font-weight': 'bold',
                'text-align': 'center'
            };
        }
        if (params.data.타입 === 'separator') {
            return {
                'background-color': '#f0f0f0',
                'height': '5px'
            };
        }
        if (params.data.타입 === 'basic') {
            return {
                'background-color': '#e8f4f8',
                'font-weight': 'bold'
            };
        }
        return {};
    }
    """)
    
    # AgGrid 표시
    st.markdown("### 📋 통합 공사일보 양식")
    st.markdown("*아래 양식에서 직접 내용을 편집하거나, 행/열을 추가하고, 카카오톡 데이터를 분석하여 자동으로 채울 수 있습니다.*")
    
    grid_response = AgGrid(
        df,
        gridOptions=grid_options,
        height=600,  # 고정 높이
        theme='streamlit',
        allow_unsafe_jscode=True,
        reload_data=False,
        update_mode=GridUpdateMode.VALUE_CHANGED,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        custom_css={
            ".ag-header-cell": {"background-color": "#2c3e50", "color": "white", "font-weight": "bold"},
            ".ag-cell": {"border": "1px solid #ddd"},
            ".ag-root-wrapper": {"border": "2px solid #3498db", "border-radius": "10px"}
        }
    )
    
    # 변경된 데이터 저장
    if grid_response['data'] is not None:
        st.session_state.unified_form_data = pd.DataFrame(grid_response['data'])

def add_row_to_unified_form():
    """통합 양식에 새 행을 추가합니다"""
    if 'unified_form_data' not in st.session_state:
        initialize_unified_form_data()
        return
    
    df = st.session_state.unified_form_data
    
    # 새 행 추가 (기본 타입으로)
    new_row = {
        "섹션": "사용자추가",
        "구분": "새 항목",
        "내용": "",
        "타입": "custom"
    }
    
    # 데이터프레임에 추가
    new_df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    st.session_state.unified_form_data = new_df

def add_section_to_unified_form(section_type):
    """통합 양식에 새 섹션을 추가합니다"""
    if 'unified_form_data' not in st.session_state:
        initialize_unified_form_data()
        return
    
    df = st.session_state.unified_form_data
    
    # 섹션 헤더 및 기본 항목들
    section_data = []
    
    # 구분선 추가
    section_data.append({"섹션": "", "구분": "", "내용": "", "타입": "separator"})
    
    # 섹션별 기본 항목들
    if section_type == "날씨정보":
        section_data.append({"섹션": "날씨정보", "구분": "섹션 헤더", "내용": "🌤️ 추가 날씨 현황", "타입": "header"})
        section_data.extend([
            {"섹션": "날씨정보", "구분": "습도", "내용": "", "타입": "weather"},
            {"섹션": "날씨정보", "구분": "풍속", "내용": "", "타입": "weather"}
        ])
    elif section_type == "시공현황":
        section_data.append({"섹션": "시공현황", "구분": "섹션 헤더", "내용": "🏗️ 추가 시공 현황", "타입": "header"})
        section_data.extend([
            {"섹션": "시공현황", "구분": "추가 공사 항목 1", "내용": "", "타입": "construction"},
            {"섹션": "시공현황", "구분": "추가 공사 항목 2", "내용": "", "타입": "construction"}
        ])
    elif section_type == "작업내용":
        section_data.append({"섹션": "작업내용", "구분": "섹션 헤더", "내용": "📝 추가 작업 내용", "타입": "header"})
        section_data.extend([
            {"섹션": "작업내용", "구분": "추가 작업 1", "내용": "", "타입": "work"},
            {"섹션": "작업내용", "구분": "추가 작업 2", "내용": "", "타입": "work"}
        ])
    elif section_type == "인원현황":
        section_data.append({"섹션": "인원현황", "구분": "섹션 헤더", "내용": "👥 추가 인원 현황", "타입": "header"})
        section_data.extend([
            {"섹션": "인원현황", "구분": "추가 직종 1", "내용": "0", "타입": "personnel"},
            {"섹션": "인원현황", "구분": "추가 직종 2", "내용": "0", "타입": "personnel"}
        ])
    elif section_type == "장비현황":
        section_data.append({"섹션": "장비현황", "구분": "섹션 헤더", "내용": "🚛 추가 장비 현황", "타입": "header"})
        section_data.extend([
            {"섹션": "장비현황", "구분": "추가 장비 1", "내용": "0", "타입": "equipment"},
            {"섹션": "장비현황", "구분": "추가 장비 2", "내용": "0", "타입": "equipment"}
        ])
    
    # 데이터프레임에 추가
    new_df = pd.concat([df, pd.DataFrame(section_data)], ignore_index=True)
    st.session_state.unified_form_data = new_df

def fill_unified_form_with_ai_data():
    """AI 분석 결과로 통합 양식의 빈칸을 자동으로 채웁니다"""
    try:
        if 'unified_form_data' not in st.session_state:
            initialize_unified_form_data()
        
        processed_tables = st.session_state.get('processed_tables', [])
        work_date = st.session_state.get('work_date')
        
        if not processed_tables:
            st.warning("⚠️ 처리된 테이블 데이터가 없습니다.")
            return
        
        df = st.session_state.unified_form_data.copy()
        
        # 작업 날짜 업데이트
        if work_date:
            df.loc[df['구분'] == '보고일', '내용'] = work_date
        
        # 스마트 매칭으로 AI 데이터를 템플릿에 채우기
        filled_count = 0
        
        if len(processed_tables) >= 1 and processed_tables[0] is not None:
            # 시공현황 - 스마트 매칭
            construction_df = processed_tables[0]
            for _, ai_row in construction_df.iterrows():
                ai_key = str(ai_row['구분']).strip().lower()
                ai_value = str(ai_row.get('누계', ai_row.get('금일', ''))).strip()
                
                if ai_value and ai_value != 'nan':
                    for idx, template_row in df.iterrows():
                        if template_row['타입'] in ['construction', 'custom']:
                            template_key = str(template_row['구분']).strip().lower()
                            
                            # 시공현황 키워드 매칭 (더 유연하게)
                            if (ai_key in template_key or template_key in ai_key or
                                similarity_match(ai_key, template_key)):
                                df.loc[idx, '내용'] = ai_value
                                filled_count += 1
                                break
        
        if len(processed_tables) >= 3 and processed_tables[2] is not None:
            # 작업내용 - 스마트 매칭
            work_df = processed_tables[2]
            for _, ai_row in work_df.iterrows():
                ai_key = str(ai_row['구분']).strip().lower()
                ai_value = str(ai_row.get('금일작업', ai_row.get('작업내용', ''))).strip()
                
                if ai_value and ai_value != 'nan':
                    for idx, template_row in df.iterrows():
                        if template_row['타입'] in ['work', 'custom']:
                            template_key = str(template_row['구분']).strip().lower()
                            
                            if (ai_key in template_key or template_key in ai_key or
                                similarity_match(ai_key, template_key)):
                                df.loc[idx, '내용'] = ai_value
                                filled_count += 1
                                break
        
        if len(processed_tables) >= 4 and processed_tables[3] is not None:
            # 인원현황 - 스마트 매칭 및 합계 계산
            personnel_df = processed_tables[3]
            for _, ai_row in personnel_df.iterrows():
                ai_key = str(ai_row['구분']).strip().lower()
                
                # 합계 계산 (구분 컬럼 제외)
                total = 0
                for col in personnel_df.columns[1:]:
                    val = pd.to_numeric(ai_row[col], errors='coerce')
                    if pd.notna(val):
                        total += val
                
                if total > 0:
                    for idx, template_row in df.iterrows():
                        if template_row['타입'] in ['personnel', 'custom']:
                            template_key = str(template_row['구분']).strip().lower()
                            
                            if (ai_key in template_key or template_key in ai_key or
                                similarity_match(ai_key, template_key)):
                                df.loc[idx, '내용'] = str(int(total))
                                filled_count += 1
                                break
        
        if len(processed_tables) >= 5 and processed_tables[4] is not None:
            # 장비현황 - 스마트 매칭 및 합계 계산
            equipment_df = processed_tables[4]
            for _, ai_row in equipment_df.iterrows():
                ai_key = str(ai_row['구분']).strip().lower()
                
                # 합계 계산 (구분 컬럼 제외)
                total = 0
                for col in equipment_df.columns[1:]:
                    val = pd.to_numeric(ai_row[col], errors='coerce')
                    if pd.notna(val):
                        total += val
                
                if total > 0:
                    for idx, template_row in df.iterrows():
                        if template_row['타입'] in ['equipment', 'custom']:
                            template_key = str(template_row['구분']).strip().lower()
                            
                            if (ai_key in template_key or template_key in ai_key or
                                similarity_match(ai_key, template_key)):
                                df.loc[idx, '내용'] = str(int(total))
                                filled_count += 1
                                break
        
        # 업데이트된 데이터 저장
        st.session_state.unified_form_data = df
        
        st.success(f"✅ AI 분석 결과로 양식이 자동으로 채워졌습니다! ({filled_count}개 항목 매칭)")
        if filled_count > 0:
            st.balloons()
        else:
            st.warning("⚠️ 매칭된 항목이 없습니다. 템플릿 구조를 확인해주세요.")
        
    except Exception as e:
        st.error(f"❌ 양식 자동 채우기 중 오류: {e}")
        import traceback
        st.code(traceback.format_exc())

def similarity_match(ai_key, template_key):
    """AI 키와 템플릿 키의 유사성을 검사합니다"""
    try:
        # 기본 키워드 매칭
        common_keywords = [
            # 날씨 관련
            ['온도', '기온', 'temp'], ['습도', 'humidity'], ['강수', '비', '우량', 'rain'],
            # 시공 관련  
            ['터널', 'tunnel'], ['굴착', '발파', 'excavation'], ['라이닝', 'lining'], 
            ['콘크리트', 'concrete'], ['정거장', 'station'],
            # 작업 관련
            ['작업', 'work'], ['시공', 'construction'], ['공사', 'project'],
            # 인원 관련
            ['반장', 'foreman'], ['기능공', 'worker'], ['인부', 'laborer'], ['기사', 'operator'],
            # 장비 관련
            ['덤프', 'dump'], ['크레인', 'crane'], ['굴삭기', 'excavator'], ['로더', 'loader']
        ]
        
        ai_key_lower = ai_key.lower()
        template_key_lower = template_key.lower()
        
        # 공통 키워드 그룹에서 매칭 확인
        for keyword_group in common_keywords:
            ai_has_keyword = any(keyword in ai_key_lower for keyword in keyword_group)
            template_has_keyword = any(keyword in template_key_lower for keyword in keyword_group)
            
            if ai_has_keyword and template_has_keyword:
                return True
        
        # 숫자 패턴 매칭 (예: "1.", "2)", "(1)" 등)
        import re
        ai_numbers = re.findall(r'\d+', ai_key)
        template_numbers = re.findall(r'\d+', template_key)
        
        if ai_numbers and template_numbers:
            if any(num in template_numbers for num in ai_numbers):
                return True
        
        return False
        
    except Exception:
        return False

def create_excel_from_unified_data():
    """통합 양식 데이터로 엑셀 파일을 생성합니다 - 화면과 100% 동일하게"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "공사일보"
    
    # 스타일 정의 (AgGrid와 동일)
    header_font = Font(bold=True, size=12, name='맑은 고딕', color='FFFFFF')
    normal_font = Font(size=10, name='맑은 고딕')
    title_font = Font(bold=True, size=16, name='맑은 고딕')
    
    # AgGrid와 동일한 색상 사용
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    basic_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid") 
    separator_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 통합 양식 데이터 가져오기
    df = st.session_state.get('unified_form_data', pd.DataFrame())
    if df.empty:
        # 빈 워크북 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    current_row = 1
    
    # 제목
    ws.merge_cells(f'A{current_row}:C{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value="🏗️ 터널 건설공사 일일작업보고서")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2
    
    # 테이블 헤더 생성 (AgGrid와 동일)
    header_row = current_row
    ws.cell(row=header_row, column=1, value="섹션").font = header_font
    ws.cell(row=header_row, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.cell(row=header_row, column=1).border = thin_border
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(row=header_row, column=2, value="구분").font = header_font
    ws.cell(row=header_row, column=2).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.cell(row=header_row, column=2).border = thin_border
    ws.cell(row=header_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(row=header_row, column=3, value="내용").font = header_font
    ws.cell(row=header_row, column=3).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.cell(row=header_row, column=3).border = thin_border
    ws.cell(row=header_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    
    # 데이터 순회하며 엑셀에 기록 (화면과 동일하게)
    for _, row in df.iterrows():
        if row['타입'] == 'separator':
            # 구분선 - 빈 행으로 표시
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col, value="")
                cell.fill = separator_fill
                cell.border = thin_border
        elif row['타입'] == 'header':
            # 섹션 헤더 - 3열 병합
            ws.merge_cells(f'A{current_row}:C{current_row}')
            cell = ws.cell(row=current_row, column=1, value=row['내용'])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif row['타입'] == 'basic':
            # 기본 정보 - 특별한 스타일
            ws.cell(row=current_row, column=1, value=row['섹션']).font = normal_font
            ws.cell(row=current_row, column=1).fill = basic_fill
            ws.cell(row=current_row, column=1).border = thin_border
            
            ws.cell(row=current_row, column=2, value=row['구분']).font = normal_font
            ws.cell(row=current_row, column=2).fill = basic_fill
            ws.cell(row=current_row, column=2).border = thin_border
            
            ws.cell(row=current_row, column=3, value=row['내용']).font = normal_font
            ws.cell(row=current_row, column=3).border = thin_border
        else:
            # 일반 데이터 - 3컬럼 모두 표시 (화면과 동일)
            ws.cell(row=current_row, column=1, value=row['섹션']).font = normal_font
            ws.cell(row=current_row, column=1).border = thin_border
            
            ws.cell(row=current_row, column=2, value=row['구분']).font = normal_font
            ws.cell(row=current_row, column=2).border = thin_border
            
            ws.cell(row=current_row, column=3, value=row['내용']).font = normal_font
            ws.cell(row=current_row, column=3).border = thin_border
        
        current_row += 1
    
    # 열 너비 설정 (AgGrid와 비슷하게)
    ws.column_dimensions['A'].width = 15  # 섹션
    ws.column_dimensions['B'].width = 35  # 구분
    ws.column_dimensions['C'].width = 25  # 내용
    
    # 전체 테이블에 테두리 적용
    for row in ws.iter_rows(min_row=header_row, max_row=current_row-1, min_col=1, max_col=3):
        for cell in row:
            if not cell.border.left.style:
                cell.border = thin_border
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def save_unified_form_to_supabase():
    """통합 양식을 Supabase에 저장합니다"""
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 필요합니다.")
        return False
    
    try:
        df = st.session_state.get('unified_form_data', pd.DataFrame())
        if df.empty:
            st.warning("⚠️ 저장할 데이터가 없습니다.")
            return False
        
        # 날짜 가져오기
        date_row = df[df['구분'] == '보고일']
        work_date = date_row['내용'].iloc[0] if not date_row.empty else datetime.now().strftime('%Y-%m-%d')
        
        # 데이터 변환
        save_data = {
            "unified_form": df.to_dict('records'),
            "created_at": datetime.now().isoformat()
        }
        
        result = supabase_client.table("unified_reports").insert({
            "date": work_date,
            "data": save_data,
            "created_at": datetime.now().isoformat()
        }).execute()
        
        return True
        
    except Exception as e:
        st.error(f"❌ 저장 실패: {e}")
        return False

def display_template_upload_modal():
    """템플릿 업로드 모달을 표시합니다"""
    
    # 모달 스타일
    st.markdown("""
    <div style="position: fixed; top: 0; left: 0; width: 100%; height: 100%; 
                background-color: rgba(0,0,0,0.5); z-index: 9999; display: flex;
                justify-content: center; align-items: center;">
        <div style="background: white; padding: 30px; border-radius: 15px; 
                    box-shadow: 0 10px 30px rgba(0,0,0,0.3); max-width: 600px; width: 90%;">
    """, unsafe_allow_html=True)
    
    st.markdown("### 📤 엑셀 템플릿 업로드 및 관리")
    st.markdown("---")
    
    # 현재 저장된 템플릿 목록 표시
    saved_templates = get_saved_templates()
    
    if saved_templates:
        st.markdown("#### 📋 저장된 템플릿 목록")
        for i, template in enumerate(saved_templates):
            col1, col2, col3 = st.columns([3, 1, 1])
            with col1:
                st.write(f"**{template['name']}**")
                st.caption(f"업로드: {template['created_at'][:10]}")
            with col2:
                if st.button("🔄 적용", key=f"apply_template_{i}"):
                    load_template_to_form(template)
                    st.session_state.show_template_upload = False
                    st.success(f"✅ '{template['name']}' 템플릿이 적용되었습니다!")
                    st.rerun()
            with col3:
                if st.button("🗑️ 삭제", key=f"delete_template_{i}"):
                    delete_template(template['id'])
                    st.success(f"✅ '{template['name']}' 템플릿이 삭제되었습니다!")
                    st.rerun()
        st.markdown("---")
    
    # 새 템플릿 업로드
    st.markdown("#### 📤 새 템플릿 업로드")
    
    template_name = st.text_input("템플릿 이름", placeholder="예: 터널공사 기본양식", key="template_name_input")
    uploaded_file = st.file_uploader(
        "엑셀 파일 선택", 
        type=["xlsx", "xls"],
        key="template_file_upload",
        help="업로드한 엑셀 양식이 저장되어 계속 사용할 수 있습니다."
    )
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("💾 저장", key="save_template_btn", use_container_width=True):
            if template_name and uploaded_file:
                if save_excel_template(template_name, uploaded_file):
                    st.success("✅ 템플릿이 저장되었습니다!")
                    st.session_state.show_template_upload = False
                    st.rerun()
            else:
                st.error("❌ 템플릿 이름과 파일을 모두 입력해주세요.")
    
    with col2:
        if st.button("📋 양식에 바로 적용", key="apply_direct_btn", use_container_width=True):
            if uploaded_file:
                load_excel_to_unified_form(uploaded_file)
                st.success("✅ 엑셀 파일이 양식에 적용되었습니다!")
                st.session_state.show_template_upload = False
                st.rerun()
            else:
                st.error("❌ 파일을 선택해주세요.")
    
    with col3:
        if st.button("❌ 닫기", key="close_template_modal", use_container_width=True):
            st.session_state.show_template_upload = False
            st.rerun()
    
    st.markdown("</div></div>", unsafe_allow_html=True)

def get_saved_templates():
    """저장된 템플릿 목록을 가져옵니다"""
    if not supabase_client:
        return []
    
    try:
        result = supabase_client.table("excel_templates").select("*").order("created_at", desc=True).execute()
        return result.data
    except Exception as e:
        st.error(f"❌ 템플릿 목록 조회 실패: {e}")
        return []

def save_excel_template(template_name, uploaded_file):
    """엑셀 템플릿을 저장합니다"""
    if not supabase_client:
        st.warning("⚠️ Supabase 연결이 필요합니다.")
        return False
    
    try:
        # 파일 내용을 base64로 인코딩
        file_content = uploaded_file.read()
        import base64
        encoded_content = base64.b64encode(file_content).decode('utf-8')
        
        # 메타데이터 추출
        file_size = len(file_content)
        file_type = uploaded_file.type
        
        # Supabase에 저장
        result = supabase_client.table("excel_templates").insert({
            "name": template_name,
            "file_name": uploaded_file.name,
            "file_content": encoded_content,
            "file_size": file_size,
            "file_type": file_type,
            "created_at": datetime.now().isoformat()
        }).execute()
        
        return True
        
    except Exception as e:
        st.error(f"❌ 템플릿 저장 실패: {e}")
        return False

def load_template_to_form(template):
    """저장된 템플릿을 양식에 로드합니다"""
    try:
        import base64
        
        # base64 디코딩
        file_content = base64.b64decode(template['file_content'])
        
        # BytesIO 객체로 변환
        file_obj = io.BytesIO(file_content)
        file_obj.name = template['file_name']
        
        # 양식에 적용
        load_excel_to_unified_form(file_obj)
        
    except Exception as e:
        st.error(f"❌ 템플릿 로드 실패: {e}")

def delete_template(template_id):
    """템플릿을 삭제합니다"""
    if not supabase_client:
        return False
    
    try:
        supabase_client.table("excel_templates").delete().eq("id", template_id).execute()
        return True
    except Exception as e:
        st.error(f"❌ 템플릿 삭제 실패: {e}")
        return False

def load_excel_to_unified_form(uploaded_file):
    """업로드된 엑셀 파일을 통합 양식으로 변환합니다 - 엑셀과 100% 동일한 구조"""
    try:
        # 엑셀 파일 읽기
        df_dict = pd.read_excel(uploaded_file, sheet_name=None, header=None)
        
        # 첫 번째 시트 사용
        sheet_name = list(df_dict.keys())[0]
        df = df_dict[sheet_name]
        
        # 통합 양식 데이터로 변환
        unified_data = []
        
        st.info(f"📊 엑셀 파일 분석 중... 총 {len(df)} 행 발견")
        
        # 제목 찾기 및 처리
        title_found = False
        for i, row in df.iterrows():
            row_values = [str(cell) for cell in row if pd.notna(cell)]
            if any('공사' in val and ('보고서' in val or '일보' in val) for val in row_values):
                title_found = True
                # 제목은 기본정보에 포함하지 않고 건너뜀
                continue
        
        # 모든 행을 순차적으로 처리하여 엑셀과 동일한 구조 유지
        current_section = ""
        skip_title = True if title_found else False
        
        for i, row in df.iterrows():
            # 제목 행 건너뛰기
            if skip_title and i < 3:
                continue
            
            # 행의 모든 값 추출
            col_values = []
            for j in range(min(10, len(row))):  # 최대 10열까지만 확인
                if j < len(row):
                    val = str(row.iloc[j]).strip() if pd.notna(row.iloc[j]) else ""
                    col_values.append(val)
                else:
                    col_values.append("")
            
            # 완전히 빈 행인지 확인
            if all(val == "" or val == "nan" for val in col_values):
                unified_data.append({"섹션": "", "구분": "", "내용": "", "타입": "separator"})
                continue
            
            # 첫 번째 열에 값이 있는 경우
            col1_val = col_values[0] if col_values[0] != "nan" else ""
            col2_val = col_values[1] if len(col_values) > 1 and col_values[1] != "nan" else ""
            col3_val = col_values[2] if len(col_values) > 2 and col_values[2] != "nan" else ""
            
            # 섹션 헤더 판단 (더 정확한 기준)
            is_section_header = False
            if col1_val:
                # 이모지나 특정 키워드로 섹션 헤더 판단
                header_keywords = ['🌤️', '🏗️', '📝', '👥', '🚛', '날씨', '기상', '시공', '공사', 
                                 '작업', '공종', '인원', '투입', '장비', '기계', '현황', '정보']
                
                # 단일 셀에 섹션명이 있거나, 배경색이 있는 경우 (헤더로 추정)
                if any(keyword in col1_val for keyword in header_keywords):
                    is_section_header = True
                # 또는 첫 번째 열만 값이 있고 나머지가 비어있는 경우
                elif col1_val and not col2_val and not col3_val:
                    is_section_header = True
            
            if is_section_header:
                # 섹션 헤더 추가
                unified_data.append({
                    "섹션": col1_val.split()[0] if col1_val else "",
                    "구분": "섹션 헤더",
                    "내용": col1_val,
                    "타입": "header"
                })
                current_section = col1_val.split()[0] if col1_val else ""
            
            elif col1_val:  # 일반 데이터 행
                # 섹션 타입 결정
                section_type = "custom"
                if current_section:
                    if any(keyword in current_section.lower() for keyword in ['날씨', '기상']):
                        section_type = "weather"
                    elif any(keyword in current_section.lower() for keyword in ['시공', '공사']):
                        section_type = "construction"
                    elif any(keyword in current_section.lower() for keyword in ['작업', '공종']):
                        section_type = "work"
                    elif any(keyword in current_section.lower() for keyword in ['인원', '투입']):
                        section_type = "personnel"
                    elif any(keyword in current_section.lower() for keyword in ['장비', '기계']):
                        section_type = "equipment"
                    elif any(keyword in current_section.lower() for keyword in ['기본', '정보']):
                        section_type = "basic"
                
                # 내용 결정 (우선순위: col3 > col2 > col1의 일부)
                content = ""
                if col3_val:
                    content = col3_val
                elif col2_val:
                    content = col2_val
                
                unified_data.append({
                    "섹션": current_section,
                    "구분": col1_val,
                    "내용": content,
                    "타입": section_type
                })
        
        # 기본 정보 추가 (없는 경우에만)
        has_basic_info = any(item['타입'] == 'basic' for item in unified_data)
        if not has_basic_info:
            basic_info = [
                {"섹션": "기본정보", "구분": "프로젝트명", "내용": "터널 건설공사", "타입": "basic"},
                {"섹션": "기본정보", "구분": "구간명", "내용": "대림-신풍-도림", "타입": "basic"},
                {"섹션": "기본정보", "구분": "보고일", "내용": datetime.now().strftime('%Y-%m-%d'), "타입": "basic"},
                {"섹션": "", "구분": "", "내용": "", "타입": "separator"}
            ]
            unified_data = basic_info + unified_data
        
        # 세션 상태에 저장
        st.session_state.unified_form_data = pd.DataFrame(unified_data)
        
        # 성공 메시지와 함께 변환 결과 요약
        total_rows = len(unified_data)
        header_count = len([item for item in unified_data if item['타입'] == 'header'])
        data_count = len([item for item in unified_data if item['타입'] not in ['header', 'separator']])
        
        st.success(f"✅ 엑셀 템플릿이 성공적으로 로드되었습니다!")
        st.info(f"📊 변환 결과: 총 {total_rows}행 (섹션 {header_count}개, 데이터 {data_count}개)")
        
    except Exception as e:
        st.error(f"❌ 엑셀 파일 로드 실패: {e}")
        import traceback
        st.code(traceback.format_exc())

# 메인 앱 실행
# Streamlit 앱은 파일이 직접 실행되므로 별도의 main() 함수 호출이 필요하지 않습니다.

# --- 메인 애플리케이션 로직 ---
def main():
    # Supabase 초기화
    if SUPABASE_AVAILABLE:
        try:
            supabase = init_supabase()
            if supabase:
                st.success("✅ Supabase 연결 성공!")
        except Exception as e:
            st.warning(f"⚠️ Supabase 연결 실패: {e}")
    
    # 제목
    st.markdown("""
    <div style="text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                color: white; padding: 20px; border-radius: 10px; margin: 20px 0;">
        <h1 style="margin: 0; font-size: 28px; font-weight: bold;">🏗️ 터널 건설공사 일일작업보고서</h1>
    </div>
    """, unsafe_allow_html=True)
    
    # 사이드바
    with st.sidebar:
        st.markdown("### 📋 작업 단계")
        st.markdown("1. **기본 정보 입력**")
        st.markdown("2. **날씨 정보 입력**")
        st.markdown("3. **시공 현황 입력**")
        st.markdown("4. **작업 내용 입력**")
        st.markdown("5. **인원 현황 입력**")
        st.markdown("6. **장비 현황 입력**")
        st.markdown("7. **🤖 AI 분석 및 통합 보고서**")
    
    # 탭 생성
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "🌤️ 날씨정보", "🏗️ 시공현황", "📝 작업내용", "👥 인원현황", "🚛 장비현황", "🤖 AI 분석 및 통합 보고서"
    ])
    
    # 탭 1: 시공현황
    with tab1:
        st.markdown("### 🏗️ 시공 현황 입력")
        if 'construction_table' not in st.session_state:
            initialize_construction_table()
        
        construction_edited = st.data_editor(
            st.session_state.construction_table,
            key="construction_editor",
            num_rows="dynamic",
            use_container_width=True
        )
        st.session_state.construction_table = construction_edited
    
    # 탭 2: 작업내용
    with tab2:
        st.markdown("### 📝 작업 내용 입력")
        if 'work_content_table' not in st.session_state:
            initialize_work_content_table()
        
        work_content_edited = st.data_editor(
            st.session_state.work_content_table,
            key="work_content_editor",
            num_rows="dynamic",
            use_container_width=True
        )
        st.session_state.work_content_table = work_content_edited
    
    # 탭 3: 인원현황
    with tab3:
        st.markdown("### 👥 인원 현황 입력")
        if 'personnel_table' not in st.session_state:
            initialize_personnel_table()
        
        personnel_edited = st.data_editor(
            st.session_state.personnel_table,
            key="personnel_editor",
            num_rows="dynamic",
            use_container_width=True
        )
        st.session_state.personnel_table = personnel_edited
    
    # 탭 4: 장비현황
    with tab4:
        st.markdown("### 🚛 장비 현황 입력")
        if 'equipment_table' not in st.session_state:
            initialize_equipment_table()
        
        equipment_edited = st.data_editor(
            st.session_state.equipment_table,
            key="equipment_editor",
            num_rows="dynamic",
            use_container_width=True
        )
        st.session_state.equipment_table = equipment_edited
    
    # 탭 5: AI 분석 및 통합 보고서
    with tab5:
        st.markdown("### 🤖 AI 분석 및 통합 보고서")
        
        # AI 분석 섹션
        st.markdown("#### 📝 AI 분석을 위한 텍스트 입력")
        
        col1, col2 = st.columns(2)
        
        with col1:
            project_info = st.text_area(
                "프로젝트 정보",
                placeholder="프로젝트명, 구간명, 보고일 등 기본 정보를 입력하세요...",
                height=100,
                key="ai_project_info"
            )
            
            daily_work = st.text_area(
                "금일 작업 내용",
                placeholder="오늘 수행한 작업 내용을 상세히 입력하세요...",
                height=150,
                key="ai_daily_work"
            )
        
        with col2:
            issues_solutions = st.text_area(
                "문제점 및 해결방안",
                placeholder="발생한 문제점과 해결방안을 입력하세요...",
                height=100,
                key="ai_issues_solutions"
            )
            
            # 프롬프트 편집 섹션
            with st.expander("🔧 프롬프트(지시문) 수정", expanded=False):
                st.markdown("#### 프롬프트 관리")
                
                # 프롬프트 선택
                prompt_options = ["기본 프롬프트"] + [p['name'] for p in get_all_prompts_from_supabase()]
                selected_prompt = st.selectbox("프롬프트 선택:", prompt_options, key="prompt_selector")
                
                if selected_prompt == "기본 프롬프트":
                    current_prompt = DEFAULT_PROMPT
                else:
                    prompt_data = load_prompt_from_supabase(selected_prompt)
                    current_prompt = prompt_data['content'] if prompt_data else DEFAULT_PROMPT
                
                # 프롬프트 편집
                edited_prompt = st.text_area(
                    "프롬프트(지시문) 수정",
                    value=current_prompt,
                    height=300,
                    key="prompt_editor"
                )
                
                # 프롬프트 관리 버튼들
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("💾 저장", key="save_prompt_btn", use_container_width=True):
                        prompt_name = st.text_input("프롬프트 이름:", value=f"프롬프트_{datetime.now().strftime('%Y%m%d_%H%M%S')}", key="prompt_name_input")
                        if prompt_name:
                            save_prompt_to_supabase(prompt_name, edited_prompt)
                
                with col2:
                    if st.button("🔄 새로고침", key="refresh_prompts_btn", use_container_width=True):
                        st.rerun()
                
                with col3:
                    if st.button("🗑️ 삭제", key="delete_prompt_btn", use_container_width=True):
                        if selected_prompt != "기본 프롬프트":
                            delete_prompt_from_supabase(selected_prompt)
                            st.rerun()
        
        # AI 분석 실행 버튼
        st.markdown("---")
        if st.button("🤖 AI로 분석하기", key="run_ai_analysis", use_container_width=True):
            if project_info or daily_work or issues_solutions:
                # AI 분석 실행
                combined_text = f"프로젝트 정보: {project_info}\n\n금일 작업 내용: {daily_work}\n\n문제점 및 해결방안: {issues_solutions}"
                
                with st.spinner("🤖 AI가 텍스트를 분석하고 있습니다..."):
                    try:
                        # AI 분석 실행
                        result = analyze_text_with_ai(combined_text, edited_prompt)
                        if result:
                            st.session_state.processed_tables = result
                            st.success("✅ AI 분석이 완료되었습니다!")
                            st.balloons()
                        else:
                            st.error("❌ AI 분석 중 오류가 발생했습니다.")
                    except Exception as e:
                        st.error(f"❌ AI 분석 실패: {e}")
            else:
                st.warning("⚠️ 분석할 텍스트를 입력해주세요.")
        
        # 통합 보고서 섹션
        st.markdown("---")
        st.markdown("#### 📊 통합 보고서")
        
        # 통합 공사일보 표시
        display_unified_construction_report()

# 메인 함수 실행
if __name__ == "__main__":
    main()